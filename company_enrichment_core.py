from __future__ import annotations

import csv
import errno
import hashlib
import json
import logging
import os
import random
import re
import sys
import time
from collections import Counter, defaultdict
from collections.abc import Mapping
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from threading import Lock
from typing import Any, Iterable
from urllib.parse import quote_plus, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from urllib3 import disable_warnings
from urllib3.exceptions import InsecureRequestWarning

from app.discovery.address_resolution import AddressEnrichment, enrich_address_candidates
from app.discovery.company_profile import (
    COMPANY_OUTPUT_CONTRACT_VERSION,
    CompanyProfile,
    assemble_company_profile,
    company_profile_from_dict,
    company_profile_to_dict,
)
from app.discovery.geo_lookup import GeoLookupResult, lookup_settlement
from app.discovery.name_signal import (
    COMPANY_TOKEN_STOPWORDS,
    NameSignalResult,
    company_tokens,
    detect_name_signal,
    normalize_company_name,
)
from app.discovery.models import DomainCandidate, DomainResolution
from app.llm.benchmark_capture import (
    LLMBenchmarkCaptureWriter,
    describe_content_review_prod_skip_reason,
    select_content_review_benchmark_records,
)
from app.llm.content_review_compaction import (
    DEFAULT_CONTENT_REVIEW_EXCERPT_CHARS,
    build_content_review_excerpt,
)
from app.llm.pricing import calculate_usage_cost_usd
from app.llm.openai_responses import (
    OpenAIJsonParseError,
    extract_openai_json as _extract_openai_json,
    extract_openai_text as _extract_openai_text,
    parse_openai_response,
)
from app.runtime import ProgressStore, ProxyPool, ProxySelection
from app.runtime.concurrency import (
    DEFAULT_SOURCE_TRANSPORT_POLICY,
    DIRECT_DEFAULT_TRANSPORT,
    OFFLINE_ONLY_TRANSPORT,
    PROXY_BOUND_TRANSPORT,
    SESSION_BOUND_TRANSPORT,
)
from app.runtime.proxy6 import (
    PROXY_PROVIDER_INVENTORY_EMPTY_OR_EXPIRED,
    PROXY_PROVIDER_INVENTORY_HEALTHY,
    PROXY_PROVIDER_STATUS_UNKNOWN,
    Proxy6InventoryDiagnostic,
    diagnose_proxy6_inventory_from_env,
)
from app.site_intelligence import (
    SITE_AUTH_STATUS_RANK,
    ContentRecord,
    RouteStrategy,
    SiteAuthHelpers,
    SiteAuthenticityAnalyzer,
    SiteDecision,
    SiteProbe,
    content_record_from_dict,
    infer_lead_type_from_record,
    route_strategy_from_dict,
    site_probe_from_dict,
)


UTC = timezone.utc
PROGRESS_DIAGNOSTIC_LOGGER = logging.getLogger("company_research_parser")
HOST_STATS_NON_FATAL_ERRNOS = frozenset({errno.EACCES, errno.EPERM, errno.EBUSY})
HOST_STATS_NON_FATAL_WINERRORS = frozenset({5, 32, 33})
SOURCE_DOMAINS = {
    "checko": "checko.ru",
    "zachestnyibiznes": "zachestnyibiznes.ru",
    "rusprofile": "rusprofile.ru",
    "list_org": "list-org.com",
    "spark": "spark-interfax.ru",
}
COMPANY_CONTACT_FIELDS = ("phones", "emails", "websites", "addresses")
DEFAULT_ENV_FILE = ".env"
DEFAULT_LISTORG_SESSION_RELATIVE_PATH = Path("browser_sessions") / "list_org_session.json"
DEFAULT_LISTORG_STORAGE_STATE_RELATIVE_PATH = Path("browser_sessions") / "list_org_storage_state.json"
DEFAULT_LISTORG_JSON_FILE = Path("search.json")
DEFAULT_SITE_DECISION_MODEL = "gpt-5.4-nano"
DEFAULT_CONTENT_REVIEW_MODEL = "gpt-5.4-nano"
DEFAULT_CONTENT_REVIEW_FALLBACK_MODEL = "gpt-5.4-mini"
DEFAULT_SITE_DECISION_MAX_OUTPUT_TOKENS = 384
SITE_DECISION_REASON_MAX_CHARS = 160
SITE_DECISION_LIST_ITEM_MAX_CHARS = 96
SITE_DECISION_EVIDENCE_MAX_ITEMS = 3
SITE_DECISION_CONTRADICTION_MAX_ITEMS = 2
SITE_DECISION_EXCERPT_CHARS = 680
SITE_DECISION_RETRY_TRIGGER = "incomplete_max_output_tokens"
SITE_DECISION_RETRY_REASON_TARGET_CHARS = 96
SITE_DECISION_RETRY_LIST_ITEM_TARGET_CHARS = 72
SITE_DECISION_RETRY_EVIDENCE_TARGET_ITEMS = 2
SITE_DECISION_RETRY_CONTRADICTION_TARGET_ITEMS = 1
SITE_DECISION_RETRY_EXCERPT_CHARS = 360
CONTENT_REVIEW_FALLBACK_PARSER_REASONS = {
    "incomplete_max_output_tokens",
    "non_json_text",
    "refusal",
    "malformed_json",
    "missing_output",
}
LLM_SITE_DECISION_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "belongs_to_company": {"type": "boolean"},
        "website_type": {
            "type": "string",
            "enum": [
                "official_corporate",
                "subsidiary_or_related_brand",
                "marketplace_or_catalog",
                "directory_or_listing",
                "social_or_messenger",
                "unrelated",
                "uncertain",
            ],
        },
        "industrial_relevance": {"type": "string", "enum": ["high", "medium", "low", "none"]},
        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
        "reason": {"type": "string", "minLength": 1, "maxLength": SITE_DECISION_REASON_MAX_CHARS},
        "evidence": {
            "type": "array",
            "items": {"type": "string", "maxLength": SITE_DECISION_LIST_ITEM_MAX_CHARS},
            "maxItems": SITE_DECISION_EVIDENCE_MAX_ITEMS,
        },
        "contradictions": {
            "type": "array",
            "items": {"type": "string", "maxLength": SITE_DECISION_LIST_ITEM_MAX_CHARS},
            "maxItems": SITE_DECISION_CONTRADICTION_MAX_ITEMS,
        },
    },
    "required": [
        "belongs_to_company",
        "website_type",
        "industrial_relevance",
        "confidence",
        "reason",
        "evidence",
        "contradictions",
    ],
}
LLM_CONTENT_RELEVANCE_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "relevance_label": {"type": "string", "enum": ["irrelevant", "maybe_relevant", "likely_relevant"]},
        "lead_type": {"type": "string", "enum": ["tender", "realization", "direct_sale", "document", "news", "unknown"]},
        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
        "summary": {"type": "string", "minLength": 1},
        "evidence": {"type": "array", "items": {"type": "string"}, "maxItems": 5},
    },
    "required": ["relevance_label", "lead_type", "confidence", "summary", "evidence"],
}
GENERIC_EMAIL_DOMAINS = {
    "gmail.com",
    "googlemail.com",
    "mail.ru",
    "bk.ru",
    "inbox.ru",
    "list.ru",
    "yandex.ru",
    "ya.ru",
    "yandex.com",
    "rambler.ru",
    "lenta.ru",
    "icloud.com",
    "outlook.com",
    "hotmail.com",
    "live.com",
    "yahoo.com",
    "proton.me",
    "protonmail.com",
    "ru.net",
}
ADDRESS_STOP_PHRASES = (
    "на карте",
    "телефон",
    "тел.",
    "email",
    "e-mail",
    "электронная почта",
    "форма обратной связи",
    "политика конфиденциальности",
    "время работы",
    "copyright",
    "разработка сайта",
    "заказать звонок",
    "горячая линия",
    "инн",
    "кпп",
    "огрн",
    "окпо",
    "оквэд",
    "корр. счет",
    "корр счет",
    "р/с",
    "расчетный счет",
    "наименование банка",
    "бик",
    "написать директору",
    "смотреть на google карте",
    "смотреть на карте",
    "установите приложение",
    "узнать подробнее",
    "новости",
    "вконтакте",
    "telegram",
    "whatsapp",
)
ADDRESS_STOP_PHRASE_PATTERNS = tuple(
    re.compile(
        r"(?<![a-zа-яё0-9])" + re.escape(phrase).replace(r"\ ", r"\s+") + r"(?![a-zа-яё0-9])",
        flags=re.IGNORECASE,
    )
    for phrase in ADDRESS_STOP_PHRASES
)
ADDRESS_TOKEN_STOPWORDS = {
    "ао",
    "г",
    "город",
    "дом",
    "домовладение",
    "д",
    "каб",
    "ком",
    "корп",
    "корпус",
    "край",
    "лит",
    "литер",
    "москва",
    "муниципальный",
    "обл",
    "область",
    "office",
    "оф",
    "офис",
    "пер",
    "переулок",
    "пом",
    "помещение",
    "поселок",
    "почтамт",
    "пр",
    "пр-д",
    "проезд",
    "просп",
    "проспект",
    "пр-кт",
    "район",
    "республика",
    "рф",
    "россия",
    "санкт",
    "спб",
    "стр",
    "строение",
    "территория",
    "ул",
    "улица",
    "ш",
    "шоссе",
    "этаж",
}
CONTACT_PATH_HINTS = [
    "contact",
    "contacts",
    "kontakt",
    "kontakty",
    "about",
    "company",
    "o-kompanii",
    "about-us",
    "production",
    "manufacturing",
    "factory",
    "plant",
]
CONTACT_LINK_TEXT_HINTS = [
    "контакт",
    "contact",
    "о компании",
    "about",
    "компания",
    "предприят",
    "производ",
]
SITE_PROBE_ROUTE_HINTS = [
    ("/contacts", "contacts"),
    ("/kontakt", "contacts"),
    ("/kontakty", "contacts"),
    ("/about", "about"),
    ("/o-kompanii", "about"),
    ("/company", "about"),
    ("/procurement", "procurement"),
    ("/zakupki", "procurement"),
    ("/tenders", "tenders"),
    ("/torgi", "tenders"),
    ("/news", "news"),
    ("/documents", "documents"),
    ("/docs", "documents"),
]
DOCUMENT_EXTENSIONS = (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".zip", ".rar")
SPA_MARKERS = (
    "data-reactroot",
    "__next_data__",
    "id=\"__next\"",
    "id=\"root\"",
    "id=\"app\"",
    "ng-version",
    "webpack",
    "vite",
    "nuxt",
    "vue",
)
CMS_SIGNATURES: list[tuple[str, str]] = [
    ("1c-bitrix", "bitrix"),
    ("bitrix", "bitrix"),
    ("wp-content", "wordpress"),
    ("wordpress", "wordpress"),
    ("drupal-settings-json", "drupal"),
    ("joomla", "joomla"),
    ("/typo3/", "typo3"),
    (".aspx", "aspnet"),
    ("asp.net", "aspnet"),
]
INDUSTRIAL_POSITIVE_KEYWORDS = {
    "завод": 3,
    "предприят": 2,
    "производ": 3,
    "изготов": 2,
    "переработ": 3,
    "цех": 2,
    "комбинат": 3,
    "фабрик": 3,
    "литейн": 3,
    "металл": 2,
    "металлоконструк": 3,
    "сталь": 2,
    "труба": 2,
    "арматур": 2,
    "листовой": 2,
    "сырье": 1,
    "отход": 1,
    "вторсыр": 2,
    "лом": 2,
    "демонтаж": 1,
    "переплав": 2,
    "станк": 2,
    "оборудован": 1,
    "машиностро": 3,
    "промышлен": 2,
    "конвейер": 1,
    "логистич": 1,
    "склад": 1,
    "road": 1,
    "construction": 1,
    "manufactur": 3,
    "factory": 3,
    "plant": 3,
    "processing": 3,
}
INDUSTRIAL_NEGATIVE_KEYWORDS = {
    "казино": 3,
    "ставк": 3,
    "bet": 3,
    "seo": 2,
    "маркетингов": 2,
    "digital agency": 3,
    "таро": 3,
    "ресторан": 2,
    "кафе": 2,
    "отель": 2,
    "hotel": 2,
    "blog": 1,
    "news portal": 2,
    "онлайн-школ": 2,
    "курсы": 2,
    "casino": 3,
    "gambling": 3,
}
LEAD_POSITIVE_KEYWORDS = {
    "тендер": 3,
    "торги": 3,
    "закупк": 3,
    "реализац": 3,
    "продаж": 2,
    "аукцион": 3,
    "конкурс": 2,
    "запрос предложений": 2,
    "лом": 2,
    "вторсыр": 2,
    "отход": 2,
    "неликвид": 3,
    "списан": 2,
    "демонтаж": 2,
    "металлоконструк": 3,
    "имущество": 1,
    "оборудован": 1,
    "лот": 2,
    "извещение": 2,
    "документация": 1,
    "протокол": 1,
}
LEAD_NEGATIVE_KEYWORDS = {
    "ваканс": 3,
    "политика конфиденциальности": 3,
    "пользовательское соглашение": 3,
    "cookie": 2,
    "о компании": 1,
    "контакты": 1,
    "новости компании": 1,
}
BOT_GATE_PATTERNS = [
    "too many requests",
    "captcha",
    "turnstile",
    "cloudflare",
    "verify you are human",
    "checking if the site connection is secure",
    "проверка, что вы не робот",
    "вы слишком часто обращались к сайту",
    "распознана как автоматическая",
    "автоматическая активность",
    "/bot?from=",
]
SHARED_CONTACT_FIELDS = ("phones", "emails", "websites", "addresses")
NON_CONTACT_AVAILABILITY_FIELDS = ("management", "founders")
IMPORTANT_FIELDS = SHARED_CONTACT_FIELDS + NON_CONTACT_AVAILABILITY_FIELDS
SOURCE_AVAILABILITY_STATUSES = ("open", "masked", "absent", "blocked", "unknown")
REQUEST_STATUS_BLOCKED_NO_PROXY = "blocked_no_proxy"
REQUEST_BLOCKED_NO_PROXY_NO_POOL_REASON = "usable proxy pool count == 0 before outbound request"
REQUEST_BLOCKED_NO_PROXY_SELECTION_UNAVAILABLE_REASON = "proxy selection unavailable before outbound request"
CANONICAL_REQUIRED_SOURCE_NAMES = frozenset({"list_org", "rusprofile", "spark", "zachestnyibiznes", "checko"})
SOURCE_BLOCKED_RESULT_STATUSES = frozenset(
    {
        "auth_failed",
        "blocked",
        REQUEST_STATUS_BLOCKED_NO_PROXY,
        "bot_gate",
        "cooldown_active",
        "rate_limited",
        "source_disabled_after_block",
    }
)
SOURCE_ABSENT_CONTACT_RESULT_STATUSES = frozenset({"mismatch", "not_found"})
SOURCE_OPERATIONAL_RESULT_STATUSES = frozenset({"success", "ok", "not_found", "mismatch"})
DIRECT_OPERATIONAL_RESULT_STATUSES = frozenset({"guest"})
NON_CORPORATE_DOMAINS = {
    "115",
    "facebook.com",
    "instagram.com",
    "linkedin.com",
    "rutube.ru",
    "tiktok.com",
    "twitter.com",
    "wa.me",
    "whatsapp.com",
    "x.com",
    "yandex.ru",
    "yandex.com",
    "vk.com",
    "t.me",
    "telegram.me",
    "youtu.be",
    "youtube.com",
    "dzen.ru",
    "ok.ru",
    "ru.net",
    "digital.gov.ru",
    "reestr.digital.gov.ru",
}
ACTIVITY_TOKEN_STOPWORDS = {
    "компания",
    "официальный",
    "производство",
    "услуги",
    "решения",
    "продукция",
    "продукты",
    "главная",
    "контакты",
    "новости",
    "о компании",
    "group",
    "company",
    "official",
    "services",
    "solutions",
    "products",
    "homepage",
    "contacts",
    "news",
}
RUN_DISABLE_ON_BLOCK_STATUSES = frozenset({"blocked", "rate_limited", "bot_gate"})
LLM_BENCHMARK_CAPTURE_SUMMARY_FIELDS = {
    "site_decision": ("captured_site_decision_count", "captured_site_decision_company_count"),
    "content_review": ("captured_content_review_count", "captured_content_review_company_count"),
}


def utc_now_iso() -> str:
    return datetime.now(tz=UTC).replace(microsecond=0).isoformat()


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    try:
        raw_text = path.read_text(encoding="utf-8")
    except OSError:
        return
    for raw_line in raw_text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key or key in os.environ:
            continue
        value = value.strip()
        if value and value[0] == value[-1] and value[0] in {'"', "'"}:
            value = value[1:-1]
        os.environ[key] = value


def cookie_header_from_items(cookies: list[dict[str, Any]]) -> str:
    pairs: list[str] = []
    for cookie in cookies:
        name = normalize_whitespace(str(cookie.get("name", "")))
        value = str(cookie.get("value", ""))
        if not name:
            continue
        pairs.append(f"{name}={value}")
    return "; ".join(pairs)


def load_session_profile(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    return payload


def save_session_profile(path: Path, payload: dict[str, Any]) -> None:
    ensure_dir(path.parent)
    atomic_write_json(path, payload)


def slugify(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9а-яА-ЯёЁ]+", "-", value.lower(), flags=re.IGNORECASE)
    return value.strip("-")


def normalize_whitespace(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


MOJIBAKE_SENTINELS = ("Ð", "Ñ", "Ã", "â", "ð", "\ufffd")


def _mojibake_score(value: str) -> int:
    return sum(value.count(marker) for marker in MOJIBAKE_SENTINELS) + sum(
        1 for char in value if 0x80 <= ord(char) <= 0x9F
    )


def _cyrillic_score(value: str) -> int:
    return sum(1 for char in value if "А" <= char <= "я" or char in {"Ё", "ё"})


def repair_mojibake_text(value: str) -> str:
    repaired = value
    for _ in range(2):
        current_score = _mojibake_score(repaired)
        if current_score == 0:
            break
        candidates: list[str] = []
        for source_encoding in ("latin-1", "cp1252"):
            try:
                candidate = repaired.encode(source_encoding).decode("utf-8")
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
            if candidate != repaired:
                candidates.append(candidate)
        if not candidates:
            break
        best_candidate = min(
            candidates,
            key=lambda item: (_mojibake_score(item), -_cyrillic_score(item), len(item)),
        )
        best_score = _mojibake_score(best_candidate)
        if best_score > current_score:
            break
        if best_score == current_score and _cyrillic_score(best_candidate) <= _cyrillic_score(repaired):
            break
        repaired = best_candidate
    return repaired


def repair_output_value(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: repair_output_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [repair_output_value(item) for item in value]
    if isinstance(value, tuple):
        return tuple(repair_output_value(item) for item in value)
    if isinstance(value, str):
        return repair_mojibake_text(value)
    return value


def normalize_inn(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))
    text = normalize_whitespace(str(value))
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits or text


def is_valid_russian_inn(inn: str) -> bool:
    digits = normalize_inn(inn)
    if not digits.isdigit():
        return False
    if len(digits) == 10:
        coeffs = (2, 4, 10, 3, 5, 9, 4, 6, 8)
        checksum = sum(int(d) * c for d, c in zip(digits[:9], coeffs)) % 11 % 10
        return checksum == int(digits[9])
    if len(digits) == 12:
        coeffs11 = (7, 2, 4, 10, 3, 5, 9, 4, 6, 8)
        coeffs12 = (3, 7, 2, 4, 10, 3, 5, 9, 4, 6, 8)
        checksum11 = sum(int(d) * c for d, c in zip(digits[:10], coeffs11)) % 11 % 10
        checksum12 = sum(int(d) * c for d, c in zip(digits[:11], coeffs12)) % 11 % 10
        return checksum11 == int(digits[10]) and checksum12 == int(digits[11])
    return False


def normalized_phone_digits(value: str) -> str:
    normalized = normalize_phone_candidate(value)
    if not normalized:
        return ""
    return re.sub(r"\D+", "", normalized)


def normalize_phone_candidate(value: str) -> str:
    raw = normalize_whitespace(value)
    if not raw:
        return ""
    raw = re.sub(r"^(?:tel:|тел\.?:?)\s*", "", raw, flags=re.IGNORECASE)
    raw = re.split(r"(?i)\b(?:доб\.?|ext\.?|extension|внутр\.?|доп\.?)\b", raw, maxsplit=1)[0]
    digits = re.sub(r"\D+", "", raw)
    if len(digits) == 10:
        if not any(symbol in raw for symbol in "+()- "):
            return ""
        digits = "7" + digits
    elif len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) != 11 or not digits.startswith("7"):
        return ""
    if digits[1] in {"0", "1", "2"}:
        return ""
    if digits[1] == "8" and digits[1:4] != "800":
        return ""
    return format_phone(digits)


def normalize_phone_values(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = normalize_phone_candidate(value)
        if not normalized:
            continue
        digits = normalized_phone_digits(normalized)
        if not digits or digits in seen:
            continue
        seen.add(digits)
        result.append(normalized)
    return result


def extract_phones(text: str) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for match in re.finditer(r"(?<!\d)(?:(?:\+7|8)[\s\-()]*)?(?:\d[\s\-()]*){10,11}(?!\d)", text):
        raw = match.group(0)
        context_before = text[max(0, match.start() - 24):match.start()].lower()
        context_after = text[match.end(): min(len(text), match.end() + 12)].lower()
        if any(marker in context_before or marker in context_after for marker in ("инн", "кпп", "огрн", "бик", "окпо", "оквэд", "счет", "р/с", "корр", "лиценз")):
            continue
        normalized = normalize_phone_candidate(raw)
        if not normalized:
            continue
        digits = normalized_phone_digits(normalized)
        if digits in seen:
            continue
        seen.add(digits)
        result.append(normalized)
    return result


def format_phone(digits: str) -> str:
    if len(digits) == 11 and digits.startswith("7"):
        return f"+7 {digits[1:4]} {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    return digits


def extract_emails(text: str) -> list[str]:
    raw = re.findall(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", text, flags=re.IGNORECASE)
    emails: list[str] = []
    seen: set[str] = set()
    for email in raw:
        lowered = email.lower().strip(" .,:;")
        if lowered in seen:
            continue
        seen.add(lowered)
        emails.append(lowered)
    return emails


def guess_registered_domain(host: str) -> str:
    host = host.lower().strip(".")
    if not host:
        return ""
    parts = host.split(".")
    if len(parts) <= 2:
        return host
    compound_suffixes = {"com.ru", "net.ru", "org.ru", "gov.ru", "edu.ru", "spb.ru", "msk.ru", "co.uk", "com.tr"}
    suffix = ".".join(parts[-2:])
    if suffix in compound_suffixes and len(parts) >= 3:
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


def extract_urls(text: str) -> list[str]:
    urls = re.findall(r"https?://[^\s<>\"]+", text, flags=re.IGNORECASE)
    result: list[str] = []
    seen: set[str] = set()
    for url in urls:
        clean = url.rstrip(").,;]}>\"'")
        if clean in seen:
            continue
        seen.add(clean)
        result.append(clean)
    return result


def normalize_url(url: str) -> str:
    url = normalize_whitespace(url)
    if not url:
        return ""
    if url.startswith("//"):
        url = "https:" + url
    if not re.match(r"^[a-z]+://", url, flags=re.IGNORECASE):
        url = "https://" + url
    parsed = urlparse(url)
    if not parsed.netloc:
        return ""
    return parsed._replace(fragment="").geturl()


def domain_matches_blocklist(domain: str, blocklist: set[str]) -> bool:
    domain = domain.lower().strip(".")
    if not domain:
        return False
    return any(domain == item or domain.endswith("." + item) for item in blocklist)


def is_probably_corporate_domain(domain: str) -> bool:
    domain = domain.lower().strip(".")
    if not domain or "." not in domain:
        return False
    if "░" in domain or "*" in domain:
        return False
    if re.fullmatch(r"(?:www\.)?\d+(?::\d+)?", domain):
        return False
    if not re.search(r"[a-zа-яё]", domain, flags=re.IGNORECASE):
        return False
    if domain_matches_blocklist(domain, NON_CORPORATE_DOMAINS):
        return False
    if domain_matches_blocklist(domain, set(SOURCE_DOMAINS.values())):
        return False
    return True


def sanitize_website_url(url: str, *, keep_path: bool = True) -> str:
    normalized = normalize_url(url)
    if not normalized:
        return ""
    parsed = urlparse(normalized)
    host = parsed.netloc.lower().split("@")[-1].strip(".")
    domain = guess_registered_domain(host)
    if not is_probably_corporate_domain(domain):
        return ""
    path = parsed.path or ""
    if not keep_path:
        path = ""
    return parsed._replace(netloc=host, path=path, params="", query="", fragment="").geturl()


def email_to_candidate_website(email: str) -> str:
    email = normalize_whitespace(email).lower()
    if "@" not in email:
        return ""
    domain = email.split("@", 1)[-1]
    if domain in GENERIC_EMAIL_DOMAINS:
        return ""
    return sanitize_website_url(domain, keep_path=False)


def split_list_org_www_candidates(value: str | None) -> list[str]:
    raw = normalize_whitespace(value)
    if not raw:
        return []
    raw = raw.replace(" | ", "|").replace("| ", "|").replace(" |", "|")
    parts = re.split(r"[|,]+", raw)
    candidates: list[str] = []
    for part in parts:
        cleaned = sanitize_website_url(part, keep_path=True)
        if cleaned:
            candidates.append(cleaned)
    return dedupe_websites_preserve_order(candidates)


def dedupe_websites_preserve_order(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen_domains: set[str] = set()
    for value in values:
        cleaned = sanitize_website_url(value)
        if not cleaned:
            continue
        domain = guess_registered_domain(urlparse(cleaned).netloc)
        if not domain or domain in seen_domains:
            continue
        seen_domains.add(domain)
        result.append(cleaned)
    return result


def dedupe_preserve_order(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value:
            continue
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def extract_text_snippet(text: str, needle: str, span: int = 200) -> str:
    if not text or not needle:
        return ""
    idx = text.lower().find(needle.lower())
    if idx < 0:
        return ""
    return normalize_whitespace(text[max(0, idx - span): idx + span])


def looks_like_bot_gate(response: requests.Response, text: str) -> bool:
    combined = f"{response.url}\n{text[:6000]}".lower()
    return any(pattern in combined for pattern in BOT_GATE_PATTERNS)


def sanitize_for_json(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: sanitize_for_json(val) for key, val in value.items()}
    if isinstance(value, list):
        return [sanitize_for_json(item) for item in value]
    return value


def atomic_write_json(path: Path, data: Any) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(sanitize_for_json(data), ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def _cleanup_temp_file(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        return
    except OSError:
        return


def _is_non_fatal_host_stats_replace_error(error: OSError) -> bool:
    if isinstance(error, PermissionError):
        return True
    winerror = getattr(error, "winerror", None)
    if isinstance(winerror, int) and winerror in HOST_STATS_NON_FATAL_WINERRORS:
        return True
    return error.errno in HOST_STATS_NON_FATAL_ERRNOS


def atomic_write_host_stats_json_best_effort(path: Path, data: Any) -> bool:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(sanitize_for_json(data), ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        tmp.replace(path)
    except OSError as error:
        if not _is_non_fatal_host_stats_replace_error(error):
            raise
        _cleanup_temp_file(tmp)
        PROGRESS_DIAGNOSTIC_LOGGER.warning(
            "Non-fatal host_stats persistence failure; continuing run: target=%s tmp=%s errno=%s winerror=%s error=%s",
            path,
            tmp,
            getattr(error, "errno", None),
            getattr(error, "winerror", None),
            error,
        )
        return False
    return True


def atomic_write_text(path: Path, text: str) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def append_jsonl(path: Path, item: Any) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(sanitize_for_json(item), ensure_ascii=False) + "\n")


def build_list_org_session_profile(
    *,
    cookies: list[dict[str, Any]],
    user_agent: str,
    referer: str,
    storage_state_path: Path,
) -> dict[str, Any]:
    return {
        "source": "list_org",
        "created_at": utc_now_iso(),
        "user_agent": normalize_whitespace(user_agent),
        "referer": normalize_whitespace(referer) or "https://www.list-org.com/",
        "cookies": cookies,
        "cookie_header": cookie_header_from_items(cookies),
        "storage_state_path": str(storage_state_path),
    }


def run_list_org_bootstrap(
    *,
    logger: logging.Logger,
    session_file: Path,
    storage_state_file: Path,
    headless: bool = False,
) -> None:
    try:
        from playwright.sync_api import sync_playwright
    except Exception as exc:
        raise RuntimeError(
            "Playwright Ð½ÐµÐ´Ð¾ÑÑ‚ÑƒÐ¿ÐµÐ½. Ð£ÑÑ‚Ð°Ð½Ð¾Ð²Ð¸ Ð·Ð°Ð²Ð¸ÑÐ¸Ð¼Ð¾ÑÑ‚Ð¸ Ð¸ Ð±Ñ€Ð°ÑƒÐ·ÐµÑ€: `pip install -r requirements.txt` Ð¸ `python -m playwright install chromium`."
        ) from exc

    ensure_dir(session_file.parent)
    ensure_dir(storage_state_file.parent)
    bootstrap_url = os.getenv("LISTORG_BOOTSTRAP_URL", "https://www.list-org.com/search?val=7707083893").strip()
    referer = os.getenv("LISTORG_REFERER", "https://www.list-org.com/").strip() or "https://www.list-org.com/"
    configured_user_agent = os.getenv("LISTORG_USER_AGENT", "").strip() or (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )

    logger.info("Ð—Ð°Ð¿ÑƒÑÐºÐ°ÑŽ Playwright bootstrap Ð´Ð»Ñ List-Org: %s", bootstrap_url)
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=headless)
        context = browser.new_context(
            user_agent=configured_user_agent,
            locale="ru-RU",
            ignore_https_errors=True,
            viewport={"width": 1440, "height": 960},
        )
        page = context.new_page()
        page.goto(bootstrap_url, wait_until="domcontentloaded", timeout=90_000)
        logger.info("Ð’ Ð±Ñ€Ð°ÑƒÐ·ÐµÑ€Ðµ Ð¾Ñ‚ÐºÑ€Ñ‹Ñ‚ List-Org. ÐŸÑ€Ð¾Ð¹Ð´Ð¸ challenge/login Ð¸ Ð²ÐµÑ€Ð½Ð¸ÑÑŒ Ð² Ñ‚ÐµÑ€Ð¼Ð¸Ð½Ð°Ð».")
        print(
            "\nList-Org bootstrap:\n"
            "1. Ð’ Ð¾Ñ‚ÐºÑ€Ñ‹Ñ‚Ð¾Ð¼ Chromium Ð¿Ñ€Ð¾Ð¹Ð´Ð¸ Cloudflare challenge / Ð»Ð¾Ð³Ð¸Ð½.\n"
            "2. Ð£Ð±ÐµÐ´Ð¸ÑÑŒ, Ñ‡Ñ‚Ð¾ ÑÑ‚Ñ€Ð°Ð½Ð¸Ñ†Ð° List-Org Ð·Ð°Ð³Ñ€ÑƒÐ·Ð¸Ð»Ð°ÑÑŒ Ð½Ð¾Ñ€Ð¼Ð°Ð»ÑŒÐ½Ð¾.\n"
            "3. Ð’ÐµÑ€Ð½Ð¸ÑÑŒ Ð² Ñ‚ÐµÑ€Ð¼Ð¸Ð½Ð°Ð» Ð¸ Ð½Ð°Ð¶Ð¼Ð¸ Enter, Ñ‡Ñ‚Ð¾Ð±Ñ‹ ÑÐ¾Ñ…Ñ€Ð°Ð½Ð¸Ñ‚ÑŒ cookies.\n"
        )
        input()
        current_user_agent = page.evaluate("() => navigator.userAgent")
        storage = context.storage_state(path=str(storage_state_file))
        cookies = storage.get("cookies") or []
        profile = build_list_org_session_profile(
            cookies=cookies,
            user_agent=current_user_agent or configured_user_agent,
            referer=referer,
            storage_state_path=storage_state_file,
        )
        save_session_profile(session_file, profile)
        browser.close()

    has_cf = any(str(cookie.get("name", "")).strip() == "cf_clearance" for cookie in cookies)
    logger.info(
        "Ð¡Ð¾Ñ…Ñ€Ð°Ð½Ð¸Ð» List-Org session profile: %s | cookies=%s | cf_clearance=%s",
        session_file,
        len(cookies),
        "yes" if has_cf else "no",
    )


def compact_text(value: str | None, limit: int = 220) -> str:
    text = normalize_whitespace(value)
    if not text:
        return ""
    if len(text) <= limit:
        return text
    return text[: max(limit - 1, 0)].rstrip() + "…"


def escape_code_text(value: str) -> str:
    return value.replace("`", "'")


def markdown_link(url: str, label: str | None = None) -> str:
    url = normalize_whitespace(url)
    if not url:
        return "—"
    shown = label or compact_text(url, 80) or url
    shown = shown.replace("[", "\\[").replace("]", "\\]")
    return f"[{shown}]({url})"


def markdown_inline_list(values: Iterable[str], *, code: bool = False) -> str:
    prepared = [normalize_whitespace(value) for value in values if normalize_whitespace(value)]
    if not prepared:
        return "—"
    if code:
        return ", ".join(f"`{escape_code_text(item)}`" for item in prepared)
    return ", ".join(prepared)


def report_file_name(row_index: int, inn: str, company_name: str) -> str:
    return f"{row_index:04d}-{inn}.md"


def render_contact_items_markdown(items: list[dict[str, Any]]) -> list[str]:
    if not items:
        return ["- —"]
    lines: list[str] = []
    for item in items:
        value = normalize_whitespace(str(item.get("value", "")))
        if not value:
            continue
        suffix_parts: list[str] = []
        if item.get("masked"):
            suffix_parts.append("под маской")
        if item.get("note"):
            suffix_parts.append(compact_text(str(item.get("note")), 160))
        source_url = normalize_whitespace(str(item.get("source_url", "")))
        suffix = ""
        if source_url:
            suffix_parts.append(markdown_link(source_url, "source"))
        if suffix_parts:
            suffix = " | " + " | ".join(suffix_parts)
        lines.append(f"- `{value}`{suffix}")
    return lines or ["- —"]


def render_plain_list_markdown(values: Iterable[str], *, code: bool = False) -> list[str]:
    prepared = [normalize_whitespace(value) for value in values if normalize_whitespace(value)]
    if not prepared:
        return ["- —"]
    if code:
        return [f"- `{escape_code_text(item)}`" for item in prepared]
    return [f"- {item}" for item in prepared]


def render_site_decision_markdown(site: dict[str, Any]) -> list[str]:
    url = normalize_whitespace(str(site.get("url", ""))) or "—"
    final_url = normalize_whitespace(str(site.get("final_url", "")))
    reasons = [compact_text(item, 160) for item in (site.get("reasons") or []) if compact_text(item, 160)]
    errors = [compact_text(item, 200) for item in (site.get("errors") or []) if compact_text(item, 200)]
    llm_result = site.get("llm_result") or {}
    lines = [
        f"- Сайт: {markdown_link(url) if url != '—' else '—'}",
        f"  Статус: `{site.get('status', '') or 'unknown'}` | belongs_to_company: `{'yes' if site.get('belongs_to_company') else 'no'}` | industrial_relevance: `{site.get('industrial_relevance', 'unknown')}`",
        f"  Scores: identity=`{site.get('identity_score', 0)}` industrial=`{site.get('industrial_score', 0)}` | decision_source: `{site.get('decision_source', 'heuristics')}`",
    ]
    if final_url:
        lines.append(f"  Final URL: {markdown_link(final_url)}")
    decision_status = normalize_whitespace(str(site.get("decision_status", "")))
    if decision_status:
        lines.append(
            "  Site decision: status=`{}` authenticity=`{}` viability=`{}` conflict_penalty=`{}`".format(
                decision_status,
                site.get("authenticity_score", 0),
                site.get("viability_score", 0),
                site.get("conflict_penalty", 0),
            )
        )
    if reasons:
        lines.append(f"  Reasons: {markdown_inline_list(reasons)}")
    evidence = site.get("evidence") or []
    if evidence:
        lines.append(f"  Evidence: {markdown_inline_list([compact_text(str(item), 120) for item in evidence])}")
    hard_negative_hits = site.get("hard_negative_hits") or []
    if hard_negative_hits:
        lines.append(f"  Hard negatives: {markdown_inline_list([compact_text(str(item), 80) for item in hard_negative_hits])}")
    if errors:
        lines.append(f"  Errors: {markdown_inline_list(errors)}")
    if llm_result:
        llm_reason = compact_text(str(llm_result.get("reason", "")), 220)
        llm_confidence = llm_result.get("confidence", "—")
        llm_type = llm_result.get("website_type", "unknown")
        evidence = llm_result.get("evidence") or []
        contradictions = llm_result.get("contradictions") or []
        lines.append(f"  LLM: website_type=`{llm_type}` confidence=`{llm_confidence}`")
        if llm_reason:
            lines.append(f"  LLM reason: {llm_reason}")
        if evidence:
            lines.append(f"  LLM evidence: {markdown_inline_list([compact_text(str(item), 120) for item in evidence])}")
        if contradictions:
            lines.append(
                f"  LLM contradictions: {markdown_inline_list([compact_text(str(item), 120) for item in contradictions])}"
            )
    extracted_phones = site.get("extracted_phones") or []
    extracted_emails = site.get("extracted_emails") or []
    extracted_addresses = site.get("extracted_addresses") or []
    if extracted_phones:
        lines.append(f"  Phones: {markdown_inline_list(extracted_phones, code=True)}")
    if extracted_emails:
        lines.append(f"  Emails: {markdown_inline_list(extracted_emails, code=True)}")
    if extracted_addresses:
        lines.append(f"  Addresses: {markdown_inline_list(extracted_addresses)}")
    fetched_pages = site.get("fetched_pages") or []
    if fetched_pages:
        lines.append(f"  Fetched pages: {markdown_inline_list([markdown_link(item) for item in fetched_pages])}")
    return lines


def render_domain_resolution_markdown(payload: dict[str, Any]) -> list[str]:
    if not payload:
        return ["- —"]
    lines = [
        f"- Статус: `{payload.get('status', 'not_found')}`",
        f"- Основной домен: {markdown_link(str(payload.get('selected_primary_domain', '') or ''))}",
        f"- Статус основного домена: `{payload.get('selected_primary_status', '—') or '—'}`",
    ]
    notes = payload.get("notes") or []
    if notes:
        lines.append(f"- Notes: {markdown_inline_list(notes)}")
    candidates = payload.get("candidates") or []
    if not candidates:
        lines.append("- Кандидаты: —")
        return lines
    lines.append("- Кандидаты:")
    for item in candidates:
        lines.append(
            f"  - {markdown_link(str(item.get('url', '') or ''))} | status=`{item.get('status', 'candidate')}` | confidence=`{item.get('confidence', 0)}` | source=`{item.get('source', 'unknown')}`"
        )
        evidence = item.get("evidence") or []
        if evidence:
            lines.append(f"    evidence: {markdown_inline_list([compact_text(str(x), 120) for x in evidence])}")
    return lines


def render_site_probe_markdown(probe: dict[str, Any]) -> list[str]:
    url = normalize_whitespace(str(probe.get("url", ""))) or "—"
    final_url = normalize_whitespace(str(probe.get("final_url", "")))
    notes = [compact_text(item, 160) for item in (probe.get("notes") or []) if compact_text(item, 160)]
    errors = [compact_text(item, 160) for item in (probe.get("errors") or []) if compact_text(item, 160)]
    lines = [
        f"- Сайт: {markdown_link(url) if url != '—' else '—'}",
        f"  Probe: status=`{probe.get('status', 'unknown')}` http=`{probe.get('http_status', '—') if probe.get('http_status') is not None else '—'}` class=`{probe.get('site_class', 'F')}` worth_crawling=`{probe.get('worth_crawling', 'false')}` browser_required=`{'yes' if probe.get('browser_required_default') else 'no'}`",
        f"  Metrics: html_ok=`{'yes' if probe.get('html_ok') else 'no'}` anti_bot=`{'yes' if probe.get('anti_bot_detected') else 'no'}` links=`{probe.get('internal_links_count', 0)}` docs=`{probe.get('document_links_count', 0)}` text=`{probe.get('text_length', 0)}`",
    ]
    if final_url:
        lines.append(f"  Final URL: {markdown_link(final_url)}")
    if probe.get("content_type") or probe.get("encoding"):
        lines.append(
            f"  Response: content_type=`{probe.get('content_type', '') or '—'}` encoding=`{probe.get('encoding', '') or '—'}` redirects=`{probe.get('redirect_count', 0)}`"
        )
    if probe.get("robots_found") or probe.get("sitemap_found") or probe.get("cms_guess"):
        lines.append(
            f"  Extra: robots=`{'yes' if probe.get('robots_found') else 'no'}` sitemap=`{'yes' if probe.get('sitemap_found') else 'no'}` cms=`{probe.get('cms_guess', 'unknown') or 'unknown'}`"
        )
    if probe.get("failure_reason") or probe.get("timeout_reason"):
        lines.append(
            f"  Failure: reason=`{probe.get('failure_reason', '') or '—'}` timeout_reason=`{probe.get('timeout_reason', '') or '—'}`"
        )
    obvious_routes = probe.get("obvious_routes_attempted") or []
    if obvious_routes:
        lines.append(f"  Obvious routes checked: `{len(obvious_routes)}`")
    key_sections = probe.get("key_sections") or []
    if key_sections:
        lines.append(f"  Key sections: {markdown_inline_list(key_sections)}")
    sampled_urls = probe.get("sampled_urls") or []
    if sampled_urls:
        lines.append(f"  Sampled URLs: {markdown_inline_list([markdown_link(item) for item in sampled_urls])}")
    if notes:
        lines.append(f"  Notes: {markdown_inline_list(notes)}")
    if errors:
        lines.append(f"  Errors: {markdown_inline_list(errors)}")
    return lines


def render_route_strategy_markdown(item: dict[str, Any]) -> list[str]:
    route = normalize_whitespace(str(item.get("route_pattern", ""))) or "—"
    reasons = [compact_text(reason, 120) for reason in (item.get("reasons") or []) if compact_text(reason, 120)]
    lines = [
        f"- Route: {markdown_link(route) if route.startswith('http') else route}",
        f"  section=`{item.get('section_guess', 'unknown') or 'unknown'}` | mode=`{item.get('mode', 'skip') or 'skip'}` | confidence=`{item.get('confidence', 0)}`",
    ]
    if reasons:
        lines.append(f"  Reasons: {markdown_inline_list(reasons)}")
    return lines


def render_content_record_markdown(item: dict[str, Any]) -> list[str]:
    url = normalize_whitespace(str(item.get("url", ""))) or "—"
    notes = [compact_text(note, 120) for note in (item.get("notes") or []) if compact_text(note, 120)]
    lines = [
        f"- Record: {markdown_link(url) if url != '—' else '—'}",
        f"  type=`{item.get('source_type', 'unknown')}` | section=`{item.get('section_guess', 'unknown') or 'unknown'}` | extraction=`{item.get('extraction_method', 'unknown') or 'unknown'}` | fetch_status=`{item.get('fetch_status', 'unknown') or 'unknown'}` | relevance=`{item.get('relevance_label', 'unknown')}` score=`{item.get('relevance_score', 0)}`",
    ]
    if item.get("title"):
        lines.append(f"  Title: {compact_text(str(item.get('title')), 160)}")
    if item.get("date"):
        lines.append(f"  Date: `{item.get('date')}`")
    if item.get("cleaned_text"):
        lines.append(f"  Text: {compact_text(str(item.get('cleaned_text')), 260)}")
    if item.get("relevance_reasons"):
        lines.append(f"  Relevance reasons: {markdown_inline_list([compact_text(str(x), 120) for x in (item.get('relevance_reasons') or [])])}")
    if item.get("llm_result"):
        llm_result = item.get("llm_result") or {}
        lines.append(
            f"  LLM: label=`{llm_result.get('relevance_label', 'unknown')}` confidence=`{llm_result.get('confidence', '—')}` summary={compact_text(str(llm_result.get('summary', '')), 180) or '—'}"
        )
    if notes:
        lines.append(f"  Notes: {markdown_inline_list(notes)}")
    return lines


def render_lead_card_markdown(item: dict[str, Any]) -> list[str]:
    source_urls = item.get("source_urls") or []
    contacts = item.get("contacts") or {}
    lines = [
        f"- Lead: {compact_text(str(item.get('title', '') or 'Без названия'), 160)}",
        f"  type=`{item.get('lead_type', 'unknown')}` | confidence=`{item.get('confidence', 0)}` | status=`{item.get('status', 'new')}`",
        f"  Why: {compact_text(str(item.get('why_relevant', '') or ''), 200) or '—'}",
    ]
    if item.get("date"):
        lines.append(f"  Date: `{item.get('date')}`")
    if source_urls:
        lines.append(f"  Sources: {markdown_inline_list([markdown_link(url) for url in source_urls])}")
    if contacts:
        lines.append(
            f"  Contacts: phones={markdown_inline_list(contacts.get('phones') or [], code=True)} | emails={markdown_inline_list(contacts.get('emails') or [], code=True)}"
        )
    return lines


def render_refresh_plan_markdown(item: dict[str, Any]) -> list[str]:
    return [
        f"- {markdown_link(str(item.get('site_url', '') or ''))} | cadence=`{item.get('cadence', '') or 'unknown'}` | next_due=`{item.get('next_due_at', '') or '—'}` | reason={compact_text(str(item.get('reason', '') or ''), 140) or '—'}"
    ]


def render_leads_report_markdown(
    ordered_results: list[dict[str, Any]],
    *,
    summary: dict[str, Any] | None = None,
) -> str:
    ordered_results = [repair_output_value(result) for result in ordered_results]
    lines = [
        "# Leads Report",
        "",
        f"- Updated: `{utc_now_iso()}`",
    ]
    status_lines = render_run_status_markdown(summary)
    if status_lines:
        lines.extend(status_lines)
    lines.append("")
    total_leads = 0
    for result in ordered_results:
        profile = company_profile_payload_from_result(result)
        summary = profile.get("summary") or {}
        lead_cards = result.get("lead_cards") or []
        if not lead_cards:
            continue
        total_leads += len(lead_cards)
        lines.append(f"## {summary.get('company_name', result.get('company_name', ''))} ({summary.get('inn', result.get('inn', ''))})")
        for item in lead_cards:
            lines.extend(render_lead_card_markdown(item))
        lines.append("")
    if total_leads == 0:
        lines.append("Лидов пока нет.")
    else:
        lines.insert(3, f"- Total leads: `{total_leads}`")
    return "\n".join(lines).strip() + "\n"


def render_availability_markdown(availability: dict[str, Any]) -> list[str]:
    if not availability:
        return ["- —"]
    lines: list[str] = []
    for field_name in IMPORTANT_FIELDS:
        payload = availability.get(field_name) or {}
        status = normalize_source_availability_status(str(payload.get("status", "")))
        parts = [f"`{status}`"]
        if payload.get("open_count") is not None:
            parts.append(f"open_count={payload.get('open_count')}")
        if payload.get("masked_examples"):
            parts.append(f"masked_examples={markdown_inline_list(payload.get('masked_examples') or [], code=True)}")
        if payload.get("reason"):
            parts.append(compact_text(str(payload.get("reason")), 220))
        lines.append(f"- `{field_name}`: " + " | ".join(parts))
    return lines


def render_company_report_markdown(result: dict[str, Any]) -> str:
    result = repair_output_value(result)
    profile = company_profile_payload_from_result(result)
    summary = profile.get("summary") or {}
    contacts = profile.get("contacts") or {}
    sites = profile.get("sites") or {}
    merged = contacts.get("raw") or {}
    trusted = contacts.get("trusted") or {}
    domain_resolution = result.get("domain_resolution") or {}
    site_probes = result.get("site_probes") or []
    route_strategies = result.get("route_strategies") or []
    content_records = result.get("content_records") or []
    lead_cards = result.get("lead_cards") or []
    site_refresh_plans = result.get("site_refresh_plans") or []
    validated_sites = result.get("validated_sites") or []
    candidate_sites = result.get("candidate_sites") or []
    confirmed_sites = [
        site
        for site in validated_sites
        if site.get("decision_status") == "verified" or (not site.get("decision_status") and site.get("belongs_to_company"))
    ]

    lines = [
        f"# {summary.get('company_name', result.get('company_name', ''))} ({summary.get('inn', result.get('inn', ''))})",
        "",
        f"- Строка XLSX: `{result.get('row_index', 0)}`",
        f"- Статус: `{summary.get('processing_status', result.get('status', 'unknown'))}`",
        f"- Started: `{result.get('started_at', '') or '—'}`",
        f"- Finished: `{result.get('finished_at', '') or '—'}`",
        f"- Input site: {markdown_link(str(result.get('input_site', '') or ''))}",
        f"- Input phone: {markdown_inline_list([str(result.get('input_phone', '') or '')], code=True)}",
        f"- Комментарий: {compact_text(str(result.get('input_comment', '') or ''), 400) or '—'}",
        "",
        "## Domain Resolution",
    ]
    lines.extend(render_domain_resolution_markdown(domain_resolution))
    lines.extend(
        [
            "",
            "## Site Probes",
        ]
    )
    if site_probes:
        for probe in site_probes:
            lines.extend(render_site_probe_markdown(probe))
    else:
        lines.append("- —")

    lines.extend(
        [
            "",
            "## Route Strategies",
        ]
    )
    if route_strategies:
        for item in route_strategies:
            lines.extend(render_route_strategy_markdown(item))
    else:
        lines.append("- —")

    lines.extend(
        [
            "",
            "## Content Records",
        ]
    )
    if content_records:
        for item in content_records[:12]:
            lines.extend(render_content_record_markdown(item))
        if len(content_records) > 12:
            lines.append(f"- … еще `{len(content_records) - 12}` records")
    else:
        lines.append("- —")

    lines.extend(
        [
            "",
            "## Lead Cards",
        ]
    )
    if lead_cards:
        for item in lead_cards:
            lines.extend(render_lead_card_markdown(item))
    else:
        lines.append("- —")

    lines.extend(
        [
            "",
            "## Refresh Plan",
        ]
    )
    if site_refresh_plans:
        for item in site_refresh_plans:
            lines.extend(render_refresh_plan_markdown(item))
    else:
        lines.append("- —")

    lines.extend(
        [
            "",
            "## Trusted Contacts",
            f"- Телефоны: {markdown_inline_list(trusted.get('phones') or [], code=True)}",
            f"- Email: {markdown_inline_list(trusted.get('emails') or [], code=True)}",
            f"- Сайты: {markdown_inline_list([markdown_link(item) for item in (trusted.get('websites') or [])])}",
            f"- Адреса: {markdown_inline_list(trusted.get('addresses') or [])}",
            "",
            "## Raw Merged Contacts",
            f"- Телефоны: {markdown_inline_list(merged.get('phones') or [], code=True)}",
            f"- Email: {markdown_inline_list(merged.get('emails') or [], code=True)}",
            f"- Сайты: {markdown_inline_list([markdown_link(item) for item in (merged.get('websites') or [])])}",
            f"- Адреса: {markdown_inline_list(merged.get('addresses') or [])}",
            "",
            "## Сайты",
            f"- Кандидаты: {markdown_inline_list([markdown_link(item) for item in (sites.get('candidate_sites') or candidate_sites)])}",
            f"- Подтвержденные сайты: {markdown_inline_list([markdown_link(item) for item in (sites.get('confirmed_sites') or [site.get('final_url') or site.get('url') or '' for site in confirmed_sites])])}",
            "",
            "### Проверка сайтов",
        ]
    )
    if validated_sites:
        for site in validated_sites:
            lines.extend(render_site_decision_markdown(site))
    else:
        lines.append("- —")

    notes = result.get("notes") or []
    if notes:
        lines.extend(["", "## Notes"])
        lines.extend(render_plain_list_markdown(notes))

    lines.extend(["", "## Источники"])
    for source_name, source_payload in (result.get("sources") or {}).items():
        lines.extend(
            [
                "",
                f"### {source_name}",
                f"- Статус: `{source_payload.get('status', 'unknown')}` | http_status: `{source_payload.get('http_status') if source_payload.get('http_status') is not None else '—'}`",
                f"- Search: {markdown_link(str(source_payload.get('search_url', '') or ''))}",
                f"- Listing: {markdown_link(str(source_payload.get('listing_url', '') or ''))}",
                f"- Entity: {markdown_link(str(source_payload.get('entity_url', '') or ''))}",
                f"- Название в источнике: {compact_text(str(source_payload.get('company_name_found', '') or ''), 220) or '—'}",
                "- Availability:",
            ]
        )
        lines.extend(render_availability_markdown(source_payload.get("availability") or {}))
        if source_payload.get("masked_rows"):
            lines.append("- Masked rows:")
            lines.extend(render_plain_list_markdown(source_payload.get("masked_rows") or []))
        lines.append("- Phones:")
        lines.extend(render_contact_items_markdown(source_payload.get("phones") or []))
        lines.append("- Emails:")
        lines.extend(render_contact_items_markdown(source_payload.get("emails") or []))
        lines.append("- Websites:")
        lines.extend(render_contact_items_markdown(source_payload.get("websites") or []))
        lines.append("- Addresses:")
        lines.extend(render_contact_items_markdown(source_payload.get("addresses") or []))
        if source_payload.get("links"):
            lines.append("- Links:")
            lines.extend(render_plain_list_markdown([markdown_link(item) for item in (source_payload.get("links") or [])]))
        primary_okved_display = okved_entry_display(source_payload.get("primary_okved"))
        additional_okved_displays = okved_entries_display(source_payload.get("additional_okveds"), limit=3, max_len=220)
        if primary_okved_display:
            lines.append(f"- Основной ОКВЭД: {primary_okved_display}")
        if additional_okved_displays:
            lines.append("- Дополнительные ОКВЭД:")
            lines.extend(render_plain_list_markdown(additional_okved_displays))
        if source_payload.get("notes"):
            lines.append("- Notes:")
            lines.extend(render_plain_list_markdown(source_payload.get("notes") or []))
        if source_payload.get("errors"):
            lines.append("- Errors:")
            lines.extend(render_plain_list_markdown(source_payload.get("errors") or []))
        if source_payload.get("snippets"):
            lines.append("- Snippets:")
            lines.extend(render_plain_list_markdown([compact_text(item, 260) for item in (source_payload.get("snippets") or [])]))

    return "\n".join(lines).strip() + "\n"


def source_issue_summary(source_payload: dict[str, Any]) -> str:
    parts: list[str] = []
    status = source_payload.get("status", "")
    if status and status != "success":
        parts.append(status)
    masked_fields: list[str] = []
    blocked_fields: list[str] = []
    for field_name, payload in (source_payload.get("availability") or {}).items():
        field_status = normalize_source_availability_status(str(payload.get("status", "")))
        if field_status == "masked":
            masked_fields.append(field_name)
        if field_status == "blocked":
            blocked_fields.append(field_name)
    if masked_fields:
        parts.append("masked: " + ", ".join(masked_fields))
    if blocked_fields:
        parts.append("blocked: " + ", ".join(blocked_fields))
    return "; ".join(parts)


def render_run_status_markdown(summary: dict[str, Any] | None) -> list[str]:
    summary = summary or {}
    run_status = normalize_whitespace(str(summary.get("run_status", "") or ""))
    if not run_status:
        return []
    lines = [f"- Run status: `{run_status}`"]
    finish_reason = normalize_whitespace(str(summary.get("finish_reason", "") or ""))
    if finish_reason:
        lines.append(f"- Finish reason: `{finish_reason}`")
    finished_at = normalize_whitespace(str(summary.get("finished_at", "") or ""))
    if finished_at:
        lines.append(f"- Finished at: `{finished_at}`")
    stop_requested_at = normalize_whitespace(str(summary.get("stop_requested_at", "") or ""))
    if stop_requested_at:
        lines.append(f"- Stop requested at: `{stop_requested_at}`")
    stop_reason = normalize_whitespace(str(summary.get("stop_reason", "") or ""))
    if stop_reason:
        lines.append(f"- Stop note: {compact_text(stop_reason, 220)}")
    terminal_source = normalize_whitespace(str(summary.get("terminal_source", "") or ""))
    if terminal_source:
        lines.append(f"- Terminal source: `{terminal_source}`")
    terminal_source_status = normalize_whitespace(str(summary.get("terminal_source_status", "") or ""))
    if terminal_source_status:
        lines.append(f"- Terminal source status: `{terminal_source_status}`")
    terminal_source_access_mode = normalize_whitespace(str(summary.get("terminal_source_access_mode", "") or ""))
    if terminal_source_access_mode:
        lines.append(f"- Terminal source access mode: `{terminal_source_access_mode}`")
    terminal_error_type = normalize_whitespace(str(summary.get("terminal_error_type", "") or ""))
    if terminal_error_type:
        lines.append(f"- Terminal error type: `{terminal_error_type}`")
    terminal_error_message = normalize_whitespace(str(summary.get("terminal_error_message", "") or ""))
    if terminal_error_message:
        lines.append(f"- Terminal error: {compact_text(terminal_error_message, 220)}")
    try:
        unresolved_required_source_rows = int(summary.get("unresolved_required_source_rows", 0) or 0)
    except (TypeError, ValueError):
        unresolved_required_source_rows = 0
    if unresolved_required_source_rows:
        lines.append(f"- Deferred required-source rows: `{unresolved_required_source_rows}`")
        by_source = summary.get("required_source_deferred_rows_by_source") or {}
        if isinstance(by_source, Mapping) and by_source:
            rendered = ", ".join(f"{source}={count}" for source, count in sorted(by_source.items()))
            lines.append(f"- Deferred required-source rows by source: {rendered}")
    if run_status == "controlled_stop":
        lines.append("- Public outputs include only completed companies; pending work remains in runtime state.")
    elif unresolved_required_source_rows:
        lines.append("- Public outputs include only completed companies; deferred required-source rows are unresolved.")
    elif run_status != "completed":
        lines.append("- Public outputs include only completed companies up to the terminal stop boundary.")
    return lines


def render_index_report_markdown(
    ordered_results: list[dict[str, Any]],
    *,
    summary: dict[str, Any] | None,
    availability_summary: dict[str, Any],
    host_stats: dict[str, Any],
) -> str:
    ordered_results = [repair_output_value(result) for result in ordered_results]
    missing_value = "—"
    domain_status_counts = Counter(
        (company_profile_payload_from_result(result).get("summary") or {}).get("domain_resolution_status", "not_found")
        or "not_found"
        for result in ordered_results
    )
    site_class_counts: Counter[str] = Counter()
    for result in ordered_results:
        for probe in result.get("site_probes") or []:
            site_class_counts[str(probe.get("site_class", "F") or "F")] += 1
    lines = [
        "# Company Research Report",
        "",
        f"- Updated: `{utc_now_iso()}`",
        f"- Rows in report: `{len(ordered_results)}`",
        f"- Summary total_rows: `{(summary or {}).get('total_rows', missing_value)}`",
        f"- Summary processed_rows: `{(summary or {}).get('processed_rows', missing_value)}`",
        f"- Summary completed_rows: `{(summary or {}).get('completed_rows', len(ordered_results))}`",
        f"- Summary remaining_rows: `{(summary or {}).get('remaining_rows', missing_value)}`",
    ]
    status_lines = render_run_status_markdown(summary)
    if status_lines:
        lines.extend(status_lines)
    lines.extend(
        [
            "",
            "## Domain Resolution",
            f"- verified: `{domain_status_counts.get('verified', 0)}` | candidate: `{domain_status_counts.get('candidate', 0)}` | not_found: `{domain_status_counts.get('not_found', 0)}`",
            "",
            "## Site Classes",
            f"- A: `{site_class_counts.get('A', 0)}` | B: `{site_class_counts.get('B', 0)}` | C: `{site_class_counts.get('C', 0)}` | D: `{site_class_counts.get('D', 0)}` | E: `{site_class_counts.get('E', 0)}` | F: `{site_class_counts.get('F', 0)}`",
            "",
            "## Доступность полей по агрегаторам",
        ]
    )
    sources_block = availability_summary.get("sources") or {}
    if sources_block:
        for source_name, source_fields in sources_block.items():
            lines.append(f"- `{source_name}`")
            for field_name in IMPORTANT_FIELDS:
                counts = source_fields.get(field_name) or {}
                rendered = ", ".join(f"{key}={counts.get(key, 0)}" for key in SOURCE_AVAILABILITY_STATUSES)
                lines.append(f"  - `{field_name}`: {rendered}")
    else:
        lines.append(f"- {missing_value}")

    lines.extend(["", "## Поведение хостов"])
    if host_stats:
        for host, payload in sorted(host_stats.items()):
            event_types = payload.get("event_types") or {}
            interval_stats = payload.get("interval_seconds") or {}
            cooldown_stats = payload.get("cooldown_seconds") or {}
            lines.append(
                f"- `{host}` | events={payload.get('total_events', 0)} | ok={event_types.get('request_ok', 0)} | 429={event_types.get('rate_limited', 0)} | bot_gate={event_types.get('bot_gate', 0)} | avg_interval={interval_stats.get('avg', missing_value)} | max_cooldown={cooldown_stats.get('max', 0)}"
            )
    else:
        lines.append(f"- {missing_value}")

    lines.extend(["", "## Компании"])
    if not ordered_results:
        lines.append(f"- {missing_value}")
        return "\n".join(lines).strip() + "\n"

    for result in ordered_results:
        profile = company_profile_payload_from_result(result)
        summary = profile.get("summary") or {}
        contacts = profile.get("contacts") or {}
        sites = profile.get("sites") or {}
        merged = contacts.get("raw") or {}
        trusted = contacts.get("trusted") or {}
        report_name = report_file_name(
            int(result.get("row_index", 0) or 0),
            str(summary.get("inn", result.get("inn", "")) or ""),
            str(summary.get("company_name", result.get("company_name", "")) or ""),
        )
        domain_status = summary.get("domain_resolution_status", "not_found") or "not_found"
        primary_domain = sites.get("primary_domain", "")
        probe_classes = ",".join(sites.get("site_classes") or []) or missing_value
        primary_domain_markdown = markdown_link(str(primary_domain or "")) if primary_domain else missing_value
        lead_count = int(summary.get("lead_count", 0) or 0)
        lines.append(
            f"- [{summary.get('inn', result.get('inn', ''))} — {summary.get('company_name', result.get('company_name', ''))}](company_reports/{report_name}) | статус: `{summary.get('processing_status', result.get('status', 'unknown'))}` | domain=`{domain_status}` | primary={primary_domain_markdown} | probes=`{probe_classes}` | leads=`{lead_count}` | trusted телефоны: `{len(trusted.get('phones') or [])}` | trusted email: `{len(trusted.get('emails') or [])}` | trusted сайты: `{len(trusted.get('websites') or [])}` | raw сайты: `{len(merged.get('websites') or [])}` | подтверждено сайтов: `{len(sites.get('confirmed_sites') or [])}`"
        )
        if summary.get("issues"):
            lines.append(f"  issues: {markdown_inline_list(summary.get('issues') or [])}")

    return "\n".join(lines).strip() + "\n"


def stored_contact_values(source_payload: dict[str, Any], kind: str) -> list[str]:
    values: list[str] = []
    for item in source_payload.get(kind) or []:
        if isinstance(item, dict):
            value = normalize_whitespace(str(item.get("value", "")))
        else:
            value = normalize_whitespace(str(item))
        if value:
            values.append(value)
    return values


def contact_value_matches(kind: str, left: str, right: str) -> bool:
    if kind == "phones":
        return normalized_phone_digits(left) and normalized_phone_digits(left) == normalized_phone_digits(right)
    if kind == "emails":
        return normalize_whitespace(left).lower() == normalize_whitespace(right).lower()
    if kind == "websites":
        left_url = sanitize_website_url(left)
        right_url = sanitize_website_url(right)
        if not left_url or not right_url:
            return False
        return guess_registered_domain(urlparse(left_url).netloc) == guess_registered_domain(urlparse(right_url).netloc)
    if kind == "addresses":
        return sanitize_address_candidate(left).lower() == sanitize_address_candidate(right).lower()
    return normalize_whitespace(left) == normalize_whitespace(right)


def sources_for_contact_value(result: dict[str, Any], kind: str, value: str) -> list[str]:
    labels: list[str] = []
    if not value:
        return labels
    for source_name, source_payload in (result.get("sources") or {}).items():
        if any(contact_value_matches(kind, source_value, value) for source_value in stored_contact_values(source_payload, kind)):
            labels.append(source_name)
            continue
        if kind == "websites":
            for email_value in stored_contact_values(source_payload, "emails"):
                inferred_website = email_to_candidate_website(email_value)
                if inferred_website and contact_value_matches("websites", inferred_website, value):
                    labels.append(f"{source_name}:email_domain")
                    break
    if kind == "websites":
        domain_resolution = result.get("domain_resolution") or {}
        for candidate in domain_resolution.get("candidates") or []:
            candidate_url = normalize_whitespace(str(candidate.get("url", "")))
            if candidate_url and contact_value_matches("websites", candidate_url, value):
                source_name = normalize_whitespace(str(candidate.get("source", "")))
                if source_name:
                    labels.append(f"domain_resolution:{source_name}")
    return dedupe_preserve_order(labels)


ADDRESS_REGION_SURFACE_RE = re.compile(
    r"\b(?:обл(?:асть)?|край|респ(?:ублика)?|автоном(?:ный|ная)\s+округ|ао)\b",
    flags=re.IGNORECASE,
)
ADDRESS_DISTRICT_SURFACE_RE = re.compile(
    r"\b(?:р-?н|район|муниципальный\s+округ|муниципальный\s+район)\b",
    flags=re.IGNORECASE,
)
ADDRESS_CITY_SURFACE_RE = re.compile(r"(?:\bг\.\s*[a-zа-яё0-9-]+|\bгород\b)", flags=re.IGNORECASE)
ADDRESS_SETTLEMENT_SURFACE_RE = re.compile(
    r"(?:\b(?:д|дер|деревня|с|село|пос|поселок|пгт|хутор|станица|слобода|аул|кп)\.?\s*[a-zа-яё0-9-]+|\b(?:деревня|село|поселок|хутор|станица|слобода|аул)\b)",
    flags=re.IGNORECASE,
)
ADDRESS_TERRITORY_SURFACE_RE = re.compile(
    r"\b(?:тер\.?|территория|кпо|промзона|мкр\.?|микрорайон|квартал|массив|площадка)\b",
    flags=re.IGNORECASE,
)
ADDRESS_STREET_SURFACE_RE = re.compile(
    r"\b(?:ул\.?|улица|пр-кт|просп\.?|проспект|пер\.?|переулок|ш\.?|шоссе|проезд|б-р|бульвар|наб\.?|набережная|тракт|аллея|тупик)\b",
    flags=re.IGNORECASE,
)
ADDRESS_HOUSE_SURFACE_RE = re.compile(
    r"\b(?:д\.?|дом|влд\.?|владение)\s*[0-9а-яa-z/-]+",
    flags=re.IGNORECASE,
)
ADDRESS_BUILDING_SURFACE_RE = re.compile(
    r"\b(?:к\.|корп\.?|корпус|стр\.?|строение|лит\.?|литер|литера)\s*[0-9а-яa-z/-]+",
    flags=re.IGNORECASE,
)
ADDRESS_VALID_SINGLE_LETTER_SUFFIX_RE = re.compile(
    r"\b(?:к\.|корп\.?|корпус|стр\.?|строение|лит\.?|литер|литера|оф\.?|офис|пом\.?|помещ\.?|помещение)\s*[а-яa-z]$",
    flags=re.IGNORECASE,
)
ADDRESS_NUMBERED_SINGLE_LETTER_SUFFIX_RE = re.compile(
    r"(\b(?:д\.?|дом|влд\.?|владение|корп\.?|корпус|стр\.?|строение|лит\.?|литер|литера|офис|оф\.?|каб\.?|кабинет|комн\.?|комната|пом\.?|помещ\.?|помещение)\s*[0-9/-]+)\s*(-?)\s*([а-яa-z])\b",
    flags=re.IGNORECASE,
)
ADDRESS_UNIT_SURFACE_RE = re.compile(
    r"\b(?:офис|оф\.?|каб\.?|кабинет|комн\.?|комната|пом\.?|помещение|этаж)\s*[0-9а-яa-z/-]+",
    flags=re.IGNORECASE,
)
ADDRESS_TRUNCATED_SURFACE_RE = re.compile(
    r"(?:[,;:/-]\s*|\s)(?:обл\.?|область|край|респ\.?|республика|р-?н|район|г\.?|город|ул\.?|улица|пр-кт|просп\.?|проспект|д\.?|дом|тер\.?|территория)$",
    flags=re.IGNORECASE,
)
ADDRESS_GARBAGE_SURFACE_PATTERNS = (
    re.compile(r"https?://|www\.|@[a-z0-9._%+-]+\.[a-z]{2,}", flags=re.IGNORECASE),
    re.compile(r"\b(?:тел(?:ефон)?|e-?mail|почта|сайт|www)\b", flags=re.IGNORECASE),
    re.compile(
        r"\b(?:производительност[ьи]|мощност[ьи]|диаметр|длина|ширина|высота|об[ъеё]м|масса|вес|артикул|модель|серия|гост|характеристик[аи]|т/год|т/сутки|тонн?|кг|квт|м2|м3|мм|см)\b",
        flags=re.IGNORECASE,
    ),
)
ADDRESS_COMPONENT_WEIGHTS = {
    "postal": 2,
    "region": 2,
    "district": 2,
    "city": 4,
    "settlement": 4,
    "territory": 3,
    "street": 5,
    "house": 6,
    "building": 2,
    "unit": 1,
}


@dataclass(frozen=True)
class AddressSurfaceScore:
    garbage_penalty: int
    lookup_status_rank: int
    specificity_score: int
    locality_score: int
    oracle_locality_detail: tuple[int, int, int, int]
    source_count: int
    segment_count: int
    length: int

    def sort_key(self) -> tuple[int, ...]:
        return (
            -self.garbage_penalty,
            self.lookup_status_rank,
            self.specificity_score,
            self.locality_score,
            *self.oracle_locality_detail,
            self.source_count,
            self.segment_count,
            self.length,
        )


def _address_surface_component_hits(value: str) -> dict[str, bool]:
    return {
        "postal": bool(re.search(r"\b\d{6}\b", value)),
        "region": bool(ADDRESS_REGION_SURFACE_RE.search(value)),
        "district": bool(ADDRESS_DISTRICT_SURFACE_RE.search(value)),
        "city": bool(ADDRESS_CITY_SURFACE_RE.search(value)),
        "settlement": bool(ADDRESS_SETTLEMENT_SURFACE_RE.search(value)),
        "territory": bool(ADDRESS_TERRITORY_SURFACE_RE.search(value)),
        "street": bool(ADDRESS_STREET_SURFACE_RE.search(value)),
        "house": bool(ADDRESS_HOUSE_SURFACE_RE.search(value)),
        "building": bool(ADDRESS_BUILDING_SURFACE_RE.search(value)),
        "unit": bool(ADDRESS_UNIT_SURFACE_RE.search(value)),
    }


def _address_surface_garbage_penalty(value: str) -> int:
    penalty = sum(1 for pattern in ADDRESS_GARBAGE_SURFACE_PATTERNS if pattern.search(value))
    if re.search(r"\b\d{2,}\s*[xх*]\s*\d{2,}\b", value, flags=re.IGNORECASE):
        penalty += 1
    return penalty


def _score_address_surface(result: dict[str, Any], candidate: AddressEnrichment) -> AddressSurfaceScore:
    cleaned = candidate.sanitized_value
    if not cleaned:
        return AddressSurfaceScore(
            garbage_penalty=99,
            lookup_status_rank=0,
            specificity_score=0,
            locality_score=0,
            oracle_locality_detail=(0, 0, 0, 0),
            source_count=0,
            segment_count=0,
            length=0,
        )

    component_hits = _address_surface_component_hits(cleaned)
    identity = address_identity_tokens(cleaned)
    specificity_score = sum(
        weight
        for component_name, weight in ADDRESS_COMPONENT_WEIGHTS.items()
        if component_hits.get(component_name)
    )
    if component_hits["street"] and component_hits["house"]:
        specificity_score += 4
    if component_hits["territory"] and any(
        component_hits[name] for name in ("district", "city", "settlement")
    ):
        specificity_score += 2
    if component_hits["region"] and not any(
        component_hits[name]
        for name in ("district", "city", "settlement", "territory", "street", "house")
    ):
        specificity_score -= 4
    if ADDRESS_TRUNCATED_SURFACE_RE.search(cleaned):
        specificity_score -= 3
    specificity_score += min(len(identity["tokens"]), 6)

    locality_score = (
        len(identity["tokens"])
        + len(identity["postals"])
        + int(component_hits["city"])
        + int(component_hits["settlement"])
        + int(component_hits["street"])
        + int(component_hits["house"])
    )
    segment_count = sum(1 for item in re.split(r"[,;/]", cleaned) if normalize_whitespace(item))
    source_count = len(sources_for_contact_value(result, "addresses", cleaned))
    garbage_penalty = _address_surface_garbage_penalty(cleaned)

    return AddressSurfaceScore(
        garbage_penalty=garbage_penalty,
        lookup_status_rank=candidate.lookup_status_rank,
        specificity_score=specificity_score,
        locality_score=locality_score,
        oracle_locality_detail=candidate.locality_detail_rank,
        source_count=source_count,
        segment_count=segment_count,
        length=len(cleaned),
    )


def pick_best_address_value(result: dict[str, Any], values: Iterable[str]) -> str:
    candidates = enrich_address_candidates(values, sanitizer=sanitize_address_candidate)
    if not candidates:
        return ""
    # Address surface ranking should prefer geo-useful specificity over short coarse strings.
    return max(candidates, key=lambda candidate: _score_address_surface(result, candidate).sort_key()).sanitized_value


def pick_best_contact_value(result: dict[str, Any], kind: str, values: list[str]) -> str:
    prepared = [normalize_whitespace(value) for value in values if normalize_whitespace(value)]
    if not prepared:
        return ""
    scored = sorted(
        prepared,
        key=lambda value: (
            -len(sources_for_contact_value(result, kind, value)),
            len(value),
        ),
    )
    return scored[0]


def _site_export_decision_status(site: dict[str, Any]) -> str:
    status = normalize_whitespace(str(site.get("decision_status") or ""))
    if status == "ambiguous":
        return "suspicious"
    if status:
        return status
    if site.get("belongs_to_company"):
        return "verified"
    return "rejected"


def _site_export_status_rank(site: dict[str, Any]) -> int:
    status = _site_export_decision_status(site)
    return SITE_AUTH_STATUS_RANK.get(status, 9)


def _ranked_validated_sites(validated_sites: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        validated_sites,
        key=lambda site: (
            _site_export_status_rank(site),
            -float(site.get("authenticity_score", 0.0) or 0.0),
            -float(site.get("identity_score", 0.0) or 0.0),
        ),
    )


def choose_best_company_contacts(result: dict[str, Any]) -> dict[str, Any]:
    trusted = result.get("trusted_contacts") or {}
    merged = result.get("merged_contacts") or {}
    validated_sites = result.get("validated_sites") or []
    domain_resolution = result.get("domain_resolution") or {}

    ranked_validated_sites = _ranked_validated_sites(validated_sites)
    confirmed_site_domains = {
        guess_registered_domain(urlparse(sanitize_website_url(str(site.get("final_url") or site.get("url") or ""))).netloc)
        for site in ranked_validated_sites
        if _site_export_decision_status(site) == "verified"
    }
    confirmed_site_domains.discard("")

    best_site = ""
    best_site_status = "none"
    best_validated = ranked_validated_sites[0] if ranked_validated_sites else None
    if best_validated:
        best_site = sanitize_website_url(str(best_validated.get("final_url") or best_validated.get("url") or ""))
        best_site_status = _site_export_decision_status(best_validated)
    else:
        trusted_sites = trusted.get("websites") or []
        if trusted_sites:
            best_site = trusted_sites[0]
            domain = guess_registered_domain(urlparse(sanitize_website_url(best_site)).netloc)
            best_site_status = "trusted_validated" if domain in confirmed_site_domains else "trusted"
        elif domain_resolution.get("selected_primary_domain"):
            best_site = normalize_whitespace(str(domain_resolution.get("selected_primary_domain", "")))
            best_site_status = "domain_resolution"
        elif merged.get("websites"):
            best_site = str((merged.get("websites") or [""])[0])
            best_site_status = "merged_only"

    best_phone = pick_best_contact_value(result, "phones", list(trusted.get("phones") or merged.get("phones") or []))
    best_email = pick_best_contact_value(result, "emails", list(trusted.get("emails") or merged.get("emails") or []))
    best_address = pick_best_address_value(
        result,
        [
            *(trusted.get("addresses") or []),
            *(merged.get("addresses") or []),
        ],
    )

    return {
        "best_site": best_site,
        "best_site_status": best_site_status,
        "best_site_sources": ", ".join(sources_for_contact_value(result, "websites", best_site)) if best_site else "",
        "best_phone": best_phone,
        "best_phone_sources": ", ".join(sources_for_contact_value(result, "phones", best_phone)) if best_phone else "",
        "best_email": best_email,
        "best_email_sources": ", ".join(sources_for_contact_value(result, "emails", best_email)) if best_email else "",
        "best_address": best_address,
        "best_address_sources": ", ".join(sources_for_contact_value(result, "addresses", best_address)) if best_address else "",
    }


def summarize_company_decision(result: dict[str, Any]) -> str:
    validated_sites = _ranked_validated_sites(result.get("validated_sites") or [])
    for site in validated_sites:
        if _site_export_decision_status(site) in {"verified", "candidate"}:
            reasons = site.get("reasons") or []
            return compact_text("; ".join(str(item) for item in reasons[:3]), 240)
    domain_resolution = result.get("domain_resolution") or {}
    notes = domain_resolution.get("notes") or []
    if notes:
        return compact_text("; ".join(str(item) for item in notes[:3]), 240)
    issues = [
        f"{source_name}: {issue}"
        for source_name, source_payload in (result.get("sources") or {}).items()
        if (issue := source_issue_summary(source_payload))
    ]
    return compact_text("; ".join(issues[:3]), 240)


def _normalized_company_contact_lists(payload: dict[str, Any] | None) -> dict[str, list[str]]:
    source = payload if isinstance(payload, dict) else {}
    normalized: dict[str, list[str]] = {}
    for field_name in COMPANY_CONTACT_FIELDS:
        values: list[str] = []
        for item in source.get(field_name) or []:
            text = str(item or "")
            if text:
                values.append(text)
        normalized[field_name] = values
    return normalized


def _split_profile_sources(value: Any) -> list[str]:
    if isinstance(value, str):
        text = normalize_whitespace(value)
        if not text:
            return []
        delimiter = "|" if "|" in text else ","
        parts = [normalize_whitespace(item) for item in text.split(delimiter)]
        return dedupe_preserve_order(item for item in parts if item)
    if isinstance(value, (list, tuple, set)):
        return dedupe_preserve_order(normalize_whitespace(str(item)) for item in value if normalize_whitespace(str(item)))
    return []


def _confirmed_site_urls(result: dict[str, Any]) -> list[str]:
    confirmed: list[str] = []
    for site in result.get("validated_sites") or []:
        is_confirmed = site.get("decision_status") == "verified" or (
            not site.get("decision_status") and site.get("belongs_to_company")
        )
        if not is_confirmed:
            continue
        url = sanitize_website_url(str(site.get("final_url") or site.get("url") or ""))
        if url and url not in confirmed:
            confirmed.append(url)
    return confirmed


def _normalized_profile_contact_payload(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, Mapping):
        return {}
    return company_profile_to_dict(company_profile_from_dict(payload)).get("contacts") or {}


def _normalized_profile_summary_payload(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, Mapping):
        return {}
    return company_profile_to_dict(company_profile_from_dict(payload)).get("summary") or {}


def _first_nonempty_text(values: Iterable[Any]) -> str:
    for value in values:
        text = normalize_whitespace(str(value or ""))
        if text:
            return text
    return ""


def _geo_signal_source_address(result: dict[str, Any], *, best_address: str = "") -> str:
    selected_best_address = normalize_whitespace(str(best_address or ""))
    if selected_best_address:
        return selected_best_address
    profile_contacts = _normalized_profile_contact_payload(result.get("profile"))
    profile_best_address = normalize_whitespace(str((profile_contacts.get("best_address") or {}).get("value") or ""))
    trusted = _normalized_company_contact_lists(result.get("trusted_contacts") or {})
    merged = _normalized_company_contact_lists(result.get("merged_contacts") or {})
    profile_trusted = _normalized_company_contact_lists(profile_contacts.get("trusted") or {})
    profile_raw = _normalized_company_contact_lists(profile_contacts.get("raw") or {})
    return _first_nonempty_text(
        [
            *(trusted.get("addresses") or []),
            *(merged.get("addresses") or []),
            profile_best_address,
            *(profile_trusted.get("addresses") or []),
            *(profile_raw.get("addresses") or []),
        ]
    )


def _geo_lookup_result_to_profile_payload(value: GeoLookupResult) -> dict[str, Any]:
    return {
        "match_status": value.match_status,
        "source_address": value.source_address,
        "matched_settlement": value.matched_settlement,
        "matched_municipality": value.matched_municipality,
        "matched_region": value.matched_region,
        "geo_bucket": value.geo_bucket or "",
        "geo_weight": value.geo_weight,
        "inside_outer_polygon": value.inside_outer_polygon,
        "inside_inner_polygon": value.inside_inner_polygon,
        "distance_to_moscow_km": value.distance_to_moscow_km,
        "candidate_count": int(value.candidate_count or 0),
        "variant_count": int(value.variant_count or 0),
        "distance_spread_km": value.distance_spread_km,
        "ambiguous_geo_buckets": [str(item) for item in (value.ambiguous_geo_buckets or ()) if str(item)],
    }


def build_geo_signal_payload(result: dict[str, Any], *, best_address: str = "") -> dict[str, Any]:
    source_address = _geo_signal_source_address(result, best_address=best_address)
    return _geo_lookup_result_to_profile_payload(lookup_settlement(source_address))


def _naming_signal_source_name(result: dict[str, Any]) -> str:
    profile_summary = _normalized_profile_summary_payload(result.get("profile"))
    source_results = result.get("sources") or {}
    return _first_nonempty_text(
        [
            result.get("company_name"),
            profile_summary.get("company_name"),
            *(
                source_payload.get("company_name_found")
                for source_payload in source_results.values()
                if isinstance(source_payload, Mapping)
            ),
        ]
    )


def _name_signal_result_to_profile_payload(value: NameSignalResult) -> dict[str, Any]:
    return {
        "signal_status": value.signal_status,
        "source_name": value.source_name,
        "verdict": value.verdict,
        "risk_weight": int(value.risk_weight or 0),
        "matched_markers": list(value.matched_markers or []),
        "reason_codes": list(value.reason_codes or []),
    }


def build_naming_signal_payload(
    result: dict[str, Any],
    *,
    geo_signal: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    signal_geo = geo_signal if isinstance(geo_signal, Mapping) else {}
    geo_match_status = normalize_whitespace(str(signal_geo.get("match_status") or ""))
    source_name = _naming_signal_source_name(result)
    return _name_signal_result_to_profile_payload(
        detect_name_signal(
            source_name,
            geo_match_status=geo_match_status,
        )
    )


def _has_meaningful_profile_summary_inputs(result: dict[str, Any]) -> bool:
    domain_resolution = result.get("domain_resolution") or {}
    return any(
        [
            normalize_whitespace(str(result.get("inn", "") or "")),
            normalize_whitespace(str(result.get("company_name", "") or "")),
            normalize_whitespace(str(result.get("status", "") or "")),
            bool(result.get("lead_cards")),
            bool(result.get("sources")),
            normalize_whitespace(str(domain_resolution.get("status", "") or "")),
        ]
    )


def _has_meaningful_profile_contact_inputs(result: dict[str, Any]) -> bool:
    trusted = result.get("trusted_contacts") or {}
    merged = result.get("merged_contacts") or {}
    return any(bool((trusted.get(field_name) or [])) or bool((merged.get(field_name) or [])) for field_name in COMPANY_CONTACT_FIELDS)


def _has_meaningful_profile_site_inputs(result: dict[str, Any]) -> bool:
    domain_resolution = result.get("domain_resolution") or {}
    return any(
        [
            normalize_whitespace(str(domain_resolution.get("selected_primary_domain", "") or "")),
            bool(result.get("candidate_sites")),
            bool(result.get("validated_sites")),
            bool(result.get("site_probes")),
        ]
    )


def _has_meaningful_profile_signal_inputs(result: dict[str, Any]) -> bool:
    return bool(_geo_signal_source_address(result) or _naming_signal_source_name(result))


def assemble_company_profile_payload(result: dict[str, Any]) -> dict[str, Any]:
    best = choose_best_company_contacts(result)
    trusted = _normalized_company_contact_lists(result.get("trusted_contacts") or {})
    merged = _normalized_company_contact_lists(result.get("merged_contacts") or {})
    domain_resolution = result.get("domain_resolution") or {}
    geo_signal = build_geo_signal_payload(result, best_address=best.get("best_address", ""))
    naming_signal = build_naming_signal_payload(result, geo_signal=geo_signal)
    site_classes = dedupe_preserve_order(
        str(probe.get("site_class", ""))
        for probe in (result.get("site_probes") or [])
        if probe.get("site_class")
    )
    worth_crawling = dedupe_preserve_order(
        str(probe.get("worth_crawling", ""))
        for probe in (result.get("site_probes") or [])
        if probe.get("worth_crawling")
    )
    issues = [
        f"{source_name}: {issue}"
        for source_name, source_payload in (result.get("sources") or {}).items()
        if (issue := source_issue_summary(source_payload))
    ]
    profile = assemble_company_profile(
        summary={
            "inn": result.get("inn", ""),
            "company_name": result.get("company_name", ""),
            "processing_status": result.get("status", ""),
            "domain_resolution_status": domain_resolution.get("status", ""),
            "lead_count": len(result.get("lead_cards") or []),
            "decision_summary": summarize_company_decision(result),
            "issues": issues,
        },
        contacts={
            "trusted": trusted,
            "raw": merged,
            "best_phone": {
                "value": best["best_phone"],
                "sources": _split_profile_sources(best["best_phone_sources"]),
            },
            "best_email": {
                "value": best["best_email"],
                "sources": _split_profile_sources(best["best_email_sources"]),
            },
            "best_address": {
                "value": best["best_address"],
                "sources": _split_profile_sources(best["best_address_sources"]),
            },
        },
        sites={
            "primary_domain": domain_resolution.get("selected_primary_domain", ""),
            "best_site": best["best_site"],
            "best_site_status": best["best_site_status"],
            "best_site_sources": _split_profile_sources(best["best_site_sources"]),
            "candidate_sites": list(result.get("candidate_sites") or []),
            "confirmed_sites": _confirmed_site_urls(result),
            "site_classes": site_classes,
            "worth_crawling": worth_crawling,
        },
        signals={
            "geo": geo_signal,
            "naming": naming_signal,
        },
    )
    return company_profile_to_dict(profile)


def _legacy_payload_from_profile(profile_payload: dict[str, Any]) -> dict[str, Any]:
    summary = profile_payload.get("summary") or {}
    contacts = profile_payload.get("contacts") or {}
    sites = profile_payload.get("sites") or {}
    primary_domain = str(sites.get("primary_domain", "") or "")
    confirmed_sites = [str(item) for item in (sites.get("confirmed_sites") or []) if str(item)]
    legacy_payload: dict[str, Any] = {
        "inn": summary.get("inn", ""),
        "company_name": summary.get("company_name", ""),
        "status": summary.get("processing_status", ""),
        "trusted_contacts": contacts.get("trusted") or {},
        "merged_contacts": contacts.get("raw") or {},
        "candidate_sites": list(sites.get("candidate_sites") or []),
        "validated_sites": [
            {
                "url": url,
                "final_url": url,
                "belongs_to_company": True,
                "decision_status": "verified",
            }
            for url in confirmed_sites
        ],
    }
    if summary.get("domain_resolution_status") or primary_domain:
        legacy_payload["domain_resolution"] = {
            "inn": summary.get("inn", ""),
            "company_name": summary.get("company_name", ""),
            "status": summary.get("domain_resolution_status", ""),
            "selected_primary_domain": primary_domain,
            "selected_primary_status": sites.get("best_site_status", ""),
            "candidates": [],
            "notes": [],
        }
    return legacy_payload


def _profile_merge_value_is_present(value: Any) -> bool:
    if isinstance(value, str):
        return bool(normalize_whitespace(value))
    if isinstance(value, dict):
        return any(_profile_merge_value_is_present(item) for item in value.values())
    if isinstance(value, (list, tuple, set)):
        return any(_profile_merge_value_is_present(item) for item in value)
    return bool(value)


def _merge_profile_section(preferred: Any, fallback: Any) -> Any:
    if isinstance(preferred, dict) or isinstance(fallback, dict):
        preferred_dict = preferred if isinstance(preferred, dict) else {}
        fallback_dict = fallback if isinstance(fallback, dict) else {}
        keys = list(fallback_dict)
        keys.extend(key for key in preferred_dict if key not in fallback_dict)
        merged: dict[str, Any] = {}
        for key in keys:
            preferred_value = preferred_dict.get(key)
            fallback_value = fallback_dict.get(key)
            if isinstance(preferred_value, dict) or isinstance(fallback_value, dict):
                merged[key] = _merge_profile_section(preferred_value, fallback_value)
            elif _profile_merge_value_is_present(preferred_value):
                merged[key] = preferred_value
            else:
                merged[key] = fallback_value
        return merged
    if _profile_merge_value_is_present(preferred):
        return preferred
    return fallback


def _merge_profile_signal_section(preferred: Any, fallback: Any) -> Any:
    preferred_dict = preferred if isinstance(preferred, dict) else {}
    fallback_dict = fallback if isinstance(fallback, dict) else {}
    keys = list(fallback_dict)
    keys.extend(key for key in preferred_dict if key not in fallback_dict)
    merged: dict[str, Any] = {}
    for key in keys:
        preferred_value = preferred_dict.get(key)
        fallback_value = fallback_dict.get(key)
        if key in {"geo", "naming"} and isinstance(preferred_value, dict):
            status_field = "match_status" if key == "geo" else "signal_status"
            preferred_status = preferred_value.get(status_field)
            if _profile_merge_value_is_present(preferred_status):
                merged[key] = preferred_value
                continue
        merged[key] = _merge_profile_section(preferred_value, fallback_value)
    return merged


def company_profile_payload_from_result(result: dict[str, Any]) -> dict[str, Any]:
    profile_payload = result.get("profile")
    normalized_profile = (
        company_profile_to_dict(company_profile_from_dict(profile_payload))
        if isinstance(profile_payload, dict)
        else {}
    )
    if not normalized_profile:
        return assemble_company_profile_payload(result)
    assembled_profile: dict[str, Any] | None = None
    if _has_meaningful_profile_summary_inputs(result):
        assembled_profile = assembled_profile or assemble_company_profile_payload(result)
        normalized_profile["summary"] = _merge_profile_section(
            assembled_profile.get("summary") or {},
            normalized_profile.get("summary") or {},
        )
    if _has_meaningful_profile_contact_inputs(result):
        assembled_profile = assembled_profile or assemble_company_profile_payload(result)
        normalized_profile["contacts"] = _merge_profile_section(
            assembled_profile.get("contacts") or {},
            normalized_profile.get("contacts") or {},
        )
    if _has_meaningful_profile_site_inputs(result):
        assembled_profile = assembled_profile or assemble_company_profile_payload(result)
        normalized_profile["sites"] = _merge_profile_section(
            assembled_profile.get("sites") or {},
            normalized_profile.get("sites") or {},
        )
    normalized_signal_blocks = normalized_profile.get("signals") or {}
    normalized_geo = normalized_signal_blocks.get("geo") or {}
    normalized_naming = normalized_signal_blocks.get("naming") or {}
    if (
        _has_meaningful_profile_signal_inputs(result)
        or not _profile_merge_value_is_present(normalized_geo.get("match_status"))
        or not _profile_merge_value_is_present(normalized_naming.get("signal_status"))
    ):
        assembled_profile = assembled_profile or assemble_company_profile_payload(result)
        normalized_profile["signals"] = _merge_profile_signal_section(
            assembled_profile.get("signals") or {},
            normalized_profile.get("signals") or {},
        )
    return normalized_profile


def normalize_company_result_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = repair_output_value(dict(payload))
    profile_payload = company_profile_payload_from_result(normalized)
    legacy_from_profile = _legacy_payload_from_profile(profile_payload)
    normalized["output_contract_version"] = str(payload.get("output_contract_version") or COMPANY_OUTPUT_CONTRACT_VERSION)
    normalized["profile"] = repair_output_value(profile_payload)
    for key, value in legacy_from_profile.items():
        current_value = normalized.get(key)
        if current_value in (None, "", [], {}):
            normalized[key] = repair_output_value(value)
    return normalized


def flatten_company_result_for_export(result: dict[str, Any]) -> dict[str, Any]:
    profile = company_profile_payload_from_result(result)
    summary = profile.get("summary") or {}
    contacts = profile.get("contacts") or {}
    trusted = (contacts.get("trusted") or {})
    merged = (contacts.get("raw") or {})
    sites = profile.get("sites") or {}
    signals = profile.get("signals") or {}
    geo = signals.get("geo") or {}
    naming = signals.get("naming") or {}
    best_phone = contacts.get("best_phone") or {}
    best_email = contacts.get("best_email") or {}
    best_address = contacts.get("best_address") or {}
    source_statuses = {f"{name}_status": (payload.get("status") or "") for name, payload in (result.get("sources") or {}).items()}
    return {
        "row_index": result.get("row_index", ""),
        "inn": result.get("inn", ""),
        "inn_valid": "yes" if is_valid_russian_inn(str(result.get("inn", ""))) else "no",
        "company_name": summary.get("company_name", result.get("company_name", "")),
        "processing_status": summary.get("processing_status", result.get("status", "")),
        "domain_resolution_status": summary.get(
            "domain_resolution_status",
            (result.get("domain_resolution") or {}).get("status", ""),
        ),
        "primary_domain": sites.get("primary_domain", ""),
        "best_site": sites.get("best_site", ""),
        "best_site_status": sites.get("best_site_status", ""),
        "best_site_sources": ", ".join(sites.get("best_site_sources") or []),
        "best_phone": best_phone.get("value", ""),
        "best_phone_sources": ", ".join(best_phone.get("sources") or []),
        "best_email": best_email.get("value", ""),
        "best_email_sources": ", ".join(best_email.get("sources") or []),
        "best_address": best_address.get("value", ""),
        "best_address_sources": ", ".join(best_address.get("sources") or []),
        "geo_match_status": geo.get("match_status", ""),
        "geo_bucket": geo.get("geo_bucket", ""),
        "geo_weight": geo.get("geo_weight"),
        "geo_distance_to_moscow_km": geo.get("distance_to_moscow_km"),
        "geo_candidate_count": geo.get("candidate_count", 0),
        "geo_variant_count": geo.get("variant_count", 0),
        "geo_distance_spread_km": geo.get("distance_spread_km"),
        "geo_ambiguous_buckets": ",".join(geo.get("ambiguous_geo_buckets") or []),
        "geo_source_address": geo.get("source_address", ""),
        "geo_matched_region": geo.get("matched_region", ""),
        "geo_matched_municipality": geo.get("matched_municipality", ""),
        "geo_matched_settlement": geo.get("matched_settlement", ""),
        "naming_signal_status": naming.get("signal_status", ""),
        "naming_verdict": naming.get("verdict", ""),
        "naming_risk_weight": naming.get("risk_weight", 0),
        "naming_source_name": naming.get("source_name", ""),
        "naming_matched_markers": ",".join(naming.get("matched_markers") or []),
        "naming_reason_codes": ",".join(naming.get("reason_codes") or []),
        "trusted_phone_count": len(trusted.get("phones") or []),
        "trusted_email_count": len(trusted.get("emails") or []),
        "trusted_site_count": len(trusted.get("websites") or []),
        "trusted_address_count": len(trusted.get("addresses") or []),
        "raw_phone_count": len(merged.get("phones") or []),
        "raw_email_count": len(merged.get("emails") or []),
        "raw_site_count": len(merged.get("websites") or []),
        "raw_address_count": len(merged.get("addresses") or []),
        "all_trusted_phones": " | ".join(trusted.get("phones") or []),
        "all_trusted_emails": " | ".join(trusted.get("emails") or []),
        "all_trusted_sites": " | ".join(trusted.get("websites") or []),
        "all_trusted_addresses": " | ".join(trusted.get("addresses") or []),
        "all_raw_phones": " | ".join(merged.get("phones") or []),
        "all_raw_emails": " | ".join(merged.get("emails") or []),
        "all_raw_sites": " | ".join(merged.get("websites") or []),
        "all_raw_addresses": " | ".join(merged.get("addresses") or []),
        "confirmed_site_count": len(sites.get("confirmed_sites") or []),
        "candidate_site_count": len(sites.get("candidate_sites") or []),
        "lead_count": int(summary.get("lead_count", 0) or 0),
        "site_classes": ",".join(sites.get("site_classes") or []),
        "worth_crawling": ",".join(sites.get("worth_crawling") or []),
        "issues": " | ".join(summary.get("issues") or []),
        "decision_summary": summary.get("decision_summary", ""),
        "input_site": result.get("input_site", ""),
        "input_phone": result.get("input_phone", ""),
        "input_comment": result.get("input_comment", ""),
        **source_statuses,
    }


def write_flat_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    ensure_dir(path.parent)
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})
    tmp.replace(path)


def write_flat_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    ensure_dir(path.parent)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    if not fieldnames:
        fieldnames = ["status"]
        rows = [{"status": "empty"}]
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(key, "") for key in fieldnames])
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column[:50]]
        width = max((len(value) for value in values), default=10)
        sheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 48)
    tmp = path.with_name(path.stem + ".tmp" + path.suffix)
    workbook.save(tmp)
    tmp.replace(path)


def _make_source_mask_metrics() -> dict[str, int]:
    return {
        "companies_with_any_mask": 0,
        "companies_with_contact_mask": 0,
        "masked_fields_total": 0,
        "masked_contact_fields_total": 0,
        "companies_blocked": 0,
    }


def _collect_run_source_names(
    ordered_results: list[dict[str, Any]],
    availability_summary: Mapping[str, Any] | None,
) -> list[str]:
    discovered_sources = {
        normalize_whitespace(str(source_name))
        for result in ordered_results
        for source_name in (result.get("sources") or {})
        if normalize_whitespace(str(source_name))
    }
    discovered_sources.update(
        normalize_whitespace(str(source_name))
        for source_name in ((availability_summary or {}).get("sources") or {})
        if normalize_whitespace(str(source_name))
    )
    return sorted(discovered_sources, key=str.casefold)


def render_run_insights_markdown(
    ordered_results: list[dict[str, Any]],
    *,
    summary: dict[str, Any] | None,
    availability_summary: dict[str, Any],
    host_stats: dict[str, Any],
) -> str:
    ordered_results = [repair_output_value(result) for result in ordered_results]
    run_summary = summary or {}
    total_rows = len(ordered_results)
    valid_inn = 0
    invalid_inn = 0
    companies_with_any_contact = 0
    companies_without_any_contact = 0
    companies_with_phone = 0
    companies_with_email = 0
    companies_with_site = 0
    companies_with_confirmed_site = 0
    companies_with_leads = 0
    companies_with_nothing = 0
    domain_status_counts: Counter[str] = Counter()
    site_class_counts: Counter[str] = Counter()
    worth_counts: Counter[str] = Counter()
    run_source_names = _collect_run_source_names(ordered_results, availability_summary)
    source_status_counts: dict[str, Counter[str]] = defaultdict(Counter)
    source_mask_metrics: dict[str, dict[str, int]] = defaultdict(_make_source_mask_metrics)
    invalid_inn_examples: list[str] = []
    no_contact_examples: list[str] = []
    masked_company_lines: list[str] = []

    for result in ordered_results:
        profile = company_profile_payload_from_result(result)
        company_summary = profile.get("summary") or {}
        contacts = profile.get("contacts") or {}
        sites = profile.get("sites") or {}
        trusted = contacts.get("trusted") or {}
        merged = contacts.get("raw") or {}
        inn = str(company_summary.get("inn", result.get("inn", "")))
        company_name = str(company_summary.get("company_name", result.get("company_name", "")))
        if is_valid_russian_inn(inn):
            valid_inn += 1
        else:
            invalid_inn += 1
            invalid_inn_examples.append(f"{inn} — {company_name}")

        any_phone = bool((trusted.get("phones") or []) or (merged.get("phones") or []))
        any_email = bool((trusted.get("emails") or []) or (merged.get("emails") or []))
        any_site = bool((trusted.get("websites") or []) or (merged.get("websites") or []) or sites.get("primary_domain"))
        if any_phone:
            companies_with_phone += 1
        if any_email:
            companies_with_email += 1
        if any_site:
            companies_with_site += 1
        if any_phone or any_email or any_site:
            companies_with_any_contact += 1
        else:
            companies_without_any_contact += 1
            no_contact_examples.append(f"{inn} — {company_name}")
        if not any_phone and not any_email and not any_site:
            companies_with_nothing += 1
        if sites.get("confirmed_sites"):
            companies_with_confirmed_site += 1
        if int(company_summary.get("lead_count", 0) or 0) > 0:
            companies_with_leads += 1

        domain_status_counts[str(company_summary.get("domain_resolution_status", "not_found") or "not_found")] += 1
        for probe in result.get("site_probes") or []:
            site_class_counts[str(probe.get("site_class", "F"))] += 1
            worth_counts[str(probe.get("worth_crawling", "false"))] += 1

        company_mask_parts: list[str] = []
        for source_name, source_payload in (result.get("sources") or {}).items():
            normalized_source_name = normalize_whitespace(str(source_name))
            if not normalized_source_name:
                continue
            source_status_counts[normalized_source_name][str(source_payload.get("status", "unknown"))] += 1
            availability = source_payload.get("availability") or {}
            masked_fields = [
                field_name
                for field_name, payload in availability.items()
                if normalize_source_availability_status(str(payload.get("status", ""))) == "masked"
            ]
            masked_contact_fields = [field_name for field_name in masked_fields if field_name in {"phones", "emails", "websites", "addresses"}]
            blocked_fields = [
                field_name
                for field_name, payload in availability.items()
                if normalize_source_availability_status(str(payload.get("status", ""))) == "blocked"
            ]
            metrics = source_mask_metrics[normalized_source_name]
            if masked_fields:
                metrics["companies_with_any_mask"] += 1
                metrics["masked_fields_total"] += len(masked_fields)
            if masked_contact_fields:
                metrics["companies_with_contact_mask"] += 1
                metrics["masked_contact_fields_total"] += len(masked_contact_fields)
            if blocked_fields or normalize_whitespace(str(source_payload.get("status", ""))) in SOURCE_BLOCKED_RESULT_STATUSES:
                metrics["companies_blocked"] += 1
            if masked_fields:
                company_mask_parts.append(f"{normalized_source_name}: {', '.join(masked_fields)}")
        if company_mask_parts:
            masked_company_lines.append(f"{inn} — {company_name} | " + " | ".join(company_mask_parts))

    lines = [
        "# Run Insights",
        "",
        f"- Updated: `{utc_now_iso()}`",
        f"- Total companies: `{total_rows}`",
        f"- Completed rows: `{run_summary.get('completed_rows', total_rows)}`",
    ]
    status_lines = render_run_status_markdown(run_summary)
    if status_lines:
        lines.extend(status_lines)
    lines.extend(
        [
            f"- Valid INN: `{valid_inn}`",
            f"- Invalid INN: `{invalid_inn}`",
            "",
            "## Coverage",
            f"- Companies with any phone/email/site: `{companies_with_any_contact}`",
            f"- Companies with phone: `{companies_with_phone}`",
            f"- Companies with email: `{companies_with_email}`",
            f"- Companies with site candidate: `{companies_with_site}`",
            f"- Companies with confirmed site: `{companies_with_confirmed_site}`",
            f"- Companies with leads: `{companies_with_leads}`",
            f"- Companies with nothing useful: `{companies_with_nothing}`",
            "",
            "## Domain Resolution",
            f"- verified: `{domain_status_counts.get('verified', 0)}` | candidate: `{domain_status_counts.get('candidate', 0)}` | not_found: `{domain_status_counts.get('not_found', 0)}`",
            "",
            "## Site Probes",
            f"- A: `{site_class_counts.get('A', 0)}` | B: `{site_class_counts.get('B', 0)}` | C: `{site_class_counts.get('C', 0)}` | D: `{site_class_counts.get('D', 0)}` | E: `{site_class_counts.get('E', 0)}` | F: `{site_class_counts.get('F', 0)}`",
            f"- worth_crawling=true: `{worth_counts.get('true', 0)}` | limited: `{worth_counts.get('limited', 0)}` | false: `{worth_counts.get('false', 0)}`",
            "",
            "## Source Statuses",
        ]
    )
    if run_source_names:
        for source_name in run_source_names:
            counter = source_status_counts.get(source_name) or {}
            if not counter:
                lines.append(f"- `{source_name}`: —")
                continue
            rendered = ", ".join(f"{status}={count}" for status, count in sorted(counter.items()))
            lines.append(f"- `{source_name}`: {rendered}")
    else:
        lines.append("- —")

    lines.extend(["", "## Mask Summary By Source"])
    if run_source_names:
        for source_name in run_source_names:
            metrics = source_mask_metrics.get(source_name) or _make_source_mask_metrics()
            contact_mask_companies = metrics.get("companies_with_contact_mask", 0)
            blocked_companies = metrics.get("companies_blocked", 0)
            if contact_mask_companies >= max(10, int(total_rows * 0.15)):
                subscription_signal = "high"
            elif contact_mask_companies > 0:
                subscription_signal = "medium"
            else:
                subscription_signal = "low"
            lines.append(
                f"- `{source_name}` | companies_with_any_mask=`{metrics.get('companies_with_any_mask', 0)}` | companies_with_contact_mask=`{contact_mask_companies}` | masked_fields_total=`{metrics.get('masked_fields_total', 0)}` | masked_contact_fields_total=`{metrics.get('masked_contact_fields_total', 0)}` | blocked_companies=`{blocked_companies}` | subscription_signal=`{subscription_signal}`"
            )
    else:
        lines.append("- —")

    lines.extend(["", "## Field Availability By Source"])
    sources_block = availability_summary.get("sources") or {}
    if run_source_names:
        for source_name in run_source_names:
            source_fields = sources_block.get(source_name) or {}
            lines.append(f"- `{source_name}`")
            for field_name in IMPORTANT_FIELDS:
                counts = source_fields.get(field_name) or {}
                rendered = ", ".join(f"{key}={counts.get(key, 0)}" for key in SOURCE_AVAILABILITY_STATUSES)
                lines.append(f"  - `{field_name}`: {rendered}")
    else:
        lines.append("- —")

    lines.extend(["", "## Host Telemetry"])
    if host_stats:
        for host, payload in sorted(host_stats.items()):
            event_types = payload.get("event_types") or {}
            lines.append(
                f"- `{host}` | ok={event_types.get('request_ok', 0)} | 429={event_types.get('rate_limited', 0)} | bot_gate={event_types.get('bot_gate', 0)} | total_events={payload.get('total_events', 0)}"
            )
    else:
        lines.append("- —")

    if invalid_inn_examples:
        lines.extend(["", "## Invalid INN"])
        lines.extend(render_plain_list_markdown(invalid_inn_examples))
    if no_contact_examples:
        lines.extend(["", "## No Contact Companies"])
        lines.extend(render_plain_list_markdown(no_contact_examples))
    if masked_company_lines:
        lines.extend(["", "## Companies With Masked Data"])
        lines.extend(render_plain_list_markdown(masked_company_lines))

    lines.extend(
        [
            "",
            "## Output Files",
            "- `final_results.xlsx` — финальная читаемая таблица по компаниям",
            "- `final_results.csv` — та же таблица в CSV",
            "- `report.md` — индексный отчет по всем компаниям",
            "- `company_reports/*.md` — подробный разбор по каждой компании",
            "- `events.jsonl` / `host_stats.json` — техлоги прогона и антибан-телеметрия",
        ]
    )
    return "\n".join(lines).strip() + "\n"


def parse_title_and_meta(soup: BeautifulSoup) -> dict[str, str]:
    title = normalize_whitespace(soup.title.get_text(" ", strip=True) if soup.title else "")
    description = ""
    for selector in ['meta[name="description"]', 'meta[property="og:description"]']:
        tag = soup.select_one(selector)
        if tag and tag.get("content"):
            description = normalize_whitespace(tag["content"])
            if description:
                break
    return {"title": title, "description": description}


def label_value(text: str, label: str) -> str:
    pattern = rf"{re.escape(label)}\s*[:\-]?\s*(.+?)(?=(?:\s{{2,}}|[А-ЯA-Z][а-яa-z].+?:|$))"
    match = re.search(pattern, text, flags=re.IGNORECASE)
    if not match:
        return ""
    return normalize_whitespace(match.group(1))


def extract_probable_addresses(text: str) -> list[str]:
    raw_matches = re.findall(r"(?<!\d)\d{6}(?!\d),?[^.\n]{10,180}", text)
    addresses: list[str] = []
    for raw in raw_matches:
        cleaned = sanitize_address_candidate(raw)
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if any(bad in lowered for bad in ("отзывы", "api", "блог", "проверка", "новости", "похожие", "риски")):
            continue
        addresses.append(cleaned)
    return dedupe_preserve_order(addresses)


def sanitize_address_candidate(value: str) -> str:
    cleaned = normalize_whitespace(value)
    if len(cleaned) < 20:
        return ""
    postal_matches = list(re.finditer(r"\b\d{6}\b", cleaned))
    if len(postal_matches) > 1:
        cleaned = cleaned[:postal_matches[1].start()].rstrip(" ,;:-")
    lowered = cleaned.lower()
    for pattern in ADDRESS_STOP_PHRASE_PATTERNS:
        match = pattern.search(cleaned)
        if match and match.start() >= 20:
            cleaned = cleaned[:match.start()].rstrip(" ,;:-")
            lowered = cleaned.lower()
    if len(cleaned) < 20:
        return ""
    if len(cleaned) > 160:
        cleaned = cleaned[:160].rstrip(" ,;:-")
        lowered = cleaned.lower()
    cleaned = ADDRESS_NUMBERED_SINGLE_LETTER_SUFFIX_RE.sub(
        lambda match: f"{match.group(1)}{'-' if match.group(2) else ''}{match.group(3)}",
        cleaned,
    )
    lowered = cleaned.lower()
    if len(re.findall(r"\b(?:ооо|ао|пао|ип)\b", lowered)) >= 2:
        return ""
    if len(re.findall(r"\d{10,}", cleaned)) >= 1:
        return ""
    if not any(marker in lowered for marker in ("ул", "улиц", "пр-кт", "просп", "д.", "дом", "обл", "область", "г.", "город", "район", "корп", "стр", "шоссе", "проезд", "пер", "офис")):
        return ""
    if re.search(r"[, ]+[а-яa-z]$", lowered) and not ADDRESS_VALID_SINGLE_LETTER_SUFFIX_RE.search(lowered):
        return ""
    if lowered.endswith((" пгт", " д", " ул", " ул.", " р-н")):
        return ""
    return cleaned


def normalize_address_values(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = sanitize_address_candidate(value)
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def address_identity_tokens(value: str) -> dict[str, set[str]]:
    cleaned = sanitize_address_candidate(value)
    if not cleaned:
        return {"postals": set(), "tokens": set()}
    lowered = cleaned.lower()
    postals = set(re.findall(r"\b\d{6}\b", lowered))
    tokens = {
        token
        for token in re.findall(r"[a-zа-яё0-9]{3,}", lowered, flags=re.IGNORECASE)
        if len(token) >= 4 and token not in ADDRESS_TOKEN_STOPWORDS and not token.isdigit()
    }
    return {"postals": postals, "tokens": tokens}


@dataclass
class RowInput:
    row_index: int
    inn: str
    company_name: str
    xlsx_site: str = ""
    xlsx_phone: str = ""
    comment: str = ""


@dataclass
class ContactItem:
    value: str
    source_url: str
    kind: str
    masked: bool = False
    note: str = ""


@dataclass
class OkvedEntry:
    code: str
    label: str
    display: str = ""

    def __post_init__(self) -> None:
        self.code = normalize_whitespace(str(self.code or ""))
        self.label = normalize_whitespace(str(self.label or ""))
        self.display = build_okved_display(self.code, self.label)


@dataclass
class SourceResult:
    source: str
    status: str
    search_url: str = ""
    listing_url: str = ""
    entity_url: str = ""
    http_status: int | None = None
    company_name_found: str = ""
    addresses: list[ContactItem] = field(default_factory=list)
    phones: list[ContactItem] = field(default_factory=list)
    emails: list[ContactItem] = field(default_factory=list)
    websites: list[ContactItem] = field(default_factory=list)
    links: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    snippets: list[str] = field(default_factory=list)
    primary_okved: OkvedEntry | None = None
    additional_okveds: list[OkvedEntry] = field(default_factory=list)
    availability: dict[str, dict[str, Any]] = field(default_factory=dict)
    masked_rows: list[str] = field(default_factory=list)


@dataclass
class ListOrgOfflineRow:
    request: str
    result_count: int
    search_count: int
    entity: dict[str, Any] | None = None


@dataclass
class LeadCard:
    company_id: str
    company_name: str
    site_url: str
    title: str
    lead_type: str
    why_relevant: str
    date: str = ""
    deadline: str = ""
    contacts: dict[str, list[str]] = field(default_factory=dict)
    source_urls: list[str] = field(default_factory=list)
    confidence: float = 0.0
    status: str = "new"


@dataclass
class SiteRefreshPlan:
    site_url: str
    cadence: str
    next_due_at: str
    reason: str


@dataclass
class CompanyResult:
    row_index: int
    inn: str
    company_name: str
    input_site: str = ""
    input_phone: str = ""
    input_comment: str = ""
    started_at: str = ""
    finished_at: str = ""
    status: str = ""
    output_contract_version: str = COMPANY_OUTPUT_CONTRACT_VERSION
    sources: dict[str, SourceResult] = field(default_factory=dict)
    merged_contacts: dict[str, list[str]] = field(default_factory=dict)
    trusted_contacts: dict[str, list[str]] = field(default_factory=dict)
    domain_resolution: DomainResolution | None = None
    candidate_sites: list[str] = field(default_factory=list)
    site_probes: list[SiteProbe] = field(default_factory=list)
    route_strategies: list[RouteStrategy] = field(default_factory=list)
    content_records: list[ContentRecord] = field(default_factory=list)
    lead_cards: list[LeadCard] = field(default_factory=list)
    site_refresh_plans: list[SiteRefreshPlan] = field(default_factory=list)
    dossier_ref: dict[str, Any] = field(default_factory=dict)
    validated_sites: list[SiteDecision] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    profile: CompanyProfile = field(default_factory=CompanyProfile)


@dataclass
class HostStatus:
    last_request_at: float = 0.0
    last_attempt_started_at: float = 0.0
    cooldown_until: float = 0.0
    consecutive_429: int = 0
    consecutive_bot: int = 0


@dataclass
class RequestOutcome:
    ok: bool
    status: str
    response: requests.Response | None = None
    error: str = ""
    host: str = ""
    cooldown_seconds: int = 0
    proxy_mode: str = "direct"
    proxy_label: str = ""
    proxy_id: str = ""
    timeout: bool = False
    blocked: bool = False
    elapsed_seconds: float = 0.0


class RateLimitedHttpClient:
    def __init__(
        self,
        logger: logging.Logger,
        progress_store: ProgressStore,
        min_delay_by_host: dict[str, float],
        request_timeout: int,
        cooldown_on_429: int,
        cooldown_on_bot: int,
        proxy_pool: ProxyPool,
        list_org_session_file: Path | None = None,
    ) -> None:
        self.logger = logger
        self.progress_store = progress_store
        self.min_delay_by_host = min_delay_by_host
        self.request_timeout = request_timeout
        self.cooldown_on_429 = cooldown_on_429
        self.cooldown_on_bot = cooldown_on_bot
        self.proxy_pool = proxy_pool
        self.host_state: dict[str, HostStatus] = defaultdict(HostStatus)
        self.host_delay_locks: dict[str, Any] = {}
        self.lock = Lock()
        self.session = requests.Session()
        self.allow_insecure_tls_fallback = os.getenv("ALLOW_INSECURE_TLS_FALLBACK", "1").strip() not in {"0", "false", "False"}
        disable_warnings(InsecureRequestWarning)
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
                ),
                "Accept-Language": "ru,en;q=0.8",
            }
        )
        self.list_org_session_file = list_org_session_file
        self.list_org_cookie = os.getenv("LISTORG_COOKIE", "").strip()
        self.list_org_user_agent = os.getenv("LISTORG_USER_AGENT", "").strip()
        self.list_org_referer = os.getenv("LISTORG_REFERER", "https://www.list-org.com/").strip()
        self._load_list_org_session_profile()

    def _load_list_org_session_profile(self) -> None:
        if not self.list_org_session_file:
            return
        profile = load_session_profile(self.list_org_session_file)
        if not profile:
            return
        profile_cookie = normalize_whitespace(str(profile.get("cookie_header", "")))
        profile_user_agent = normalize_whitespace(str(profile.get("user_agent", "")))
        profile_referer = normalize_whitespace(str(profile.get("referer", "")))
        if profile_cookie and not self.list_org_cookie:
            self.list_org_cookie = profile_cookie
        if profile_user_agent and not self.list_org_user_agent:
            self.list_org_user_agent = profile_user_agent
        if profile_referer and not self.list_org_referer:
            self.list_org_referer = profile_referer
        cookies = profile.get("cookies") or []
        if isinstance(cookies, list):
            for cookie in cookies:
                try:
                    name = str(cookie.get("name", "")).strip()
                    value = str(cookie.get("value", ""))
                    domain = str(cookie.get("domain", "")).strip() or None
                    path = str(cookie.get("path", "/")).strip() or "/"
                    if name:
                        self.session.cookies.set(name, value, domain=domain, path=path)
                except Exception:
                    continue
        self.logger.info("Ð—Ð°Ð³Ñ€ÑƒÐ¶ÐµÐ½ List-Org session profile: %s", self.list_org_session_file)
        self.progress_store.append_event(
            {
                "ts": utc_now_iso(),
                "type": "list_org_session_loaded",
                "session_file": str(self.list_org_session_file),
                "has_cookie_header": bool(self.list_org_cookie),
                "has_user_agent": bool(self.list_org_user_agent),
            }
        )

    def reload_list_org_session_profile(self, session_file: Path | None = None) -> None:
        if session_file is not None:
            self.list_org_session_file = session_file
        self._load_list_org_session_profile()

    def clear_host_cooldown(self, *hosts: str) -> None:
        with self.lock:
            for host in hosts:
                if not host:
                    continue
                state = self.host_state[host]
                state.cooldown_until = 0.0
                state.last_issue = None

    def _host_headers(self, host: str) -> dict[str, str]:
        headers: dict[str, str] = {}
        if host.endswith("list-org.com"):
            if self.list_org_cookie:
                headers["Cookie"] = self.list_org_cookie
            if self.list_org_user_agent:
                headers["User-Agent"] = self.list_org_user_agent
            if self.list_org_referer:
                headers["Referer"] = self.list_org_referer
        return headers

    def _host_delay_lock(self, host: str) -> Any:
        with self.lock:
            host_lock = self.host_delay_locks.get(host)
            if host_lock is None:
                host_lock = Lock()
                self.host_delay_locks[host] = host_lock
            return host_lock

    def _wait_if_needed(self, host: str) -> RequestOutcome | None:
        host_delay_lock = self._host_delay_lock(host)
        with host_delay_lock:
            now = time.time()
            with self.lock:
                state = self.host_state[host]
                if now < state.cooldown_until:
                    remaining = int(state.cooldown_until - now)
                    return RequestOutcome(
                        ok=False,
                        status="cooldown_active",
                        host=host,
                        cooldown_seconds=remaining,
                        error=f"Host {host} is in cooldown for {remaining}s",
                    )
                min_delay = self.min_delay_by_host.get(host, 1.0)
                wait_for = (state.last_request_at + min_delay + random.uniform(0.05, 0.4)) - now
                if wait_for <= 0:
                    state.last_request_at = now
                    return None
            time.sleep(wait_for)
            now = time.time()
            with self.lock:
                state = self.host_state[host]
                if now < state.cooldown_until:
                    remaining = int(state.cooldown_until - now)
                    return RequestOutcome(
                        ok=False,
                        status="cooldown_active",
                        host=host,
                        cooldown_seconds=remaining,
                        error=f"Host {host} is in cooldown for {remaining}s",
                    )
                state.last_request_at = now
        return None

    def _usable_proxy_pool_count(self, *, source: str | None = None) -> int:
        usable_count = getattr(self.proxy_pool, "usable_count", None)
        if callable(usable_count):
            try:
                return max(int(usable_count(source_name=source)), 0)
            except TypeError:
                try:
                    return max(int(usable_count()), 0)
                except Exception:
                    return 0
            except Exception:
                return 0
        entries = getattr(self.proxy_pool, "entries", None)
        if isinstance(entries, list):
            now = time.time()
            usable_entries = 0
            for entry in entries:
                try:
                    cooldown_until = float(getattr(entry, "cooldown_until", 0.0) or 0.0)
                except (TypeError, ValueError):
                    cooldown_until = 0.0
                if cooldown_until <= now:
                    usable_entries += 1
            return usable_entries
        return 0

    def _select_proxy_from_pool(self, *, source: str, host: str | None) -> ProxySelection:
        select_proxy = getattr(self.proxy_pool, "select", None)
        if not callable(select_proxy):
            return ProxySelection()
        try:
            return select_proxy(host, source_name=source)
        except TypeError:
            try:
                return select_proxy(host)
            except Exception:
                return ProxySelection()
        except Exception:
            return ProxySelection()

    def _proxy_provider_diagnostic(
        self,
        *,
        force_refresh: bool = False,
    ) -> Proxy6InventoryDiagnostic:
        provider_diagnostic = getattr(self.proxy_pool, "proxy_provider_diagnostic", None)
        if callable(provider_diagnostic):
            try:
                return provider_diagnostic(force_refresh=force_refresh)
            except TypeError:
                try:
                    return provider_diagnostic()
                except Exception:
                    pass
            except Exception:
                pass
        return diagnose_proxy6_inventory_from_env(force_refresh=force_refresh)

    def _proxy_pool_has_entries(self) -> bool:
        return bool(getattr(self.proxy_pool, "entries", None) or [])

    def _should_check_proxy_provider_for_direct_capable_source(self, source: str) -> bool:
        normalized_source = normalize_whitespace(source)
        if normalized_source not in {"company_site", "route_fetch"}:
            return False
        return self._proxy_pool_has_entries() or bool(os.getenv("PROXY6_API_KEY", "").strip())

    def _append_proxy_provider_guard_event(
        self,
        *,
        source: str,
        host: str,
        url: str,
        since_previous_request: float | None,
        diagnostic: Proxy6InventoryDiagnostic,
    ) -> None:
        self.progress_store.append_event(
            {
                "ts": utc_now_iso(),
                "type": "request_proxy_provider_guardrail",
                "source": source,
                "host": host,
                "url": url,
                "error": diagnostic.operator_message(),
                "request_status": "proxy_provider_guarded_direct_fallback",
                "blocked_by_policy": False,
                "access_state": "direct_fallback",
                "transport_selected": DIRECT_DEFAULT_TRANSPORT,
                "transport_final": DIRECT_DEFAULT_TRANSPORT,
                "proxy_mode": "direct",
                "since_previous_request_seconds": since_previous_request,
                **diagnostic.as_event_fields(),
            }
        )

    def _proxy_provider_diagnostic_for_direct_fallback(
        self,
        *,
        source: str,
        selection: ProxySelection | None,
    ) -> Proxy6InventoryDiagnostic | None:
        if not self._should_check_proxy_provider_for_direct_capable_source(source):
            return None
        if selection is not None and selection.via_proxy and selection.url:
            return None
        diagnostic = self._proxy_provider_diagnostic()
        if diagnostic.provider_status == PROXY_PROVIDER_INVENTORY_EMPTY_OR_EXPIRED:
            return diagnostic
        if diagnostic.provider_status == PROXY_PROVIDER_INVENTORY_HEALTHY:
            return diagnostic
        if diagnostic.provider_status == PROXY_PROVIDER_STATUS_UNKNOWN and self._proxy_pool_has_entries():
            return diagnostic
        return None

    def _proxy_provider_fields_for_proxy_failure(
        self,
        *,
        source: str,
        proxy_selection: ProxySelection | None,
    ) -> dict[str, object]:
        normalized_source = normalize_whitespace(source)
        if normalized_source not in {"checko", "company_site", "route_fetch"}:
            return {}
        if not proxy_selection or not proxy_selection.via_proxy:
            return {}
        if not self._proxy_pool_has_entries() and not os.getenv("PROXY6_API_KEY", "").strip():
            return {}
        diagnostic = self._proxy_provider_diagnostic()
        if diagnostic.provider_status in {
            PROXY_PROVIDER_INVENTORY_EMPTY_OR_EXPIRED,
            PROXY_PROVIDER_INVENTORY_HEALTHY,
            PROXY_PROVIDER_STATUS_UNKNOWN,
        }:
            return diagnostic.as_event_fields()
        return {}

    def _proxy_required_boundary_outcome(
        self,
        *,
        source: str,
        host: str,
        url: str,
        since_previous_request: float | None,
        reason: str,
    ) -> RequestOutcome:
        normalized_reason = normalize_whitespace(reason) or REQUEST_BLOCKED_NO_PROXY_SELECTION_UNAVAILABLE_REASON
        proxy_provider_fields: dict[str, object] = {}
        if normalize_whitespace(source) == "checko":
            proxy_provider_diagnostic = self._proxy_provider_diagnostic()
            proxy_provider_fields = proxy_provider_diagnostic.as_event_fields()
            normalized_reason = f"{normalized_reason}; {proxy_provider_diagnostic.operator_message()}"
        self.progress_store.append_event(
            {
                "ts": utc_now_iso(),
                "type": "request_blocked_by_policy",
                "source": source,
                "host": host,
                "url": url,
                "error": normalized_reason,
                "request_status": REQUEST_STATUS_BLOCKED_NO_PROXY,
                "blocked_by_policy": True,
                "access_state": "proxy_required",
                "transport_selected": PROXY_BOUND_TRANSPORT,
                "transport_final": REQUEST_STATUS_BLOCKED_NO_PROXY,
                "since_previous_request_seconds": since_previous_request,
                **proxy_provider_fields,
            }
        )
        return RequestOutcome(
            ok=False,
            status=REQUEST_STATUS_BLOCKED_NO_PROXY,
            host=host,
            error=normalized_reason,
            proxy_mode="blocked_no_proxy",
            blocked=True,
        )

    def _resolve_request_proxy_selection(
        self,
        *,
        source: str,
        host: str,
        url: str,
        since_previous_request: float | None,
        proxy_selection: ProxySelection | None,
    ) -> tuple[ProxySelection | None, RequestOutcome | None]:
        if not source_requires_proxy_bound_transport(source):
            selected_proxy = proxy_selection or self._select_proxy_from_pool(source=source, host=host)
            diagnostic = self._proxy_provider_diagnostic_for_direct_fallback(
                source=source,
                selection=selected_proxy,
            )
            if diagnostic is not None:
                self._append_proxy_provider_guard_event(
                    source=source,
                    host=host,
                    url=url,
                    since_previous_request=since_previous_request,
                    diagnostic=diagnostic,
                )
                if diagnostic.provider_status == PROXY_PROVIDER_INVENTORY_EMPTY_OR_EXPIRED:
                    return ProxySelection(), None
            return selected_proxy, None

        if proxy_selection is not None:
            if proxy_selection.via_proxy and proxy_selection.url:
                proxy_provider_diagnostic = self._proxy_provider_diagnostic()
                if proxy_provider_diagnostic.provider_status == PROXY_PROVIDER_INVENTORY_EMPTY_OR_EXPIRED:
                    return None, self._proxy_required_boundary_outcome(
                        source=source,
                        host=host,
                        url=url,
                        since_previous_request=since_previous_request,
                        reason=REQUEST_BLOCKED_NO_PROXY_NO_POOL_REASON,
                    )
                return proxy_selection, None
            return None, self._proxy_required_boundary_outcome(
                source=source,
                host=host,
                url=url,
                since_previous_request=since_previous_request,
                reason=REQUEST_BLOCKED_NO_PROXY_SELECTION_UNAVAILABLE_REASON,
            )

        usable_proxy_pool_count = self._usable_proxy_pool_count(source=source)
        if usable_proxy_pool_count <= 0:
            return None, self._proxy_required_boundary_outcome(
                source=source,
                host=host,
                url=url,
                since_previous_request=since_previous_request,
                reason=REQUEST_BLOCKED_NO_PROXY_NO_POOL_REASON,
            )

        selected_proxy = self._select_proxy_from_pool(source=source, host=host)
        if selected_proxy and selected_proxy.via_proxy and selected_proxy.url:
            return selected_proxy, None
        return None, self._proxy_required_boundary_outcome(
            source=source,
            host=host,
            url=url,
            since_previous_request=since_previous_request,
            reason=REQUEST_BLOCKED_NO_PROXY_SELECTION_UNAVAILABLE_REASON,
        )

    def _mark_host_issue(self, host: str, issue: str) -> int:
        with self.lock:
            state = self.host_state[host]
            if issue == "rate_limited":
                state.consecutive_429 += 1
                cooldown = self.cooldown_on_429 * max(1, state.consecutive_429)
                state.cooldown_until = max(state.cooldown_until, time.time() + cooldown)
                return cooldown
            if issue == "bot_gate":
                state.consecutive_bot += 1
                cooldown = self.cooldown_on_bot * max(1, state.consecutive_bot)
                state.cooldown_until = max(state.cooldown_until, time.time() + cooldown)
                return cooldown
            return 0

    def _mark_host_success(self, host: str) -> None:
        with self.lock:
            state = self.host_state[host]
            state.consecutive_429 = 0
            state.consecutive_bot = 0

    def _proxy_context(self, proxy_selection: ProxySelection | None) -> dict[str, str]:
        if not proxy_selection or not proxy_selection.via_proxy:
            return {
                "proxy_mode": "direct",
                "proxy_label": "",
                "proxy_id": "",
                "proxy_label_or_id": "",
            }
        return {
            "proxy_mode": "proxy",
            "proxy_label": proxy_selection.label,
            "proxy_id": proxy_selection.proxy_id,
            "proxy_label_or_id": proxy_selection.proxy_label_or_id,
        }

    def _mark_proxy_bad(self, proxy_selection: ProxySelection | None, *, source: str, reason: str) -> None:
        if not proxy_selection or not proxy_selection.via_proxy:
            return
        try:
            self.proxy_pool.mark_bad(proxy_selection.url, reason=reason, source_name=source)
        except TypeError:
            self.proxy_pool.mark_bad(proxy_selection.url, reason=reason)

    def _mark_proxy_ok(self, proxy_selection: ProxySelection | None, *, source: str) -> None:
        if not proxy_selection or not proxy_selection.via_proxy:
            return
        try:
            self.proxy_pool.mark_ok(proxy_selection.url, source_name=source)
        except TypeError:
            self.proxy_pool.mark_ok(proxy_selection.url)

    def _is_timeout_error(self, exc: BaseException) -> bool:
        if isinstance(exc, requests.exceptions.Timeout):
            return True
        error_text = normalize_whitespace(str(exc)).lower()
        return any(marker in error_text for marker in ("timed out", "timeout", "read timeout", "connect timeout"))

    def _classify_proxy_transient_error(
        self,
        exc: BaseException,
        *,
        proxy_selection: ProxySelection | None,
    ) -> str | None:
        if not proxy_selection or not proxy_selection.via_proxy:
            return None
        detail = normalize_whitespace(str(exc)).lower()
        if isinstance(exc, requests.exceptions.ProxyError):
            if self._is_timeout_error(exc):
                return "proxy_timeout"
            if "tunnel" in detail:
                return "proxy_tunnel_error"
            return "proxy_connection_error"
        explicit_proxy_markers = (
            "proxyerror",
            "proxy error",
            "cannot connect to proxy",
            "unable to connect to proxy",
            "proxy connection",
            "proxy connect",
            "proxy timeout",
            "proxy timed out",
            "proxy tunnel",
            "tunnel connection failed",
        )
        if not any(marker in detail for marker in explicit_proxy_markers):
            return None
        if self._is_timeout_error(exc):
            return "proxy_timeout"
        if "tunnel" in detail:
            return "proxy_tunnel_error"
        return "proxy_connection_error"

    def _proxy_failover_attempt_limit(self, initial_proxy_selection: ProxySelection | None) -> int:
        proxy_urls: set[str] = set()
        if initial_proxy_selection and initial_proxy_selection.via_proxy and initial_proxy_selection.url:
            proxy_urls.add(initial_proxy_selection.url)
        for entry in getattr(self.proxy_pool, "entries", []) or []:
            proxy_url = normalize_whitespace(str(getattr(entry, "url", "") or ""))
            if proxy_url:
                proxy_urls.add(proxy_url)
        return max(len(proxy_urls), 1)

    def _select_alternative_proxy(
        self,
        host: str,
        *,
        source: str,
        attempted_proxy_urls: set[str],
    ) -> ProxySelection | None:
        entry_count = max(len(getattr(self.proxy_pool, "entries", []) or []), 1)
        for _ in range(entry_count):
            candidate = self._select_proxy_from_pool(source=source, host=host)
            if not candidate or not candidate.via_proxy:
                return None
            if candidate.url and candidate.url in attempted_proxy_urls:
                continue
            return candidate
        return None

    def _classify_request_boundary_error(self, exc: BaseException) -> tuple[str, str]:
        detail = normalize_whitespace(str(exc))
        if isinstance(exc, AssertionError):
            if detail:
                return "request_error", f"low-level URL/host assertion before request dispatch: {detail}"
            return "request_error", "low-level URL/host assertion before request dispatch"
        if detail:
            return "invalid_url", f"invalid URL/host: {detail}"
        return "invalid_url", "invalid URL/host"

    def _request_failure_outcome(
        self,
        *,
        source: str,
        host: str,
        url: str,
        error: str,
        since_previous_request: float | None,
        proxy_context: Mapping[str, str],
        proxy_selection: ProxySelection | None,
        status: str = "request_error",
        timeout_error: bool = False,
        proxy_failure_reason: str | None = None,
    ) -> RequestOutcome:
        if proxy_failure_reason:
            self._mark_proxy_bad(proxy_selection, source=source, reason=proxy_failure_reason)
        proxy_provider_fields = (
            self._proxy_provider_fields_for_proxy_failure(source=source, proxy_selection=proxy_selection)
            if proxy_failure_reason
            else {}
        )
        event = {
            "ts": utc_now_iso(),
            "type": "request_error",
            "source": source,
            "host": host,
            "url": url,
            "error": error,
            "timeout": timeout_error,
            "since_previous_request_seconds": since_previous_request,
            **proxy_context,
            **proxy_provider_fields,
        }
        if status != "request_error":
            event["request_status"] = status
        self.progress_store.append_event(event)
        return RequestOutcome(
            ok=False,
            status=status,
            host=host,
            error=error,
            proxy_mode=proxy_context["proxy_mode"],
            proxy_label=proxy_context["proxy_label"],
            proxy_id=proxy_context["proxy_id"],
            timeout=timeout_error,
        )

    def request(
        self,
        url: str,
        *,
        source: str,
        allow_redirects: bool = True,
        timeout: int | None = None,
        proxy_selection: ProxySelection | None = None,
    ) -> RequestOutcome:
        url = normalize_url(url)
        if not url:
            return RequestOutcome(ok=False, status="invalid_url", error="ÐŸÑƒÑÑ‚Ð¾Ð¹ Ð¸Ð»Ð¸ Ð½ÐµÐºÐ¾Ñ€Ñ€ÐµÐºÑ‚Ð½Ñ‹Ð¹ URL")
        host = urlparse(url).netloc.lower()
        now = time.time()
        with self.lock:
            state = self.host_state[host]
            since_previous_request = None
            if state.last_attempt_started_at:
                since_previous_request = round(now - state.last_attempt_started_at, 3)
            state.last_attempt_started_at = now
        waiting = self._wait_if_needed(host)
        if waiting:
            self.progress_store.append_event(
                {
                    "ts": utc_now_iso(),
                    "type": "cooldown_skip",
                    "source": source,
                    "host": host,
                    "url": url,
                    "cooldown_seconds": waiting.cooldown_seconds,
                    "since_previous_request_seconds": since_previous_request,
                }
            )
            return waiting

        current_proxy_selection, proxy_boundary_outcome = self._resolve_request_proxy_selection(
            source=source,
            host=host,
            url=url,
            since_previous_request=since_previous_request,
            proxy_selection=proxy_selection,
        )
        if proxy_boundary_outcome is not None:
            return proxy_boundary_outcome
        proxy_context = self._proxy_context(current_proxy_selection)
        attempted_proxy_urls: set[str] = set()
        max_proxy_attempts = self._proxy_failover_attempt_limit(current_proxy_selection)
        started_at = time.time()
        insecure_tls = False
        attempt_no = 0
        while True:
            attempt_no += 1
            proxy_selection = current_proxy_selection or ProxySelection()
            proxy_context = self._proxy_context(proxy_selection)
            if proxy_selection.via_proxy and proxy_selection.url:
                attempted_proxy_urls.add(proxy_selection.url)
            proxies = proxy_selection.requests_proxies
            headers = self._host_headers(host)
            try:
                response = self.session.get(
                    url,
                    timeout=timeout or self.request_timeout,
                    allow_redirects=allow_redirects,
                    proxies=proxies,
                    headers=headers or None,
                )
            except requests.exceptions.SSLError as exc:
                if source == "company_site" and self.allow_insecure_tls_fallback:
                    try:
                        response = self.session.get(
                            url,
                            timeout=timeout or self.request_timeout,
                            allow_redirects=allow_redirects,
                            proxies=proxies,
                            headers=headers or None,
                            verify=False,
                        )
                        insecure_tls = True
                    except requests.RequestException as inner_exc:
                        timeout_error = self._is_timeout_error(inner_exc)
                        proxy_failover_reason = self._classify_proxy_transient_error(
                            inner_exc,
                            proxy_selection=proxy_selection,
                        )
                        failure_outcome = self._request_failure_outcome(
                            source=source,
                            host=host,
                            url=url,
                            error=str(inner_exc),
                            since_previous_request=since_previous_request,
                            proxy_context=proxy_context,
                            proxy_selection=proxy_selection,
                            status="request_error",
                            timeout_error=timeout_error,
                            proxy_failure_reason=proxy_failover_reason or ("timeout" if timeout_error else "request_error"),
                        )
                        if proxy_failover_reason and attempt_no < max_proxy_attempts:
                            next_proxy_selection = self._select_alternative_proxy(
                                host,
                                source=source,
                                attempted_proxy_urls=attempted_proxy_urls,
                            )
                            if next_proxy_selection is not None:
                                current_proxy_selection = next_proxy_selection
                                insecure_tls = False
                                continue
                        return failure_outcome
                    except (AssertionError, ValueError) as inner_exc:
                        status, error = self._classify_request_boundary_error(inner_exc)
                        return self._request_failure_outcome(
                            source=source,
                            host=host,
                            url=url,
                            error=error,
                            since_previous_request=since_previous_request,
                            proxy_context=proxy_context,
                            proxy_selection=proxy_selection,
                            status=status,
                        )
                else:
                    return self._request_failure_outcome(
                        source=source,
                        host=host,
                        url=url,
                        error=str(exc),
                        since_previous_request=since_previous_request,
                        proxy_context=proxy_context,
                        proxy_selection=proxy_selection,
                        status="request_error",
                    )
            except requests.RequestException as exc:
                timeout_error = self._is_timeout_error(exc)
                proxy_failover_reason = self._classify_proxy_transient_error(
                    exc,
                    proxy_selection=proxy_selection,
                )
                failure_outcome = self._request_failure_outcome(
                    source=source,
                    host=host,
                    url=url,
                    error=str(exc),
                    since_previous_request=since_previous_request,
                    proxy_context=proxy_context,
                    proxy_selection=proxy_selection,
                    status="request_error",
                    timeout_error=timeout_error,
                    proxy_failure_reason=proxy_failover_reason or ("timeout" if timeout_error else "request_error"),
                )
                if proxy_failover_reason and attempt_no < max_proxy_attempts:
                    next_proxy_selection = self._select_alternative_proxy(
                        host,
                        source=source,
                        attempted_proxy_urls=attempted_proxy_urls,
                    )
                    if next_proxy_selection is not None:
                        current_proxy_selection = next_proxy_selection
                        insecure_tls = False
                        continue
                return failure_outcome
            except (AssertionError, ValueError) as exc:
                status, error = self._classify_request_boundary_error(exc)
                return self._request_failure_outcome(
                    source=source,
                    host=host,
                    url=url,
                    error=error,
                    since_previous_request=since_previous_request,
                    proxy_context=proxy_context,
                    proxy_selection=proxy_selection,
                    status=status,
                )
            break

        elapsed = round(time.time() - started_at, 3)
        text_head = response.text[:8000] if "text" in response.headers.get("Content-Type", "") else ""
        if response.status_code == 429:
            cooldown = self._mark_host_issue(host, "rate_limited")
            self._mark_proxy_bad(proxy_selection, source=source, reason="rate_limited")
            self.progress_store.append_event(
                {
                    "ts": utc_now_iso(),
                    "type": "rate_limited",
                    "source": source,
                    "host": host,
                    "url": url,
                    "status_code": response.status_code,
                    "cooldown_seconds": cooldown,
                    "elapsed_seconds": elapsed,
                    "since_previous_request_seconds": since_previous_request,
                    **proxy_context,
                }
            )
            return RequestOutcome(
                ok=False,
                status="rate_limited",
                response=response,
                host=host,
                cooldown_seconds=cooldown,
                error="429 Too Many Requests",
                proxy_mode=proxy_context["proxy_mode"],
                proxy_label=proxy_context["proxy_label"],
                proxy_id=proxy_context["proxy_id"],
                blocked=True,
                elapsed_seconds=elapsed,
            )

        if looks_like_bot_gate(response, text_head):
            cooldown = self._mark_host_issue(host, "bot_gate")
            self._mark_proxy_bad(proxy_selection, source=source, reason="bot_gate")
            self.progress_store.append_event(
                {
                    "ts": utc_now_iso(),
                    "type": "bot_gate",
                    "source": source,
                    "host": host,
                    "url": url,
                    "status_code": response.status_code,
                    "cooldown_seconds": cooldown,
                    "elapsed_seconds": elapsed,
                    "since_previous_request_seconds": since_previous_request,
                    **proxy_context,
                }
            )
            return RequestOutcome(
                ok=False,
                status="bot_gate",
                response=response,
                host=host,
                cooldown_seconds=cooldown,
                error="Bot/captcha gate detected",
                proxy_mode=proxy_context["proxy_mode"],
                proxy_label=proxy_context["proxy_label"],
                proxy_id=proxy_context["proxy_id"],
                blocked=True,
                elapsed_seconds=elapsed,
            )

        if response.status_code >= 400:
            blocked = response.status_code == 403
            if blocked:
                self._mark_proxy_bad(proxy_selection, source=source, reason="http_403")
            self.progress_store.append_event(
                {
                    "ts": utc_now_iso(),
                    "type": "http_error",
                    "source": source,
                    "host": host,
                    "url": url,
                    "status_code": response.status_code,
                    "elapsed_seconds": elapsed,
                    "since_previous_request_seconds": since_previous_request,
                    "blocked": blocked,
                    **proxy_context,
                }
            )
            return RequestOutcome(
                ok=False,
                status=f"http_{response.status_code}",
                response=response,
                host=host,
                error=f"HTTP {response.status_code}",
                proxy_mode=proxy_context["proxy_mode"],
                proxy_label=proxy_context["proxy_label"],
                proxy_id=proxy_context["proxy_id"],
                blocked=blocked,
                elapsed_seconds=elapsed,
            )

        self._mark_host_success(host)
        self._mark_proxy_ok(proxy_selection, source=source)
        self.progress_store.append_event(
            {
                "ts": utc_now_iso(),
                "type": "request_ok_insecure_tls" if insecure_tls else "request_ok",
                "source": source,
                "host": host,
                "url": url,
                "status_code": response.status_code,
                "elapsed_seconds": elapsed,
                "since_previous_request_seconds": since_previous_request,
                **proxy_context,
            }
        )
        return RequestOutcome(
            ok=True,
            status="ok",
            response=response,
            host=host,
            proxy_mode=proxy_context["proxy_mode"],
            proxy_label=proxy_context["proxy_label"],
            proxy_id=proxy_context["proxy_id"],
            elapsed_seconds=elapsed,
        )


def dedupe_contact_items(items: list[ContactItem]) -> list[ContactItem]:
    result: list[ContactItem] = []
    seen: set[tuple[str, bool]] = set()
    for item in items:
        key = (item.value, item.masked)
        if not item.value or key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def normalize_source_availability_field_name(field_name: str) -> str:
    normalized = normalize_whitespace(field_name)
    if normalized not in IMPORTANT_FIELDS:
        raise ValueError(f"Unsupported source availability field: {field_name!r}")
    return normalized


def normalize_source_availability_status(status: str, *, default: str = "unknown") -> str:
    normalized = normalize_whitespace(status)
    if normalized in SOURCE_AVAILABILITY_STATUSES:
        return normalized
    return default


def make_source_availability_counts() -> dict[str, int]:
    return {status: 0 for status in SOURCE_AVAILABILITY_STATUSES}


def build_field_availability_payload(
    status: str,
    *,
    reason: str = "",
    masked_examples: list[str] | None = None,
    open_count: int | None = None,
) -> dict[str, Any]:
    normalized_status = normalize_source_availability_status(status)
    payload: dict[str, Any] = {"status": normalized_status}
    normalized_reason = normalize_whitespace(reason)
    if normalized_reason:
        payload["reason"] = normalized_reason
    normalized_examples = dedupe_preserve_order(
        normalize_whitespace(str(item))
        for item in (masked_examples or [])
        if normalize_whitespace(str(item))
    )
    if normalized_status == "masked" and normalized_examples:
        payload["masked_examples"] = normalized_examples[:5]
    payload["open_count"] = max(int(open_count or 0), 0) if normalized_status == "open" else 0
    return payload


def clear_source_contact_fields(result: SourceResult) -> None:
    for field_name in SHARED_CONTACT_FIELDS:
        setattr(result, field_name, [])


def source_status_forces_blocked_availability(status: str) -> bool:
    return normalize_whitespace(status) in SOURCE_BLOCKED_RESULT_STATUSES


def source_status_forces_absent_contacts(status: str) -> bool:
    return normalize_whitespace(status) in SOURCE_ABSENT_CONTACT_RESULT_STATUSES


def source_requires_proxy_bound_transport(source_name: str) -> bool:
    normalized_source_name = normalize_whitespace(source_name)
    if not normalized_source_name:
        return False
    return DEFAULT_SOURCE_TRANSPORT_POLICY.get(normalized_source_name) == PROXY_BOUND_TRANSPORT


def _resolve_blocked_availability_reason(result: SourceResult, availability: dict[str, dict[str, Any]]) -> str:
    for field_name in IMPORTANT_FIELDS:
        payload = availability.get(field_name) or {}
        field_status = normalize_source_availability_status(str(payload.get("status", "")), default="")
        reason = normalize_whitespace(str(payload.get("reason", "")))
        if field_status == "blocked" and reason:
            return reason
    for candidate in (*result.errors, *result.notes):
        reason = normalize_whitespace(candidate)
        if reason:
            return reason
    return normalize_whitespace(result.status) or "blocked"


def resolve_source_block_reason(result: SourceResult) -> str:
    availability = {
        field_name: dict(payload)
        for field_name, payload in (result.availability or {}).items()
        if field_name in IMPORTANT_FIELDS
    }
    return _resolve_blocked_availability_reason(result, availability)


def should_disable_source_for_run(status: str, *, live_mode: bool = False, offline_mode: bool = False) -> bool:
    if live_mode or offline_mode:
        return False
    normalized_status = normalize_whitespace(status)
    return normalized_status in RUN_DISABLE_ON_BLOCK_STATUSES


def _is_http_error_source_status(status: str) -> bool:
    normalized_status = normalize_whitespace(status)
    if not normalized_status.startswith("http_"):
        return False
    _, _, http_code = normalized_status.partition("_")
    return http_code.isdigit() and int(http_code) >= 400


def source_result_requires_run_fail_fast(
    source_name: str,
    status: str,
    *,
    access_mode: str = "",
) -> bool:
    normalized_source_name = normalize_whitespace(source_name)
    if normalized_source_name not in CANONICAL_REQUIRED_SOURCE_NAMES:
        return False

    normalized_status = normalize_whitespace(status)
    if not normalized_status:
        return True
    if normalized_status in SOURCE_OPERATIONAL_RESULT_STATUSES:
        return False

    normalized_access_mode = normalize_whitespace(access_mode)
    if not normalized_access_mode:
        normalized_access_mode = DEFAULT_SOURCE_TRANSPORT_POLICY.get(normalized_source_name, "")

    if normalized_access_mode == DIRECT_DEFAULT_TRANSPORT and normalized_status in DIRECT_OPERATIONAL_RESULT_STATUSES:
        return False
    if normalized_access_mode == SESSION_BOUND_TRANSPORT and normalized_status == "guest":
        return True
    if normalized_access_mode == OFFLINE_ONLY_TRANSPORT and normalized_status == "not_configured":
        return True
    if normalized_access_mode == PROXY_BOUND_TRANSPORT and normalized_status == "guest":
        return True

    if normalized_status in SOURCE_BLOCKED_RESULT_STATUSES:
        return True
    if normalized_status in {"request_error", "invalid_url", "not_configured"}:
        return True
    if _is_http_error_source_status(normalized_status):
        return True
    return False


def build_required_source_fail_fast_reason(
    source_name: str,
    status: str,
    *,
    access_mode: str = "",
    detail: str = "",
) -> str:
    normalized_source_name = normalize_whitespace(source_name) or "unknown_source"
    normalized_status = normalize_whitespace(status) or "unknown_status"
    normalized_access_mode = normalize_whitespace(access_mode)
    if not normalized_access_mode:
        normalized_access_mode = DEFAULT_SOURCE_TRANSPORT_POLICY.get(normalized_source_name, "") or "unknown_access_mode"
    normalized_detail = normalize_whitespace(detail) or normalized_status
    return (
        f"Required source `{normalized_source_name}` is not operational in `{normalized_access_mode}` access mode: "
        f"status=`{normalized_status}` detail={normalized_detail}"
    )


def _finalize_contact_field_availability(
    result: SourceResult,
    field_name: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    items: list[ContactItem] = getattr(result, field_name)
    explicit_status = normalize_source_availability_status(str(payload.get("status", "")), default="")
    open_items = [item for item in items if normalize_whitespace(item.value) and not item.masked]
    masked_examples = [
        normalize_whitespace(item.value)
        for item in items
        if normalize_whitespace(item.value) and item.masked
    ]
    if explicit_status == "blocked":
        return build_field_availability_payload("blocked", reason=str(payload.get("reason", "")))
    if open_items:
        return build_field_availability_payload("open", open_count=len(open_items))
    if masked_examples or explicit_status == "masked":
        return build_field_availability_payload(
            "masked",
            reason=str(payload.get("reason", "")) or "Поле есть на странице, но закрыто маской или подпиской",
            masked_examples=masked_examples or list(payload.get("masked_examples") or []),
        )
    if explicit_status == "absent":
        return build_field_availability_payload(
            "absent",
            reason=str(payload.get("reason", "")) or "Поле не найдено в публичной карточке",
        )
    return build_field_availability_payload("absent", reason="Поле не найдено в публичной карточке")


def _finalize_non_contact_field_availability(payload: dict[str, Any]) -> dict[str, Any]:
    return build_field_availability_payload(
        normalize_source_availability_status(str(payload.get("status", ""))),
        reason=str(payload.get("reason", "")),
        masked_examples=list(payload.get("masked_examples") or []),
        open_count=payload.get("open_count"),
    )


def set_field_availability(
    result: SourceResult,
    field_name: str,
    status: str,
    *,
    reason: str = "",
    masked_examples: list[str] | None = None,
    open_count: int | None = None,
) -> None:
    result.availability[normalize_source_availability_field_name(field_name)] = build_field_availability_payload(
        status,
        reason=reason,
        masked_examples=masked_examples,
        open_count=open_count,
    )


def mark_source_blocked(result: SourceResult, *, reason: str) -> None:
    for field_name in IMPORTANT_FIELDS:
        set_field_availability(result, field_name, "blocked", reason=reason, open_count=0)


def finalize_source_availability(result: SourceResult) -> None:
    existing_availability = {
        field_name: dict(payload)
        for field_name, payload in (result.availability or {}).items()
        if field_name in IMPORTANT_FIELDS
    }
    result.availability = {}
    source_status = normalize_whitespace(result.status)
    if source_status_forces_blocked_availability(source_status):
        blocked_reason = _resolve_blocked_availability_reason(result, existing_availability)
        for field_name in IMPORTANT_FIELDS:
            set_field_availability(result, field_name, "blocked", reason=blocked_reason, open_count=0)
        return
    if source_status_forces_absent_contacts(source_status):
        absent_reason = (
            "Карточка источника не соответствует запрошенной компании"
            if source_status == "mismatch"
            else "Компания не найдена в источнике"
        )
        for field_name in SHARED_CONTACT_FIELDS:
            set_field_availability(result, field_name, "absent", reason=absent_reason, open_count=0)
        for field_name in NON_CONTACT_AVAILABILITY_FIELDS:
            set_field_availability(result, field_name, "unknown", open_count=0)
        return
    for field_name in IMPORTANT_FIELDS:
        payload = existing_availability.get(field_name) or {}
        if field_name in SHARED_CONTACT_FIELDS:
            result.availability[field_name] = _finalize_contact_field_availability(result, field_name, payload)
        else:
            result.availability[field_name] = _finalize_non_contact_field_availability(payload)


class OpenAIDecider:
    def __init__(
        self,
        logger: logging.Logger,
        progress_store: ProgressStore,
        benchmark_capture: LLMBenchmarkCaptureWriter | None = None,
    ) -> None:
        self.logger = logger
        self.progress_store = progress_store
        self.benchmark_capture = benchmark_capture
        self.api_key = os.getenv("OPENAI_API_KEY", "").strip()
        legacy_model = os.getenv("OPENAI_MODEL", "").strip()
        self.site_decision_model = (
            os.getenv("OPENAI_SITE_DECISION_MODEL", legacy_model or DEFAULT_SITE_DECISION_MODEL).strip()
            or DEFAULT_SITE_DECISION_MODEL
        )
        self.content_review_model = (
            os.getenv("OPENAI_CONTENT_REVIEW_MODEL", legacy_model or DEFAULT_CONTENT_REVIEW_MODEL).strip()
            or DEFAULT_CONTENT_REVIEW_MODEL
        )
        self.content_review_fallback_model = (
            os.getenv("OPENAI_CONTENT_REVIEW_FALLBACK_MODEL", DEFAULT_CONTENT_REVIEW_FALLBACK_MODEL).strip()
            or DEFAULT_CONTENT_REVIEW_FALLBACK_MODEL
        )
        self.model = self.site_decision_model
        self.base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1").rstrip("/")
        self.timeout = int(os.getenv("OPENAI_TIMEOUT_SECONDS", "30"))
        self.reasoning_effort = os.getenv("OPENAI_REASONING_EFFORT", "none").strip() or "none"
        self.max_calls = int(os.getenv("OPENAI_MAX_CALLS_PER_RUN", "600"))
        self.max_calls_per_company = max(0, int(os.getenv("OPENAI_MAX_CALLS_PER_COMPANY", "12") or 12))
        self.max_input_chars = int(os.getenv("OPENAI_MAX_INPUT_CHARS", "4200"))
        self.site_decision_max_output_tokens = int(
            os.getenv("OPENAI_SITE_DECISION_MAX_OUTPUT_TOKENS", str(DEFAULT_SITE_DECISION_MAX_OUTPUT_TOKENS))
        )
        self.calls_made = 0
        self.calls_by_company: defaultdict[str, int] = defaultdict(int)
        self.call_lock = Lock()
        self.cost_lock = Lock()
        self.cache: dict[str, dict[str, Any]] = {}
        self.run_cost_usd: float | None = 0.0
        self.company_cost_usd: dict[str, float | None] = {}

    def enabled(self) -> bool:
        return bool(self.api_key)

    def should_capture_benchmark_stage(self, stage: str) -> bool:
        return bool(self.benchmark_capture and self.benchmark_capture.captures_stage(stage))

    def should_force_benchmark_stage(self, stage: str) -> bool:
        return bool(self.benchmark_capture and self.benchmark_capture.forces_stage(stage))

    def _benchmark_capture_only(self) -> bool:
        return bool(self.benchmark_capture and self.benchmark_capture.capture_only)

    def _can_prepare_capture_only_fixture(self) -> bool:
        return self._benchmark_capture_only()

    def _stage_primary_model(self, stage: str) -> str:
        if stage == "content_review":
            return self.content_review_model
        return self.site_decision_model

    def _company_budget_key(self, row: RowInput) -> str:
        inn = normalize_whitespace(row.inn)
        if inn:
            return inn
        company_name = normalize_whitespace(row.company_name).lower()
        return company_name or "unknown_company"

    def _budget_usage_snapshot(self, row: RowInput) -> dict[str, int]:
        company_key = self._company_budget_key(row)
        with self.call_lock:
            return {
                "company_calls_used": int(self.calls_by_company.get(company_key, 0) or 0),
                "company_calls_cap": self.max_calls_per_company,
                "run_calls_used": self.calls_made,
                "run_calls_cap": self.max_calls,
            }

    def _reserve_call_slot(self, row: RowInput) -> str | None:
        company_key = self._company_budget_key(row)
        with self.call_lock:
            if self.calls_made >= self.max_calls:
                return "run_cap_exhausted"
            if self.calls_by_company[company_key] >= self.max_calls_per_company:
                return "company_cap_exhausted"
            self.calls_made += 1
            self.calls_by_company[company_key] += 1
        return None

    def _llm_usage_snapshot(self, payload: Mapping[str, Any] | None) -> dict[str, int | None]:
        usage = payload.get("usage") if isinstance(payload, Mapping) else None
        if not isinstance(usage, Mapping):
            return {
                "input_tokens": None,
                "output_tokens": None,
                "cached_input_tokens": None,
            }
        input_details = usage.get("input_tokens_details")
        cached_input_tokens = (
            input_details.get("cached_tokens")
            if isinstance(input_details, Mapping)
            else None
        )
        return {
            "input_tokens": usage.get("input_tokens"),
            "output_tokens": usage.get("output_tokens"),
            "cached_input_tokens": cached_input_tokens,
        }

    def _consume_llm_cost(
        self,
        row: RowInput,
        *,
        total_cost_usd: float | None,
        cost_unknown: bool,
    ) -> dict[str, float | None]:
        company_key = self._company_budget_key(row)
        with self.cost_lock:
            company_total = self.company_cost_usd.get(company_key, 0.0)
            run_total = self.run_cost_usd
            if cost_unknown:
                self.company_cost_usd[company_key] = None
                self.run_cost_usd = None
            else:
                delta = round(float(total_cost_usd or 0.0), 8)
                if company_total is not None:
                    company_total = round(float(company_total or 0.0) + delta, 8)
                    self.company_cost_usd[company_key] = company_total
                else:
                    self.company_cost_usd[company_key] = None
                if self.run_cost_usd is not None:
                    self.run_cost_usd = round(float(self.run_cost_usd or 0.0) + delta, 8)
            return {
                "company_cost_usd_cumulative": self.company_cost_usd.get(company_key, 0.0),
                "run_cost_usd_cumulative": self.run_cost_usd,
            }

    def _llm_cost_fields(
        self,
        *,
        row: RowInput,
        stage: str,
        model: str,
        usage_payload: Mapping[str, Any] | None = None,
        input_tokens: Any = None,
        output_tokens: Any = None,
        cached_input_tokens: Any = 0,
    ) -> dict[str, Any]:
        usage_snapshot = self._llm_usage_snapshot(usage_payload)
        resolved_input_tokens = usage_snapshot["input_tokens"] if usage_payload is not None else input_tokens
        resolved_output_tokens = usage_snapshot["output_tokens"] if usage_payload is not None else output_tokens
        resolved_cached_input_tokens = (
            usage_snapshot["cached_input_tokens"]
            if usage_payload is not None
            else cached_input_tokens
        )
        cost = calculate_usage_cost_usd(
            model,
            input_tokens=resolved_input_tokens,
            output_tokens=resolved_output_tokens,
            cached_input_tokens=resolved_cached_input_tokens,
        )
        fields = {
            "stage": stage,
            "model": model,
            "input_tokens": cost.input_tokens,
            "output_tokens": cost.output_tokens,
            "input_cost_usd": cost.input_cost_usd,
            "output_cost_usd": cost.output_cost_usd,
            "total_cost_usd": cost.total_cost_usd,
            **self._consume_llm_cost(row, total_cost_usd=cost.total_cost_usd, cost_unknown=cost.cost_unknown),
        }
        if cost.cost_unknown:
            fields["cost_unknown"] = True
        return fields

    def _append_llm_event(
        self,
        *,
        event_type: str,
        row: RowInput,
        url: str,
        stage: str,
        model: str,
        usage_payload: Mapping[str, Any] | None = None,
        input_tokens: Any = None,
        output_tokens: Any = None,
        cached_input_tokens: Any = 0,
        extra: dict[str, Any] | None = None,
    ) -> None:
        event = {
            "ts": utc_now_iso(),
            "type": event_type,
            "inn": row.inn,
            "url": url,
            **self._llm_cost_fields(
                row=row,
                stage=stage,
                model=model,
                usage_payload=usage_payload,
                input_tokens=input_tokens,
                output_tokens=output_tokens,
                cached_input_tokens=cached_input_tokens,
            ),
        }
        if extra:
            event.update(extra)
        self.progress_store.append_event(event)

    def _append_llm_skip_event(
        self,
        *,
        row: RowInput,
        url: str,
        stage: str,
        reason: str,
        model: str | None = None,
        extra: dict[str, Any] | None = None,
    ) -> None:
        if extra:
            extra = {**extra}
        else:
            extra = {}
        extra.update(
            {
                "reason": reason,
                **self._budget_usage_snapshot(row),
            }
        )
        self._append_llm_event(
            event_type="llm_skip",
            row=row,
            url=url,
            stage=stage,
            model=model or self._stage_primary_model(stage),
            input_tokens=0,
            output_tokens=0,
            extra=extra,
        )

    def _append_benchmark_capture_event(self, *, stage: str, row: RowInput, fixture_hash: str) -> None:
        self.progress_store.append_event(
            {
                "ts": utc_now_iso(),
                "type": "llm_benchmark_capture",
                "stage": stage,
                "inn": row.inn,
                "fixture_hash": fixture_hash,
            }
        )

    def _capture_benchmark_fixture(
        self,
        *,
        stage: str,
        row: RowInput,
        url: str,
        request_body_template: dict[str, Any],
        would_call_in_prod: bool,
        prod_skip_reason: str,
        trust_state: str,
        decision_source_context: dict[str, Any],
        site_url: str | None = None,
        compact_context: dict[str, Any] | None = None,
        benchmark_forced_harvest: bool = False,
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
        benchmark_synthetic_candidate: bool = False,
    ) -> dict[str, Any] | None:
        if not self.should_capture_benchmark_stage(stage) or not self.benchmark_capture:
            return None
        fixture = self.benchmark_capture.append_fixture(
            stage=stage,
            row=row,
            url=url,
            site_url=site_url or url,
            request_body_template=request_body_template,
            would_call_in_prod=would_call_in_prod,
            prod_skip_reason=prod_skip_reason,
            trust_state=trust_state,
            decision_source_context=decision_source_context,
            compact_context=compact_context,
            benchmark_forced_harvest=benchmark_forced_harvest,
            benchmark_capture_path=benchmark_capture_path or f"openai_decider.{stage}.capture",
            synthetic_candidate_used=synthetic_candidate_used,
            forced_harvest_level=forced_harvest_level,
            benchmark_synthetic_candidate=benchmark_synthetic_candidate,
        )
        self._append_benchmark_capture_event(stage=stage, row=row, fixture_hash=str(fixture.get("fixture_hash", "")))
        return fixture

    def capture_site_decision_blocker(
        self,
        *,
        row: RowInput,
        blocker_reason: str,
        site_url: str = "",
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
    ) -> dict[str, Any] | None:
        if not self.should_force_benchmark_stage("site_decision") or not self.benchmark_capture:
            return None
        return self.benchmark_capture.append_blocker(
            stage="site_decision",
            row=row,
            site_url=site_url,
            blocker_reason=blocker_reason,
            would_call_in_prod=False,
            benchmark_capture_path=benchmark_capture_path or "openai_decider.site_decision.blocker",
            synthetic_candidate_used=synthetic_candidate_used,
            forced_harvest_level=forced_harvest_level,
        )

    def capture_content_review_blocker(
        self,
        *,
        row: RowInput,
        blocker_reason: str,
        site_url: str = "",
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
    ) -> dict[str, Any] | None:
        if not self.should_force_benchmark_stage("content_review") or not self.benchmark_capture:
            return None
        return self.benchmark_capture.append_blocker(
            stage="content_review",
            row=row,
            site_url=site_url,
            blocker_reason=blocker_reason,
            would_call_in_prod=False,
            benchmark_capture_path=benchmark_capture_path or "openai_decider.content_review.blocker",
            synthetic_candidate_used=synthetic_candidate_used,
            forced_harvest_level=forced_harvest_level,
        )

    def _record_trust_context(self, record: ContentRecord) -> dict[str, str]:
        trace = record.trace if isinstance(record.trace, dict) else {}
        parser_trace = trace.get("factory_site_parser") if isinstance(trace.get("factory_site_parser"), dict) else {}
        crawl_trace = parser_trace.get("crawl") if isinstance(parser_trace.get("crawl"), dict) else {}
        return {
            "state": normalize_whitespace(str(crawl_trace.get("trust_state", "") or "")).lower() or "unknown",
            "verdict": normalize_whitespace(str(crawl_trace.get("trust_verdict", "") or "")),
            "summary": compact_text(normalize_whitespace(str(crawl_trace.get("trust_summary", "") or "")), 180),
        }

    def _llm_error_diagnostics(
        self,
        payload: Mapping[str, Any] | None,
        *,
        parser_reason: str | None = None,
    ) -> dict[str, Any]:
        output = payload.get("output") if isinstance(payload, Mapping) else None
        content_items: list[Mapping[str, Any]] = []
        if isinstance(output, list):
            for item in output:
                if not isinstance(item, Mapping):
                    continue
                content = item.get("content")
                if not isinstance(content, list):
                    continue
                for entry in content:
                    if isinstance(entry, Mapping):
                        content_items.append(entry)

        response_status = payload.get("status") if isinstance(payload, Mapping) and isinstance(payload.get("status"), str) else None
        incomplete_details = payload.get("incomplete_details") if isinstance(payload, Mapping) else None
        incomplete_reason = (
            incomplete_details.get("reason")
            if isinstance(incomplete_details, Mapping) and isinstance(incomplete_details.get("reason"), str)
            else None
        )
        has_output_text = any(item.get("type") == "output_text" for item in content_items)
        has_parsed = any("parsed" in item for item in content_items)
        content_types = dedupe_preserve_order(
            normalize_whitespace(str(item.get("type") or ""))
            for item in content_items
            if normalize_whitespace(str(item.get("type") or ""))
        )
        return {
            "parser_reason": parser_reason,
            "response_status": response_status,
            "has_output": isinstance(output, list) and bool(output),
            "has_output_text": has_output_text,
            "has_parsed": has_parsed,
            "content_types": content_types,
            "has_refusal": any(item.get("type") == "refusal" for item in content_items),
            "incomplete_reason": incomplete_reason,
        }

    @staticmethod
    def _content_review_attempt_fields(
        *,
        attempt_no: int,
        attempt_kind: str,
        fallback_trigger: str | None,
    ) -> dict[str, Any]:
        return {
            "attempt_no": attempt_no,
            "attempt_kind": attempt_kind,
            "fallback_trigger": fallback_trigger,
        }

    @staticmethod
    def _content_review_should_fallback(parser_reason: str) -> bool:
        return parser_reason in CONTENT_REVIEW_FALLBACK_PARSER_REASONS

    def _update_content_review_trace(
        self,
        record: ContentRecord,
        *,
        status: str,
        reason: str,
        trust_context: dict[str, str],
        attempt_no: int,
        attempt_kind: str,
        fallback_trigger: str | None,
        model: str,
        extra: dict[str, Any] | None = None,
    ) -> None:
        record.trace.setdefault("llm_review", {})
        record.trace["llm_review"].update(
            {
                "status": status,
                "reason": reason,
                "attempt_no": attempt_no,
                "attempt_kind": attempt_kind,
                "fallback_trigger": fallback_trigger,
                "model": model,
                "site_trust_state": trust_context.get("state", "unknown"),
                "site_trust_verdict": trust_context.get("verdict", ""),
                "site_trust_summary": trust_context.get("summary", ""),
            }
        )
        if extra:
            record.trace["llm_review"].update(extra)

    def _annotate_content_review_skip(
        self,
        row: RowInput,
        record: ContentRecord,
        *,
        reason: str,
        trust_context: dict[str, str] | None = None,
        attempt_no: int = 1,
        attempt_kind: str = "primary",
        fallback_trigger: str | None = None,
        model: str | None = None,
    ) -> None:
        trust_context = trust_context or self._record_trust_context(record)
        resolved_model = model or self.content_review_model
        fallback_label = str(record.relevance_label or "unknown")
        fallback_score = round(float(record.relevance_score or 0.0), 3)
        reason_parts: list[str] = []
        if reason == "site_not_trusted":
            reason_parts.append(f"trust_state={trust_context.get('state') or 'unknown'}")
            if trust_context.get("verdict"):
                reason_parts.append(f"trust_verdict={trust_context['verdict']}")
        elif reason == "company_cap_exhausted":
            reason_parts.append(f"company_cap={self.max_calls_per_company}")
        elif reason == "run_cap_exhausted":
            reason_parts.append(f"run_cap={self.max_calls}")
        elif reason == "llm_disabled":
            reason_parts.append("OPENAI_API_KEY missing")

        reason_detail = ", ".join(reason_parts)
        note = f"LLM review skipped: {reason}"
        if reason_detail:
            note += f" ({reason_detail})"
        note += f"; fallback={fallback_label}/{fallback_score}"
        record.notes = dedupe_preserve_order([*record.notes, note])[:8]
        self._update_content_review_trace(
            record,
            status="skipped",
            reason=reason,
            trust_context=trust_context,
            attempt_no=attempt_no,
            attempt_kind=attempt_kind,
            fallback_trigger=fallback_trigger,
            model=resolved_model,
            extra={
                "reason_detail": reason_detail,
                "fallback_label": fallback_label,
                "fallback_score": fallback_score,
            },
        )
        self._append_llm_skip_event(
            row=row,
            url=record.url or record.source_url_or_file,
            stage="content_review",
            reason=reason,
            model=resolved_model,
            extra={
                "site_url": record.site_url,
                "content_fingerprint": record.content_fingerprint,
                "site_trust_state": trust_context.get("state", "unknown"),
                "heuristic_relevance_label": fallback_label,
                "heuristic_relevance_score": fallback_score,
                **self._content_review_attempt_fields(
                    attempt_no=attempt_no,
                    attempt_kind=attempt_kind,
                    fallback_trigger=fallback_trigger,
                ),
            },
        )

    def _site_decision_attempt_fields(
        self,
        *,
        attempt_no: int,
        attempt_kind: str,
        retry_trigger: str | None,
    ) -> dict[str, Any]:
        return {
            "attempt_no": attempt_no,
            "attempt_kind": attempt_kind,
            "retry_trigger": retry_trigger,
        }

    def _site_decision_prompt_max_chars(self, *, rescue: bool) -> int:
        if not rescue or self.max_input_chars <= 1200:
            return self.max_input_chars
        return max(1200, self.max_input_chars // 2)

    def _site_decision_compact_context(
        self,
        compressed_context: dict[str, Any],
        *,
        rescue: bool,
    ) -> dict[str, Any]:
        return compact_site_decision_context(compressed_context, rescue=rescue)

    def _build_site_decision_request_body(
        self,
        row: RowInput,
        site_url: str,
        compressed_context: dict[str, Any],
        *,
        rescue: bool,
        compact_context: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        target_reason_chars = SITE_DECISION_RETRY_REASON_TARGET_CHARS if rescue else SITE_DECISION_REASON_MAX_CHARS
        target_evidence_items = SITE_DECISION_RETRY_EVIDENCE_TARGET_ITEMS if rescue else SITE_DECISION_EVIDENCE_MAX_ITEMS
        target_contradiction_items = (
            SITE_DECISION_RETRY_CONTRADICTION_TARGET_ITEMS if rescue else SITE_DECISION_CONTRADICTION_MAX_ITEMS
        )
        if rescue:
            system_prompt = (
                "Return the same strict site/company decision JSON. "
                "Be extremely terse and conservative. "
                "Use only the strongest identity signals: INN, company name, brand, corporate domain, address, phones, emails. "
                "If evidence is weak or conflicting, set belongs_to_company=false. "
                f"Keep reason under {target_reason_chars} characters. "
                f"Prefer {target_evidence_items} short evidence items and at most {target_contradiction_items} short contradiction. "
                "No prose outside JSON."
            )
            task_text = "Same JSON contract. Shortest valid answer."
        else:
            system_prompt = (
                "Decide whether the candidate site belongs to this exact company. "
                "Be conservative: catalogs, dealers, suppliers, marketplaces, directories, social pages, messengers, and unrelated brands must return belongs_to_company=false. "
                "Use legal-identity signals first: INN, company name, brand, corporate email domain, address, phones, business description, and contacts. "
                "Return strict JSON only. Keep reason under "
                f"{target_reason_chars} characters. "
                f"Keep evidence to at most {target_evidence_items} short items and contradictions to at most "
                f"{target_contradiction_items} short items. "
                "Do not add narrative outside the schema."
            )
            task_text = "Return a compact site/company decision JSON."

        compact_context = compact_context or self._site_decision_compact_context(
            compressed_context,
            rescue=rescue,
        )
        user_payload = {
            "task": task_text,
            "company": {"inn": row.inn, "name": row.company_name},
            "candidate_url": site_url,
            "context": compact_context,
        }
        prompt_json = compact_llm_user_payload(
            user_payload,
            self._site_decision_prompt_max_chars(rescue=rescue),
        )
        return {
            "model": self.site_decision_model,
            "reasoning": {"effort": self.reasoning_effort},
            "max_output_tokens": self.site_decision_max_output_tokens,
            "text": {
                "verbosity": "low",
                "format": {
                    "type": "json_schema",
                    "name": "site_match_decision",
                    "strict": True,
                    "schema": LLM_SITE_DECISION_SCHEMA,
                },
            },
            "input": [
                {
                    "role": "system",
                    "content": [{"type": "input_text", "text": system_prompt}],
                },
                {
                    "role": "user",
                    "content": [{"type": "input_text", "text": prompt_json}],
                },
            ],
        }

    def _default_site_decision_capture_context(self, compressed_context: dict[str, Any]) -> dict[str, Any]:
        heuristics = compressed_context.get("heuristics") if isinstance(compressed_context, Mapping) else None
        if isinstance(heuristics, Mapping):
            return dict(heuristics)
        return {}

    def capture_site_decision_fixture(
        self,
        *,
        row: RowInput,
        site_url: str,
        compressed_context: dict[str, Any],
        trust_state: str,
        would_call_in_prod: bool,
        prod_skip_reason: str,
        decision_source_context: dict[str, Any] | None = None,
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
        benchmark_synthetic_candidate: bool = False,
    ) -> dict[str, Any] | None:
        compact_context = self._site_decision_compact_context(compressed_context, rescue=False)
        body = self._build_site_decision_request_body(
            row,
            site_url,
            compressed_context,
            rescue=False,
            compact_context=compact_context,
        )
        return self._capture_benchmark_fixture(
            stage="site_decision",
            row=row,
            url=site_url,
            site_url=site_url,
            request_body_template=body,
            compact_context=compact_context,
            would_call_in_prod=would_call_in_prod,
            prod_skip_reason=prod_skip_reason,
            trust_state=trust_state,
            decision_source_context=decision_source_context or self._default_site_decision_capture_context(compressed_context),
            benchmark_capture_path=benchmark_capture_path,
            synthetic_candidate_used=synthetic_candidate_used,
            forced_harvest_level=forced_harvest_level,
            benchmark_synthetic_candidate=benchmark_synthetic_candidate,
        )

    def capture_forced_site_decision_fixture(
        self,
        *,
        row: RowInput,
        site_url: str,
        compressed_context: dict[str, Any],
        trust_state: str,
        prod_skip_reason: str,
        decision_source_context: dict[str, Any] | None = None,
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
        benchmark_synthetic_candidate: bool = False,
    ) -> dict[str, Any] | None:
        return self.capture_site_decision_fixture(
            row=row,
            site_url=site_url,
            compressed_context=compressed_context,
            trust_state=trust_state,
            would_call_in_prod=False,
            prod_skip_reason=prod_skip_reason,
            decision_source_context=decision_source_context,
            benchmark_capture_path=benchmark_capture_path,
            synthetic_candidate_used=synthetic_candidate_used,
            forced_harvest_level=forced_harvest_level,
            benchmark_synthetic_candidate=benchmark_synthetic_candidate,
        )

    def _content_review_decision_source_context(
        self,
        record: ContentRecord,
        trust_context: dict[str, str],
    ) -> dict[str, Any]:
        return {
            "heuristic_relevance_label": str(record.relevance_label or ""),
            "heuristic_relevance_score": round(float(record.relevance_score or 0.0), 3),
            "heuristic_reasons": list(record.relevance_reasons[:6]),
            "site_url": record.site_url,
            "content_fingerprint": record.content_fingerprint,
            "site_trust_state": trust_context.get("state", "unknown"),
            "site_trust_verdict": trust_context.get("verdict", ""),
            "site_trust_summary": trust_context.get("summary", ""),
        }

    def _build_content_review_request_body(
        self,
        row: RowInput,
        record: ContentRecord,
        primary_site: str,
        *,
        trust_context: dict[str, str] | None = None,
        model: str | None = None,
    ) -> tuple[dict[str, Any], dict[str, Any], dict[str, str]]:
        trust_context = trust_context or self._record_trust_context(record)
        resolved_model = model or self.content_review_model
        system_prompt = (
            "Ты оцениваешь, стоит ли этот page/document fragment считать коммерчески интересным лидом для industrial sourcing. "
            "Не считай корпоративную новость, просто страницу контактов, вакансию, политику сайта или случайную страницу лидом. "
            "Нужны сигналы торгов, реализации, продажи неликвидов, списанного имущества, отходов, металла, вторсырья, оборудования, "
            "демонтажа или других промышленных активов."
        )
        excerpt_char_limit = max(480, min(DEFAULT_CONTENT_REVIEW_EXCERPT_CHARS, self.max_input_chars // 3))
        compacted_excerpt = build_content_review_excerpt(
            title=record.title,
            cleaned_text=record.cleaned_text,
            max_chars=excerpt_char_limit,
        )
        llm_observability = {
            "excerpt_compacted": compacted_excerpt.compacted,
            "excerpt_chars": compacted_excerpt.final_length,
            "excerpt_source_chars": compacted_excerpt.original_length,
        }
        record.trace.setdefault("llm_review", {})
        record.trace["llm_review"].update(
            {
                "excerpt_compacted": compacted_excerpt.compacted,
                "excerpt_chars": compacted_excerpt.final_length,
                "excerpt_source_chars": compacted_excerpt.original_length,
                "site_trust_state": trust_context.get("state", "unknown"),
                "site_trust_verdict": trust_context.get("verdict", ""),
                "site_trust_summary": trust_context.get("summary", ""),
            }
        )
        user_payload = {
            "company": {"inn": row.inn, "name": row.company_name, "primary_site": primary_site},
            "record": {
                "url": record.url,
                "title": record.title,
                "section_guess": record.section_guess,
                "date": record.date,
                "heuristic_relevance_label": record.relevance_label,
                "heuristic_relevance_score": record.relevance_score,
                "heuristic_reasons": record.relevance_reasons[:6],
                "text_excerpt": compacted_excerpt.text,
            },
        }
        prompt_json = compact_llm_user_payload(
            {"task": "Оцени релевантность content record для lead generation", "context": user_payload},
            self.max_input_chars,
        )
        body = {
            "model": resolved_model,
            "reasoning": {"effort": self.reasoning_effort},
            "max_output_tokens": 220,
            "text": {
                "verbosity": "low",
                "format": {
                    "type": "json_schema",
                    "name": "content_relevance_decision",
                    "strict": True,
                    "schema": LLM_CONTENT_RELEVANCE_SCHEMA,
                },
            },
            "input": [
                {"role": "system", "content": [{"type": "input_text", "text": system_prompt}]},
                {"role": "user", "content": [{"type": "input_text", "text": prompt_json}]},
            ],
        }
        return body, llm_observability, trust_context

    def capture_forced_content_review_fixture(
        self,
        *,
        row: RowInput,
        record: ContentRecord,
        primary_site: str,
        prod_skip_reason: str,
        trust_context: dict[str, str] | None = None,
        benchmark_forced_harvest: bool = False,
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
    ) -> dict[str, Any] | None:
        body, _, resolved_trust_context = self._build_content_review_request_body(
            row,
            record,
            primary_site,
            trust_context=trust_context,
        )
        return self._capture_benchmark_fixture(
            stage="content_review",
            row=row,
            url=record.url or record.source_url_or_file,
            site_url=record.site_url or primary_site or record.url or record.source_url_or_file,
            request_body_template=body,
            would_call_in_prod=False,
            prod_skip_reason=prod_skip_reason,
            trust_state=resolved_trust_context.get("state", "unknown"),
            decision_source_context=self._content_review_decision_source_context(record, resolved_trust_context),
            benchmark_forced_harvest=benchmark_forced_harvest,
            benchmark_capture_path=benchmark_capture_path,
            synthetic_candidate_used=synthetic_candidate_used,
            forced_harvest_level=forced_harvest_level,
        )

    def capture_content_review_benchmark_records(
        self,
        *,
        row: RowInput,
        records: list[ContentRecord],
        primary_site: str,
        default_prod_skip_reason: str = "",
        benchmark_forced_harvest: bool = False,
        benchmark_capture_path: str = "",
        synthetic_candidate_used: bool = False,
        forced_harvest_level: str = "none",
    ) -> int:
        captured = 0
        for record in select_content_review_benchmark_records(records):
            trust_context = self._record_trust_context(record)
            prod_skip_reason = describe_content_review_prod_skip_reason(
                record,
                default_reason=default_prod_skip_reason,
            )
            fixture = self.capture_forced_content_review_fixture(
                row=row,
                record=record,
                primary_site=record.site_url or primary_site,
                prod_skip_reason=prod_skip_reason,
                trust_context=trust_context,
                benchmark_forced_harvest=benchmark_forced_harvest,
                benchmark_capture_path=benchmark_capture_path,
                synthetic_candidate_used=synthetic_candidate_used,
                forced_harvest_level=forced_harvest_level,
            )
            if fixture is not None:
                captured += 1
        return captured

    def decide(
        self,
        row: RowInput,
        site_url: str,
        compressed_context: dict[str, Any],
        *,
        trust_state: str = "",
        decision_source_context: dict[str, Any] | None = None,
    ) -> dict[str, Any] | None:
        if not self.enabled() and not self._can_prepare_capture_only_fixture():
            self._append_llm_skip_event(
                row=row,
                url=site_url,
                stage="site_decision",
                reason="llm_disabled",
                model=self.site_decision_model,
            )
            return None

        cache_key = hashlib.sha1(
            f"{row.inn}|{site_url}|{json.dumps(compressed_context, ensure_ascii=False, sort_keys=True)}".encode("utf-8")
        ).hexdigest()
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        budget_reason = self._reserve_call_slot(row)
        if budget_reason:
            self._append_llm_skip_event(
                row=row,
                url=site_url,
                stage="site_decision",
                reason=budget_reason,
                model=self.site_decision_model,
            )
            return None

        attempt_no = 1
        attempt_kind = "primary"
        retry_trigger: str | None = None
        rescue = False

        while True:
            compact_context = self._site_decision_compact_context(
                compressed_context,
                rescue=rescue,
            )
            body = self._build_site_decision_request_body(
                row,
                site_url,
                compressed_context,
                rescue=rescue,
                compact_context=compact_context,
            )
            attempt_fields = self._site_decision_attempt_fields(
                attempt_no=attempt_no,
                attempt_kind=attempt_kind,
                retry_trigger=retry_trigger,
            )
            self._capture_benchmark_fixture(
                stage="site_decision",
                row=row,
                url=site_url,
                site_url=site_url,
                request_body_template=body,
                compact_context=compact_context,
                would_call_in_prod=True,
                prod_skip_reason="",
                trust_state=trust_state or normalize_whitespace(str((compressed_context.get("heuristics") or {}).get("decision_status", ""))),
                decision_source_context=decision_source_context or self._default_site_decision_capture_context(compressed_context),
            )
            if self._benchmark_capture_only():
                return None
            started_at = time.time()
            try:
                response = requests.post(
                    f"{self.base_url}/responses",
                    headers={
                        "Authorization": f"Bearer {self.api_key}",
                        "Content-Type": "application/json",
                    },
                    json=body,
                    timeout=self.timeout,
                )
                response.raise_for_status()
                payload = response.json()
                parse_result = parse_openai_response(payload)
                if parse_result.data is None:
                    should_retry = attempt_kind == "primary" and parse_result.reason == SITE_DECISION_RETRY_TRIGGER
                    self._append_llm_event(
                        event_type="llm_error",
                        row=row,
                        url=site_url,
                        stage="site_decision",
                        model=self.site_decision_model,
                        usage_payload=payload,
                        extra={
                            "reason": parse_result.reason,
                            "error": parse_result.message,
                            **self._site_decision_attempt_fields(
                                attempt_no=attempt_no,
                                attempt_kind=attempt_kind,
                                retry_trigger=parse_result.reason if should_retry else retry_trigger,
                            ),
                            **self._llm_error_diagnostics(payload, parser_reason=parse_result.reason),
                        },
                    )
                    if should_retry:
                        self.logger.warning(
                            "LLM decide parse failed for %s %s [%s]: %s; retrying once with rescue contract",
                            row.inn,
                            site_url,
                            parse_result.reason,
                            parse_result.message,
                        )
                        retry_trigger = parse_result.reason
                        budget_reason = self._reserve_call_slot(row)
                        if budget_reason:
                            self._append_llm_skip_event(
                                row=row,
                                url=site_url,
                                stage="site_decision",
                                reason=budget_reason,
                                model=self.site_decision_model,
                                extra=self._site_decision_attempt_fields(
                                    attempt_no=2,
                                    attempt_kind="retry",
                                    retry_trigger=retry_trigger,
                                ),
                            )
                            return None
                        attempt_no = 2
                        attempt_kind = "retry"
                        rescue = True
                        continue
                    self.logger.warning(
                        "LLM decide parse failed for %s %s [%s]: %s",
                        row.inn,
                        site_url,
                        parse_result.reason,
                        parse_result.message,
                    )
                    return None
                parsed = dict(parse_result.data)
                parsed["confidence"] = round(float(parsed.get("confidence", 0.0) or 0.0), 3)
                self.cache[cache_key] = parsed
                self._append_llm_event(
                    event_type="llm_decision",
                    row=row,
                    url=site_url,
                    stage="site_decision",
                    model=self.site_decision_model,
                    usage_payload=payload,
                    extra={
                        "elapsed_seconds": round(time.time() - started_at, 3),
                        "parser_reason": parse_result.reason,
                        "parser_source": parse_result.source,
                        "belongs_to_company": parsed.get("belongs_to_company"),
                        "confidence": parsed.get("confidence"),
                        **attempt_fields,
                        **self._budget_usage_snapshot(row),
                    },
                )
                return parsed
            except Exception as exc:
                self._append_llm_event(
                    event_type="llm_error",
                    row=row,
                    url=site_url,
                    stage="site_decision",
                    model=self.site_decision_model,
                    extra={
                        "error": str(exc),
                        **attempt_fields,
                        **self._llm_error_diagnostics(None),
                    },
                )
                self.logger.warning(
                    "LLM decide failed for %s %s [%s #%s]: %s",
                    row.inn,
                    site_url,
                    attempt_kind,
                    attempt_no,
                    exc,
                )
                return None

    def judge_content_record(self, row: RowInput, record: ContentRecord, primary_site: str) -> dict[str, Any] | None:
        if not self.enabled() and not self._can_prepare_capture_only_fixture():
            self._annotate_content_review_skip(
                row,
                record,
                reason="llm_disabled",
                model=self.content_review_model,
            )
            return None
        trust_context = self._record_trust_context(record)
        if trust_context.get("state") != "trusted":
            if self.should_force_benchmark_stage("content_review"):
                self.capture_forced_content_review_fixture(
                    row=row,
                    record=record,
                    primary_site=primary_site,
                    prod_skip_reason="site_not_trusted",
                    trust_context=trust_context,
                )
            self._annotate_content_review_skip(
                row,
                record,
                reason="site_not_trusted",
                trust_context=trust_context,
                model=self.content_review_model,
            )
            return None
        cache_key = hashlib.sha1(
            f"record|{row.inn}|{record.url}|{record.content_fingerprint}|{record.relevance_label}|{record.relevance_score}".encode("utf-8")
        ).hexdigest()
        cached = self.cache.get(cache_key)
        if cached:
            return cached
        budget_reason = self._reserve_call_slot(row)
        if budget_reason:
            self._annotate_content_review_skip(
                row,
                record,
                reason=budget_reason,
                trust_context=trust_context,
                model=self.content_review_model,
            )
            return None

        attempt_no = 1
        attempt_kind = "primary"
        fallback_trigger: str | None = None
        current_model = self.content_review_model

        while True:
            body, llm_observability, trust_context = self._build_content_review_request_body(
                row,
                record,
                primary_site,
                trust_context=trust_context,
                model=current_model,
            )
            attempt_fields = self._content_review_attempt_fields(
                attempt_no=attempt_no,
                attempt_kind=attempt_kind,
                fallback_trigger=fallback_trigger,
            )
            if attempt_no == 1:
                self._capture_benchmark_fixture(
                    stage="content_review",
                    row=row,
                    url=record.url or record.source_url_or_file,
                    site_url=record.site_url or primary_site or record.url or record.source_url_or_file,
                    request_body_template=body,
                    would_call_in_prod=True,
                    prod_skip_reason="",
                    trust_state=trust_context.get("state", "unknown"),
                    decision_source_context=self._content_review_decision_source_context(record, trust_context),
                )
                if self._benchmark_capture_only():
                    return None
            started_at = time.time()
            try:
                response = requests.post(
                    f"{self.base_url}/responses",
                    headers={"Authorization": f"Bearer {self.api_key}", "Content-Type": "application/json"},
                    json=body,
                    timeout=self.timeout,
                )
                response.raise_for_status()
                payload = response.json()
                parse_result = parse_openai_response(payload)
                if parse_result.data is None:
                    should_fallback = (
                        attempt_kind == "primary"
                        and self._content_review_should_fallback(parse_result.reason)
                    )
                    event_fallback_trigger = parse_result.reason if should_fallback else fallback_trigger
                    self._append_llm_event(
                        event_type="llm_content_error",
                        row=row,
                        url=record.url,
                        stage="content_review",
                        model=current_model,
                        usage_payload=payload,
                        extra={
                            "reason": parse_result.reason,
                            "error": parse_result.message,
                            **self._content_review_attempt_fields(
                                attempt_no=attempt_no,
                                attempt_kind=attempt_kind,
                                fallback_trigger=event_fallback_trigger,
                            ),
                            **llm_observability,
                            **self._llm_error_diagnostics(payload, parser_reason=parse_result.reason),
                        },
                    )
                    self._update_content_review_trace(
                        record,
                        status="failed",
                        reason=parse_result.reason,
                        trust_context=trust_context,
                        attempt_no=attempt_no,
                        attempt_kind=attempt_kind,
                        fallback_trigger=event_fallback_trigger,
                        model=current_model,
                        extra={
                            "parser_reason": parse_result.reason,
                            "parser_source": parse_result.source,
                            "error": parse_result.message,
                        },
                    )
                    if should_fallback:
                        self.logger.warning(
                            "LLM content judge parse failed for %s %s [%s]: %s; retrying once with fallback model %s",
                            row.inn,
                            record.url,
                            parse_result.reason,
                            parse_result.message,
                            self.content_review_fallback_model,
                        )
                        budget_reason = self._reserve_call_slot(row)
                        if budget_reason:
                            self._annotate_content_review_skip(
                                row,
                                record,
                                reason=budget_reason,
                                trust_context=trust_context,
                                attempt_no=2,
                                attempt_kind="fallback",
                                fallback_trigger=parse_result.reason,
                                model=self.content_review_fallback_model,
                            )
                            return None
                        attempt_no = 2
                        attempt_kind = "fallback"
                        fallback_trigger = parse_result.reason
                        current_model = self.content_review_fallback_model
                        continue
                    self.logger.warning(
                        "LLM content judge parse failed for %s %s [%s]: %s",
                        row.inn,
                        record.url,
                        parse_result.reason,
                        parse_result.message,
                    )
                    return None
                parsed = dict(parse_result.data)
                parsed["confidence"] = round(float(parsed.get("confidence", 0.0) or 0.0), 3)
                self.cache[cache_key] = parsed
                self._update_content_review_trace(
                    record,
                    status="completed",
                    reason="",
                    trust_context=trust_context,
                    attempt_no=attempt_no,
                    attempt_kind=attempt_kind,
                    fallback_trigger=fallback_trigger,
                    model=current_model,
                    extra={
                        "parser_reason": parse_result.reason,
                        "parser_source": parse_result.source,
                    },
                )
                self._append_llm_event(
                    event_type="llm_content_judgement",
                    row=row,
                    url=record.url,
                    stage="content_review",
                    model=current_model,
                    usage_payload=payload,
                    extra={
                        "elapsed_seconds": round(time.time() - started_at, 3),
                        "parser_reason": parse_result.reason,
                        "parser_source": parse_result.source,
                        "relevance_label": parsed.get("relevance_label"),
                        "confidence": parsed.get("confidence"),
                        **attempt_fields,
                        **llm_observability,
                        **self._budget_usage_snapshot(row),
                    },
                )
                return parsed
            except Exception as exc:
                self._append_llm_event(
                    event_type="llm_content_error",
                    row=row,
                    url=record.url,
                    stage="content_review",
                    model=current_model,
                    extra={
                        "error": str(exc),
                        **attempt_fields,
                        **llm_observability,
                        **self._llm_error_diagnostics(None),
                    },
                )
                self._update_content_review_trace(
                    record,
                    status="failed",
                    reason="request_exception",
                    trust_context=trust_context,
                    attempt_no=attempt_no,
                    attempt_kind=attempt_kind,
                    fallback_trigger=fallback_trigger,
                    model=current_model,
                    extra={"error": str(exc)},
                )
                self.logger.warning(
                    "LLM content judge failed for %s %s [%s #%s]: %s",
                    row.inn,
                    record.url,
                    attempt_kind,
                    attempt_no,
                    exc,
                )
                return None


def extract_openai_text(payload: dict[str, Any]) -> str:
    return _extract_openai_text(payload)


def extract_openai_json(payload: dict[str, Any]) -> dict[str, Any] | None:
    return _extract_openai_json(payload)


def _compact_text_hard_limit(value: str | None, limit: int) -> str:
    compacted = compact_text(value, limit)
    if len(compacted) <= limit:
        return compacted
    return compacted[:limit].rstrip()


def compact_site_decision_context(context_payload: dict[str, Any], *, rescue: bool = False) -> dict[str, Any]:
    context = json.loads(json.dumps(context_payload, ensure_ascii=False))
    title_limit = 72 if rescue else 100
    description_limit = 96 if rescue else 140
    contact_limit = 1 if rescue else 2
    list_item_limit = SITE_DECISION_RETRY_LIST_ITEM_TARGET_CHARS if rescue else 90
    snippet_limit = 72 if rescue else 100
    fetched_pages_limit = 1 if rescue else 2
    excerpt_limit = SITE_DECISION_RETRY_EXCERPT_CHARS if rescue else SITE_DECISION_EXCERPT_CHARS
    reason_list_limit = 1 if rescue else 2
    hard_negative_limit = 2 if rescue else 3
    matched_token_limit = 3 if rescue else 4
    positive_keyword_limit = 3 if rescue else 4
    negative_keyword_limit = 2 if rescue else 3

    candidate_site = dict(context.get("candidate_site") or {})
    if candidate_site:
        context["candidate_site"] = {
            "url": candidate_site.get("url"),
            "final_url": candidate_site.get("final_url"),
            "title": _compact_text_hard_limit(str(candidate_site.get("title", "")), title_limit),
            "description": _compact_text_hard_limit(str(candidate_site.get("description", "")), description_limit),
            "phones": list(candidate_site.get("phones") or [])[:contact_limit],
            "emails": list(candidate_site.get("emails") or [])[:contact_limit],
            "addresses": [
                _compact_text_hard_limit(str(item), list_item_limit) for item in (candidate_site.get("addresses") or [])[:1]
            ],
            "fetched_pages": list(candidate_site.get("fetched_pages") or [])[:fetched_pages_limit],
            "text_excerpt": _compact_text_hard_limit(
                str(candidate_site.get("text_excerpt", "")),
                excerpt_limit,
            ),
        }

    known_contacts: dict[str, list[str]] = {}
    for key, values in (context.get("known_contacts") or {}).items():
        compacted_values = [_compact_text_hard_limit(str(item), list_item_limit) for item in list(values or [])[:contact_limit]]
        compacted_values = [item for item in compacted_values if item]
        if compacted_values:
            known_contacts[key] = compacted_values
    if known_contacts:
        context["known_contacts"] = known_contacts

    compacted_aggregators: dict[str, Any] = {}
    for source_name, source_payload in (context.get("aggregator_profile") or {}).items():
        if not isinstance(source_payload, dict):
            continue
        compacted_aggregators[source_name] = {
            "status": source_payload.get("status"),
            "company_name_found": _compact_text_hard_limit(str(source_payload.get("company_name_found", "")), list_item_limit),
            "websites": list(source_payload.get("websites") or [])[:1],
            "emails": list(source_payload.get("emails") or [])[:1],
            "addresses": [_compact_text_hard_limit(str(item), list_item_limit) for item in (source_payload.get("addresses") or [])[:1]],
            "primary_okved": okved_entry_to_dict(okved_entry_from_dict(source_payload.get("primary_okved"))),
            "additional_okveds": [
                okved_entry_to_dict(item) for item in okved_entries_from_payload(source_payload.get("additional_okveds"))[:1]
            ],
            "snippets": [_compact_text_hard_limit(str(item), snippet_limit) for item in (source_payload.get("snippets") or [])[:1]],
        }
    if compacted_aggregators:
        context["aggregator_profile"] = compacted_aggregators

    heuristics = dict(context.get("heuristics") or {})
    if heuristics:
        context["heuristics"] = {
            "decision_status": heuristics.get("decision_status"),
            "authenticity_score": heuristics.get("authenticity_score"),
            "identity_score": heuristics.get("identity_score"),
            "viability_score": heuristics.get("viability_score"),
            "industrial_score": heuristics.get("industrial_score"),
            "conflict_penalty": heuristics.get("conflict_penalty"),
            "hard_negative_hits": list(heuristics.get("hard_negative_hits") or [])[:hard_negative_limit],
            "matched_name_tokens": list(heuristics.get("matched_name_tokens") or [])[:matched_token_limit],
            "positive_keywords": list(heuristics.get("positive_keywords") or [])[:positive_keyword_limit],
            "negative_keywords": list(heuristics.get("negative_keywords") or [])[:negative_keyword_limit],
            "flags": dict(heuristics.get("flags") or {}),
            "identity_reasons": [
                _compact_text_hard_limit(str(item), list_item_limit)
                for item in (heuristics.get("identity_reasons") or [])[:reason_list_limit]
            ],
            "industrial_reasons": [
                _compact_text_hard_limit(str(item), list_item_limit)
                for item in (heuristics.get("industrial_reasons") or [])[:reason_list_limit]
            ],
        }

    business_goal = _compact_text_hard_limit(str(context.get("business_goal", "")), description_limit)
    if business_goal:
        context["business_goal"] = business_goal

    return context


def compact_llm_user_payload(user_payload: dict[str, Any], max_chars: int) -> str:
    payload = json.loads(json.dumps(user_payload, ensure_ascii=False))
    rendered = json.dumps(payload, ensure_ascii=False, indent=2)
    if len(rendered) <= max_chars:
        return rendered

    context = dict(payload.get("context") or {})
    candidate_site = dict(context.get("candidate_site") or {})
    if candidate_site:
        candidate_site["title"] = compact_text(str(candidate_site.get("title", "")), 140)
        candidate_site["description"] = compact_text(str(candidate_site.get("description", "")), 220)
        candidate_site["text_excerpt"] = compact_text(str(candidate_site.get("text_excerpt", "")), 900)
        candidate_site["fetched_pages"] = (candidate_site.get("fetched_pages") or [])[:2]
        context["candidate_site"] = candidate_site

    known_contacts = {}
    for key, values in (context.get("known_contacts") or {}).items():
        known_contacts[key] = list(values or [])[:2]
    if known_contacts:
        context["known_contacts"] = known_contacts

    compacted_aggregators: dict[str, Any] = {}
    for source_name, source_payload in (context.get("aggregator_profile") or {}).items():
        if not isinstance(source_payload, dict):
            continue
        compacted_aggregators[source_name] = {
            "status": source_payload.get("status"),
            "company_name_found": compact_text(str(source_payload.get("company_name_found", "")), 90),
            "websites": list(source_payload.get("websites") or [])[:1],
            "emails": list(source_payload.get("emails") or [])[:1],
            "addresses": [compact_text(str(item), 90) for item in (source_payload.get("addresses") or [])[:1]],
            "primary_okved": okved_entry_to_dict(okved_entry_from_dict(source_payload.get("primary_okved"))),
            "additional_okveds": [
                okved_entry_to_dict(item)
                for item in okved_entries_from_payload(source_payload.get("additional_okveds"))[:2]
            ],
            "snippets": [compact_text(str(item), 120) for item in (source_payload.get("snippets") or [])[:1]],
        }
    if compacted_aggregators:
        context["aggregator_profile"] = compacted_aggregators

    heuristics = dict(context.get("heuristics") or {})
    if heuristics:
        heuristics["matched_name_tokens"] = list(heuristics.get("matched_name_tokens") or [])[:4]
        heuristics["positive_keywords"] = list(heuristics.get("positive_keywords") or [])[:6]
        heuristics["negative_keywords"] = list(heuristics.get("negative_keywords") or [])[:4]
        heuristics["identity_reasons"] = [compact_text(str(item), 110) for item in (heuristics.get("identity_reasons") or [])[:2]]
        heuristics["industrial_reasons"] = [compact_text(str(item), 110) for item in (heuristics.get("industrial_reasons") or [])[:2]]
        context["heuristics"] = heuristics

    payload["context"] = context
    rendered = json.dumps(payload, ensure_ascii=False, indent=2)
    if len(rendered) <= max_chars:
        return rendered

    minimal_context = {
        "aggregator_profile": {
            source_name: {
                "status": source_payload.get("status"),
                "company_name_found": source_payload.get("company_name_found"),
                "primary_okved": source_payload.get("primary_okved"),
                "additional_okveds": source_payload.get("additional_okveds"),
                "snippets": source_payload.get("snippets"),
            }
            for source_name, source_payload in compacted_aggregators.items()
        },
        "candidate_site": {
            "url": candidate_site.get("url"),
            "final_url": candidate_site.get("final_url"),
            "title": candidate_site.get("title"),
            "description": candidate_site.get("description"),
            "text_excerpt": compact_text(str(candidate_site.get("text_excerpt", "")), 500),
        },
        "heuristics": {
            "identity_score": heuristics.get("identity_score"),
            "industrial_score": heuristics.get("industrial_score"),
            "flags": heuristics.get("flags"),
            "matched_name_tokens": heuristics.get("matched_name_tokens"),
            "positive_keywords": heuristics.get("positive_keywords"),
            "negative_keywords": heuristics.get("negative_keywords"),
        },
        "business_goal": compact_text(str(context.get("business_goal", "")), 180),
    }
    payload["context"] = minimal_context
    rendered = json.dumps(payload, ensure_ascii=False, indent=2)
    if len(rendered) <= max_chars:
        return rendered

    minimal_context["candidate_site"]["text_excerpt"] = compact_text(str(candidate_site.get("text_excerpt", "")), 250)
    payload["context"] = minimal_context
    return json.dumps(payload, ensure_ascii=False, indent=2)


def summarize_source_context(source_results: dict[str, SourceResult]) -> dict[str, Any]:
    summaries: dict[str, Any] = {}
    for source_name, payload in source_results.items():
        snippets = [compact_text(item, 220) for item in payload.snippets[:3] if compact_text(item, 220)]
        summaries[source_name] = {
            "status": payload.status,
            "company_name_found": compact_text(payload.company_name_found, 140),
            "phones": [item.value for item in payload.phones[:3]],
            "emails": [item.value for item in payload.emails[:3]],
            "websites": [item.value for item in payload.websites[:3]],
            "addresses": [compact_text(item.value, 140) for item in payload.addresses[:2]],
            "primary_okved": okved_entry_to_dict(payload.primary_okved),
            "additional_okveds": [okved_entry_to_dict(item) for item in payload.additional_okveds[:3]],
            "snippets": snippets,
            "notes": [compact_text(item, 140) for item in payload.notes[:2]],
        }
    return summaries


def load_rows_from_xlsx(path: Path) -> list[RowInput]:
    workbook = load_workbook(path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_whitespace(str(cell or "")).lower() for cell in header_cells]
    name_to_index = {header: idx for idx, header in enumerate(headers)}

    def idx_for(*names: str) -> int | None:
        for name in names:
            key = normalize_whitespace(name).lower()
            if key in name_to_index:
                return name_to_index[key]
        return None

    inn_idx = idx_for("инн")
    company_idx = idx_for("название")
    site_idx = idx_for("сайт")
    phone_idx = idx_for("телефон")
    comment_idx = idx_for("комментарий", "коммент")

    if inn_idx is None or company_idx is None:
        raise ValueError("Не нашел обязательные колонки 'ИНН' и 'Название'")

    rows: list[RowInput] = []
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        inn = normalize_inn(values[inn_idx] if inn_idx is not None else "")
        company_name = normalize_whitespace(str(values[company_idx] or ""))
        if not inn and not company_name:
            continue
        rows.append(
            RowInput(
                row_index=row_index,
                inn=inn,
                company_name=company_name,
                xlsx_site=normalize_whitespace(str(values[site_idx] or "")) if site_idx is not None else "",
                xlsx_phone=normalize_whitespace(str(values[phone_idx] or "")) if phone_idx is not None else "",
                comment=normalize_whitespace(str(values[comment_idx] or "")) if comment_idx is not None else "",
            )
        )
    return rows


def keyword_found_in_text(text: str, keyword: str) -> bool:
    keyword = keyword.strip().lower()
    if not keyword:
        return False
    if " " in keyword:
        pattern = r"(?<![a-zа-яё0-9])" + re.escape(keyword).replace(r"\ ", r"\s+") + r"(?![a-zа-яё0-9])"
    else:
        pattern = r"(?<![a-zа-яё0-9])" + re.escape(keyword) + r"[a-zа-яё0-9-]*(?![a-zа-яё0-9])"
    return re.search(pattern, text, flags=re.IGNORECASE) is not None


def collect_contact_consensus(source_results: dict[str, SourceResult]) -> dict[str, Counter[str]]:
    counters = {
        "phones": Counter(),
        "emails": Counter(),
        "websites": Counter(),
        "addresses": Counter(),
    }
    for source in source_results.values():
        seen_per_source: dict[str, set[str]] = {key: set() for key in counters}
        for item in source.phones:
            if item.masked:
                continue
            key = normalized_phone_digits(item.value)
            if key:
                seen_per_source["phones"].add(key)
        for item in source.emails:
            if item.masked:
                continue
            key = normalize_whitespace(item.value).lower()
            if key:
                seen_per_source["emails"].add(key)
        for item in source.websites:
            if item.masked:
                continue
            cleaned = sanitize_website_url(item.value)
            if cleaned:
                seen_per_source["websites"].add(guess_registered_domain(urlparse(cleaned).netloc))
        for item in source.addresses:
            key = sanitize_address_candidate(item.value).lower()
            if key:
                seen_per_source["addresses"].add(key)
        for kind, values in seen_per_source.items():
            counters[kind].update(values)
    return counters


def merge_contacts(source_results: dict[str, SourceResult], row: RowInput) -> dict[str, list[str]]:
    merged = {"phones": [], "emails": [], "websites": [], "addresses": []}
    if row.xlsx_phone:
        merged["phones"].append(row.xlsx_phone)
    if row.xlsx_site:
        cleaned_input_site = sanitize_website_url(row.xlsx_site)
        if cleaned_input_site:
            merged["websites"].append(cleaned_input_site)
    for source in source_results.values():
        merged["phones"].extend(item.value for item in source.phones if item.value and not item.masked)
        merged["emails"].extend(item.value for item in source.emails if item.value and not item.masked)
        merged["websites"].extend(sanitize_website_url(item.value) for item in source.websites if item.value and not item.masked)
        merged["addresses"].extend(item.value for item in source.addresses if item.value)
    for email in merged["emails"]:
        website = email_to_candidate_website(email)
        if website:
            merged["websites"].append(website)
    merged["phones"] = normalize_phone_values(merged["phones"])
    merged["emails"] = dedupe_preserve_order(merged["emails"])
    merged["websites"] = dedupe_websites_preserve_order(merged["websites"])
    merged["addresses"] = normalize_address_values(merged["addresses"])
    return merged


def build_analysis_contacts(source_results: dict[str, SourceResult], row: RowInput) -> dict[str, list[str]]:
    contacts = {"phones": [], "emails": [], "websites": [], "addresses": []}
    if row.xlsx_phone:
        contacts["phones"].append(row.xlsx_phone)
    if row.xlsx_site:
        cleaned_input_site = sanitize_website_url(row.xlsx_site)
        if cleaned_input_site:
            contacts["websites"].append(cleaned_input_site)
    for source in source_results.values():
        contacts["phones"].extend(item.value for item in source.phones if item.value and not item.masked)
        contacts["emails"].extend(item.value for item in source.emails if item.value and not item.masked)
        contacts["websites"].extend(sanitize_website_url(item.value) for item in source.websites if item.value and not item.masked)
        contacts["addresses"].extend(item.value for item in source.addresses if item.value)
    contacts["phones"] = normalize_phone_values(contacts["phones"])
    contacts["emails"] = dedupe_preserve_order(contacts["emails"])
    contacts["websites"] = dedupe_websites_preserve_order(contacts["websites"])
    contacts["addresses"] = normalize_address_values(contacts["addresses"])
    return contacts


def build_trusted_contacts(
    row: RowInput,
    source_results: dict[str, SourceResult],
    merged_contacts: dict[str, list[str]],
    validated_sites: list[SiteDecision],
) -> dict[str, list[str]]:
    trusted = {"phones": [], "emails": [], "websites": [], "addresses": []}
    consensus = collect_contact_consensus(source_results)
    if row.xlsx_phone:
        trusted["phones"].append(row.xlsx_phone)
    input_site = sanitize_website_url(row.xlsx_site)
    if input_site:
        trusted["websites"].append(input_site)

    phone_samples: dict[str, str] = {}
    email_samples: dict[str, str] = {}
    address_samples: dict[str, str] = {}
    website_samples: dict[str, str] = {}
    for source in source_results.values():
        for item in source.phones:
            if item.masked:
                continue
            key = normalized_phone_digits(item.value)
            if key and key not in phone_samples:
                phone_samples[key] = normalize_phone_candidate(item.value) or item.value
        for item in source.emails:
            if item.masked:
                continue
            key = normalize_whitespace(item.value).lower()
            if key and key not in email_samples:
                email_samples[key] = item.value
        for item in source.addresses:
            cleaned_address = sanitize_address_candidate(item.value)
            key = cleaned_address.lower()
            if cleaned_address and key not in address_samples:
                address_samples[key] = cleaned_address
        for item in source.websites:
            if item.masked:
                continue
            cleaned = sanitize_website_url(item.value)
            if not cleaned:
                continue
            key = guess_registered_domain(urlparse(cleaned).netloc)
            if key and key not in website_samples:
                website_samples[key] = cleaned

    for key, count in consensus["phones"].items():
        if count >= 2 and key in phone_samples:
            trusted["phones"].append(phone_samples[key])
    for key, count in consensus["emails"].items():
        if count >= 2 and key in email_samples:
            trusted["emails"].append(email_samples[key])
    for key, count in consensus["addresses"].items():
        if count >= 2 and key in address_samples:
            trusted["addresses"].append(address_samples[key])
    for key, count in consensus["websites"].items():
        if count >= 2 and key in website_samples:
            trusted["websites"].append(website_samples[key])

    ranked_sites = sorted(
        validated_sites,
        key=lambda item: (
            SITE_AUTH_STATUS_RANK.get(item.decision_status or ("verified" if item.belongs_to_company else "rejected"), 9),
            -float(item.authenticity_score or 0.0),
            -(item.identity_score or 0.0),
        ),
    )
    for site in ranked_sites:
        if site.decision_status not in {"verified"} and not site.belongs_to_company:
            continue
        confirmed_url = sanitize_website_url(site.final_url or site.url)
        if confirmed_url:
            trusted["websites"].append(confirmed_url)
        trusted["phones"].extend(site.extracted_phones)
        trusted["emails"].extend(site.extracted_emails)
        trusted["addresses"].extend(site.extracted_addresses)

    trusted["phones"] = normalize_phone_values(trusted["phones"])
    trusted["emails"] = dedupe_preserve_order(email for email in trusted["emails"] if normalize_whitespace(email))
    trusted["websites"] = dedupe_websites_preserve_order(trusted["websites"])
    trusted["addresses"] = normalize_address_values(address for address in trusted["addresses"] if normalize_whitespace(address))

    if not trusted["emails"]:
        trusted["emails"] = dedupe_preserve_order(
            email
            for email in merged_contacts.get("emails", [])
            if guess_registered_domain(email.split("@", 1)[-1].lower()) in {
                guess_registered_domain(urlparse(url).netloc) for url in trusted["websites"]
            }
        )

    return trusted


def build_lead_cards(
    row: RowInput,
    domain_resolution: DomainResolution | None,
    trusted_contacts: dict[str, list[str]],
    merged_contacts: dict[str, list[str]],
    content_records: list[ContentRecord],
) -> list[LeadCard]:
    leads: list[LeadCard] = []
    primary_site = (domain_resolution.selected_primary_domain if domain_resolution else "") or ""
    best_contacts = trusted_contacts if any(trusted_contacts.get(key) for key in ("phones", "emails", "websites")) else merged_contacts
    seen_keys: set[str] = set()
    for record in content_records:
        if record.relevance_label not in {"maybe_relevant", "likely_relevant"}:
            continue
        if record.section_guess in {"about", "homepage", "news", "contacts"} and record.relevance_label != "likely_relevant":
            continue
        lead_type = infer_lead_type_from_record(record)
        dedupe_key = f"{lead_type}|{record.date}|{compact_text(record.title.lower(), 80)}"
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)
        leads.append(
            LeadCard(
                company_id=row.inn,
                company_name=row.company_name,
                site_url=primary_site,
                title=record.title or record.url,
                lead_type=lead_type,
                why_relevant="; ".join(record.relevance_reasons[:4]) or "rule-based relevance hit",
                date=record.date,
                contacts={
                    "phones": list(best_contacts.get("phones", [])[:3]),
                    "emails": list(best_contacts.get("emails", [])[:3]),
                    "websites": list(best_contacts.get("websites", [])[:2]),
                },
                source_urls=[record.url],
                confidence=record.relevance_score,
                status="new",
            )
        )
    return leads


def build_site_refresh_plans(candidate_sites: list[str], site_probes: list[SiteProbe]) -> list[SiteRefreshPlan]:
    plans: list[SiteRefreshPlan] = []
    by_url = {probe.url: probe for probe in site_probes}
    now = datetime.now(tz=UTC).replace(microsecond=0)
    for site_url in candidate_sites:
        probe = by_url.get(site_url)
        if not probe:
            continue
        if probe.site_class in {"A", "B", "C"} and probe.worth_crawling == "true":
            cadence = "daily"
            due = now + timedelta(days=1)
            reason = "живой и потенциально богатый сайт"
        elif probe.site_class in {"D", "E"} or probe.worth_crawling == "limited":
            cadence = "weekly"
            due = now + timedelta(days=7)
            reason = "сайт требует осторожного или частичного обхода"
        else:
            cadence = "monthly"
            due = now + timedelta(days=30)
            reason = "бедный, мертвый или нерентабельный сайт"
        plans.append(SiteRefreshPlan(site_url=site_url, cadence=cadence, next_due_at=due.isoformat(), reason=reason))
    return plans


def refresh_company_result_profile(result: CompanyResult) -> CompanyProfile:
    profile_payload = company_profile_payload_from_result(asdict(result))
    profile = company_profile_from_dict(profile_payload)
    result.output_contract_version = COMPANY_OUTPUT_CONTRACT_VERSION
    result.profile = profile
    return profile


def serialize_company_result(result: CompanyResult) -> dict[str, Any]:
    refresh_company_result_profile(result)
    payload = repair_output_value(asdict(result))
    payload["output_contract_version"] = result.output_contract_version
    payload["profile"] = repair_output_value(company_profile_to_dict(result.profile))
    return payload


def build_company_result(row: RowInput) -> CompanyResult:
    result = CompanyResult(
        row_index=row.row_index,
        inn=row.inn,
        company_name=row.company_name,
        input_site=row.xlsx_site,
        input_phone=row.xlsx_phone,
        input_comment=row.comment,
        started_at=utc_now_iso(),
        status="running",
    )
    refresh_company_result_profile(result)
    return result


def contact_item_from_dict(payload: dict[str, Any]) -> ContactItem:
    return ContactItem(
        value=payload.get("value", ""),
        source_url=payload.get("source_url", ""),
        kind=payload.get("kind", ""),
        masked=bool(payload.get("masked", False)),
        note=payload.get("note", ""),
    )


def build_okved_display(code: str, label: str) -> str:
    code_clean = normalize_whitespace(str(code or ""))
    label_clean = normalize_whitespace(str(label or ""))
    if code_clean and label_clean:
        return f"{label_clean} ({code_clean})"
    return code_clean or label_clean


def okved_entry_from_dict(payload: Any) -> OkvedEntry | None:
    if isinstance(payload, OkvedEntry):
        return OkvedEntry(code=payload.code, label=payload.label, display=payload.display)
    if not isinstance(payload, dict):
        return None
    code = normalize_whitespace(str(payload.get("code", "") or ""))
    label = normalize_whitespace(str(payload.get("label", "") or ""))
    display = normalize_whitespace(str(payload.get("display", "") or ""))
    if not code and not label and not display:
        return None
    if not code and not label:
        return None
    return OkvedEntry(code=code, label=label, display=display)


def okved_entries_from_payload(payload: Any) -> list[OkvedEntry]:
    if isinstance(payload, OkvedEntry):
        return [OkvedEntry(code=payload.code, label=payload.label, display=payload.display)]
    if not isinstance(payload, (list, tuple)):
        return []
    items: list[OkvedEntry] = []
    for item in payload:
        entry = okved_entry_from_dict(item)
        if entry:
            items.append(entry)
    return items


def okved_entry_to_dict(entry: OkvedEntry | None) -> dict[str, str] | None:
    if not entry:
        return None
    return {
        "code": entry.code,
        "label": entry.label,
        "display": entry.display,
    }


def okved_entry_display(payload: Any) -> str:
    entry = okved_entry_from_dict(payload)
    return entry.display if entry else ""


def okved_entries_display(payload: Any, *, limit: int | None = None, max_len: int = 160) -> list[str]:
    entries = okved_entries_from_payload(payload)
    if limit is not None:
        entries = entries[:limit]
    displays: list[str] = []
    for entry in entries:
        display = compact_text(entry.display, max_len)
        if display:
            displays.append(display)
    return displays


def source_result_from_dict(payload: dict[str, Any]) -> SourceResult:
    return SourceResult(
        source=payload.get("source", ""),
        status=payload.get("status", ""),
        search_url=payload.get("search_url", ""),
        listing_url=payload.get("listing_url", ""),
        entity_url=payload.get("entity_url", ""),
        http_status=payload.get("http_status"),
        company_name_found=payload.get("company_name_found", ""),
        addresses=[contact_item_from_dict(item) for item in payload.get("addresses", [])],
        phones=[contact_item_from_dict(item) for item in payload.get("phones", [])],
        emails=[contact_item_from_dict(item) for item in payload.get("emails", [])],
        websites=[contact_item_from_dict(item) for item in payload.get("websites", [])],
        links=list(payload.get("links", [])),
        notes=list(payload.get("notes", [])),
        errors=list(payload.get("errors", [])),
        snippets=list(payload.get("snippets", [])),
        primary_okved=okved_entry_from_dict(payload.get("primary_okved")),
        additional_okveds=okved_entries_from_payload(payload.get("additional_okveds")),
        availability=dict(payload.get("availability", {})),
        masked_rows=list(payload.get("masked_rows", [])),
    )


def site_decision_from_dict(payload: dict[str, Any]) -> SiteDecision:
    return SiteDecision(
        url=payload.get("url", ""),
        final_url=payload.get("final_url", ""),
        status=payload.get("status", ""),
        identity_score=float(payload.get("identity_score", 0.0) or 0.0),
        viability_score=float(payload.get("viability_score", 0.0) or 0.0),
        industrial_score=float(payload.get("industrial_score", 0.0) or 0.0),
        authenticity_score=float(payload.get("authenticity_score", 0.0) or 0.0),
        conflict_penalty=float(payload.get("conflict_penalty", 0.0) or 0.0),
        belongs_to_company=bool(payload.get("belongs_to_company", False)),
        decision_status=payload.get("decision_status") or ("verified" if payload.get("belongs_to_company") else "rejected"),
        industrial_relevance=payload.get("industrial_relevance", "unknown"),
        decision_source=payload.get("decision_source", "heuristics"),
        reasons=list(payload.get("reasons", [])),
        evidence=list(payload.get("evidence", [])),
        hard_negative_hits=list(payload.get("hard_negative_hits", [])),
        fetched_pages=list(payload.get("fetched_pages", [])),
        title=payload.get("title", ""),
        description=payload.get("description", ""),
        extracted_phones=list(payload.get("extracted_phones", [])),
        extracted_emails=list(payload.get("extracted_emails", [])),
        extracted_addresses=list(payload.get("extracted_addresses", [])),
        matched_name_tokens=list(payload.get("matched_name_tokens", [])),
        matched_keywords=list(payload.get("matched_keywords", [])),
        negative_keywords=list(payload.get("negative_keywords", [])),
        llm_result=payload.get("llm_result"),
        errors=list(payload.get("errors", [])),
    )


def domain_candidate_from_dict(payload: dict[str, Any]) -> DomainCandidate:
    return DomainCandidate(
        url=payload.get("url", ""),
        domain=payload.get("domain", ""),
        source=payload.get("source", ""),
        confidence=float(payload.get("confidence", 0.0) or 0.0),
        status=payload.get("status", "candidate"),
        evidence=list(payload.get("evidence", [])),
    )


def domain_resolution_from_dict(payload: dict[str, Any]) -> DomainResolution:
    return DomainResolution(
        inn=payload.get("inn", ""),
        company_name=payload.get("company_name", ""),
        status=payload.get("status", "not_found"),
        selected_primary_domain=payload.get("selected_primary_domain", ""),
        selected_primary_status=payload.get("selected_primary_status", ""),
        candidates=[domain_candidate_from_dict(item) for item in payload.get("candidates", [])],
        notes=list(payload.get("notes", [])),
    )


def lead_card_from_dict(payload: dict[str, Any]) -> LeadCard:
    return LeadCard(
        company_id=payload.get("company_id", ""),
        company_name=payload.get("company_name", ""),
        site_url=payload.get("site_url", ""),
        title=payload.get("title", ""),
        lead_type=payload.get("lead_type", ""),
        why_relevant=payload.get("why_relevant", ""),
        date=payload.get("date", ""),
        deadline=payload.get("deadline", ""),
        contacts={key: list(value) for key, value in (payload.get("contacts") or {}).items()},
        source_urls=list(payload.get("source_urls", [])),
        confidence=float(payload.get("confidence", 0.0) or 0.0),
        status=payload.get("status", "new"),
    )


def site_refresh_plan_from_dict(payload: dict[str, Any]) -> SiteRefreshPlan:
    return SiteRefreshPlan(
        site_url=payload.get("site_url", ""),
        cadence=payload.get("cadence", ""),
        next_due_at=payload.get("next_due_at", ""),
        reason=payload.get("reason", ""),
    )


def company_result_from_dict(payload: dict[str, Any]) -> CompanyResult:
    normalized_payload = normalize_company_result_payload(payload)
    return CompanyResult(
        row_index=int(normalized_payload.get("row_index", 0) or 0),
        inn=normalized_payload.get("inn", ""),
        company_name=normalized_payload.get("company_name", ""),
        input_site=normalized_payload.get("input_site", ""),
        input_phone=normalized_payload.get("input_phone", ""),
        input_comment=normalized_payload.get("input_comment", ""),
        started_at=normalized_payload.get("started_at", ""),
        finished_at=normalized_payload.get("finished_at", ""),
        status=normalized_payload.get("status", ""),
        output_contract_version=normalized_payload.get(
            "output_contract_version",
            COMPANY_OUTPUT_CONTRACT_VERSION,
        ),
        sources={
            key: source_result_from_dict(value)
            for key, value in (normalized_payload.get("sources") or {}).items()
        },
        merged_contacts={
            key: list(value) for key, value in (normalized_payload.get("merged_contacts") or {}).items()
        },
        trusted_contacts={
            key: list(value) for key, value in (normalized_payload.get("trusted_contacts") or {}).items()
        },
        domain_resolution=domain_resolution_from_dict(normalized_payload.get("domain_resolution") or {})
        if normalized_payload.get("domain_resolution")
        else None,
        candidate_sites=list(normalized_payload.get("candidate_sites", [])),
        site_probes=[site_probe_from_dict(item) for item in normalized_payload.get("site_probes", [])],
        route_strategies=[
            route_strategy_from_dict(item) for item in normalized_payload.get("route_strategies", [])
        ],
        content_records=[content_record_from_dict(item) for item in normalized_payload.get("content_records", [])],
        lead_cards=[lead_card_from_dict(item) for item in normalized_payload.get("lead_cards", [])],
        site_refresh_plans=[
            site_refresh_plan_from_dict(item) for item in normalized_payload.get("site_refresh_plans", [])
        ],
        dossier_ref=dict(normalized_payload.get("dossier_ref", {}))
        if isinstance(normalized_payload.get("dossier_ref"), Mapping)
        else {},
        validated_sites=[
            site_decision_from_dict(item) for item in normalized_payload.get("validated_sites", [])
        ],
        notes=list(normalized_payload.get("notes", [])),
        profile=company_profile_from_dict(normalized_payload.get("profile") or {}),
    )


def parse_count(value: str) -> int | None:
    if value.lower() == "all":
        return None
    return int(value)


def parse_source_names(value: str) -> set[str] | None:
    if not value or value.lower() == "all":
        return None
    return {item.strip() for item in value.split(",") if item.strip()}


def is_retryable_block_status(status: str) -> bool:
    return status in {
        REQUEST_STATUS_BLOCKED_NO_PROXY,
        "rate_limited",
        "bot_gate",
        "cooldown_active",
        "source_disabled_after_block",
    }


def make_blocked_source_result(source_name: str, reason: str, *, status: str = "source_disabled_after_block") -> SourceResult:
    result = SourceResult(source=source_name, status=status)
    result.notes.append(reason)
    mark_source_blocked(result, reason=reason)
    return result


def should_skip_on_resume(
    existing_payload: dict[str, Any] | None,
    active_source_names: list[str],
    *,
    retry_blocked_source: str = "",
) -> bool:
    if not existing_payload:
        return False
    if retry_blocked_source:
        return False
    sources_payload = existing_payload.get("sources") or {}
    if not sources_payload:
        return False
    for source_name in active_source_names:
        status = ((sources_payload.get(source_name) or {}).get("status") or "").strip()
        if not status:
            return False
    return True


def configure_logger(log_path: Path) -> logging.Logger:
    ensure_dir(log_path.parent)
    logger = logging.getLogger("company_research_parser")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger


def find_default_xlsx(cwd: Path) -> Path:
    matches = sorted(cwd.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError("ÐÐµ Ð½Ð°ÑˆÐµÐ» .xlsx Ñ„Ð°Ð¹Ð» Ð² Ñ‚ÐµÐºÑƒÑ‰ÐµÐ¹ Ð¿Ð°Ð¿ÐºÐµ")
    return matches[0]



