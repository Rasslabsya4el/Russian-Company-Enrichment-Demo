from __future__ import annotations

import argparse
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SOURCE_LABELS = {
    "spark": "СПАРК",
    "zachestnyibiznes": "ЗАЧЕСТНЫЙБИЗНЕС",
    "rusprofile": "Rusprofile",
    "list_org": "List-Org",
    "xlsx_input": "исходная таблица",
    "domain_resolution": "автовыбор домена",
    "email_domain": "домен из email",
}

HEADER_SPECS = [
    ("№", "Технический номер строки из полного прогона.", 8, "company"),
    ("ИНН", "Главный идентификатор компании.", 14, "company"),
    ("Компания", "Название компании из итогового прогона.", 34, "company"),
    ("ИНН корректный", "Да — ИНН выглядит валидным, Нет — проверьте вручную.", 16, "company"),
    ("Сайт (лучший вариант)", "Лучший сайт по мнению системы.", 34, "best"),
    ("Сайт подтвержден", "Да — сайт подтвержден обходом, Нет — нужна ручная проверка.", 14, "best"),
    ("Уверенность по сайту", "Высокая / Средняя / Низкая — подсказка, а не финальная истина.", 18, "best"),
    ("Почему выбран сайт", "Короткое объяснение выбора лучшего сайта.", 34, "best"),
    ("Все найденные сайты", "Все сайты и домены, найденные по компании.", 38, "all"),
    ("Телефон (лучший вариант)", "Лучший телефон по мнению системы.", 22, "best"),
    ("Все найденные телефоны", "Все телефоны, собранные по компании.", 28, "all"),
    ("Email (лучший вариант)", "Лучший email по мнению системы.", 28, "best"),
    ("Все найденные email", "Все email, собранные по компании.", 28, "all"),
    ("Адрес (лучший вариант)", "Лучший адрес для карточки компании.", 34, "best"),
    ("Все найденные адреса", "Все адреса, найденные по компании.", 42, "all"),
    ("Что найдено автоматически", "Краткая сводка по найденным данным.", 28, "signals"),
    ("Что проверить вручную", "Куда смотреть в первую очередь.", 24, "signals"),
    ("Источники данных", "Откуда система взяла данные по строке.", 24, "signals"),
    ("Проверка сайта", "Поле сотрудника для ручной проверки сайта.", 18, "review"),
    ("Проверка телефона", "Поле сотрудника для ручной проверки телефона.", 18, "review"),
    ("Проверка email", "Поле сотрудника для ручной проверки email.", 18, "review"),
    ("Итог проверки", "Общий статус ручной проверки строки.", 18, "review"),
    ("Комментарий сотрудника", "Свободное поле для замечаний и правок.", 40, "review"),
]

COLUMN_GUIDE_ROWS = [
    ("№", "Технический номер строки из полного прогона.", "Нужен только для связи с техотчетами и исходным full run.", "Любой заполненный номер.", "255", "Обычно ничего не делать."),
    ("ИНН", "Главный идентификатор компании.", "По ИНН проще всего сверять компанию с внешними источниками.", "ИНН совпадает с компанией.", "6325045051", "Если есть сомнения по компании, сначала сверить ИНН."),
    ("Компания", "Название компании из итогового прогона.", "Используйте вместе с ИНН для ручной проверки.", "Название соответствует сайту и контактам.", "ООО \"МАГИСТРАЛЬ\"", "Сверить, что сайт и контакты относятся именно к этой компании."),
    ("ИНН корректный", "Признак, что ИНН выглядит валидным.", "Если здесь Нет, строку нужно проверять особенно внимательно.", "Да — нормальный сигнал. Нет — риск ошибки.", "Да", "При значении Нет сначала перепроверьте компанию."),
    ("Сайт (лучший вариант)", "Самый вероятный сайт компании по мнению системы.", "С него начинайте проверку, но не считайте истиной без просмотра.", "Сайт открывается и совпадает с компанией. Риск — сайт чужой или устаревший.", "https://example.ru", "Открыть сайт и подтвердить или исправить."),
    ("Сайт подтвержден", "Автоматика смогла подтвердить сайт при обходе.", "Да — сильный сигнал. Нет — сайт найден, но еще не подтвержден.", "Да лучше, чем Нет.", "Да", "Если Нет, обязательно проверить сайт руками."),
    ("Уверенность по сайту", "Оценка надежности лучшего сайта.", "Высокая — сильный матч. Средняя — выглядит правдоподобно. Низкая — спорно.", "Высокая хорошо. Низкая — риск ошибки.", "Средняя", "Сначала смотреть строки с Низкой и Средней уверенностью."),
    ("Почему выбран сайт", "Короткое объяснение, почему система выбрала именно этот сайт.", "Показывает логику выбора: агрегаторы, обход сайта, домен из email.", "Чем понятнее и конкретнее причина, тем лучше.", "Статус: подтвержден", "Использовать как подсказку, а не как окончательное доказательство."),
    ("Все найденные сайты", "Все сайты и домены, которые встретились по компании.", "Нужно, если лучший сайт спорный или отсутствует.", "Хорошо, когда список короткий и согласованный. Плохо, когда там много мусора.", "• https://site1.ru\n• https://portal.site1.ru", "Если лучший сайт сомнителен, проверить альтернативы отсюда."),
    ("Телефон (лучший вариант)", "Лучший телефон по мнению системы.", "Это основной кандидат для прозвона.", "Корпоративный номер лучше случайного мобильного.", "+7 495 123-45-67", "Проверить, что номер реально относится к компании."),
    ("Все найденные телефоны", "Все телефоны, собранные по компании.", "Полезно, если лучший номер спорный или не отвечает.", "Больше одного номера — нормально. Много разнородных номеров — повод проверить.", "+7 495...\n+7 916...", "При необходимости выбрать лучший номер вручную."),
    ("Email (лучший вариант)", "Лучший email по мнению системы.", "Основной адрес для работы и сверки домена.", "Корпоративный домен лучше публичной почты.", "info@example.ru", "Проверить, что email рабочий и связан с компанией."),
    ("Все найденные email", "Все email, собранные по компании.", "Нужно, если лучший email спорный или нужен запасной.", "Корпоративные email лучше публичных.", "info@example.ru\nsales@example.ru", "При необходимости выбрать лучший email вручную."),
    ("Адрес (лучший вариант)", "Самый полезный адрес для карточки компании.", "Обычно это основной адрес компании или площадки.", "Адрес понятный и полный — хороший сигнал.", "Москва, ул. Пример, д. 1", "Проверить, что адрес относится к этой компании."),
    ("Все найденные адреса", "Все адреса, найденные по компании.", "Нужно для сверки, если адресов несколько или есть филиалы.", "Один-два понятных адреса — хорошо. Длинная мешанина — повод проверить.", "1. Москва, ...\n\n2. Химки, ...", "Сверить, какой адрес основной, а какие вторичные."),
    ("Что найдено автоматически", "Сводка, сколько система нашла лучших и всех контактов.", "Это быстрый обзор по заполненности строки.", "Чем больше полезных находок, тем проще ручная работа.", "Лучшее: ...\nВсего найдено: ...", "Использовать как быстрый ориентир по качеству строки."),
    ("Что проверить вручную", "Куда смотреть в первую очередь.", "Это список главных ручных действий по строке.", "Короткий и конкретный список — нормальный сигнал.", "• проверить сайт\n• проверить email", "Начать ручную проверку именно с этих пунктов."),
    ("Источники данных", "Откуда система взяла данные по строке.", "Показывает, какие источники реально участвовали в сборе.", "Когда источников несколько и они согласованы — это плюс.", "Rusprofile\nList-Org", "Использовать для понимания происхождения данных."),
    ("Проверка сайта", "Ручной вердикт сотрудника по сайту.", "Заполняется после ручной проверки.", "подтвержден — хорошо; неверный — сайт надо заменить.", "подтвержден", "Выбрать подходящий статус из списка."),
    ("Проверка телефона", "Ручной вердикт сотрудника по телефону.", "Заполняется после ручной проверки телефона.", "верный — хорошо; неверный — номер не использовать.", "верный", "Выбрать подходящий статус из списка."),
    ("Проверка email", "Ручной вердикт сотрудника по email.", "Заполняется после ручной проверки email.", "верный — хорошо; неверный — адрес не использовать.", "верный", "Выбрать подходящий статус из списка."),
    ("Итог проверки", "Общий статус строки после ручной проверки.", "Показывает, на каком этапе находится строка.", "проверено — строка готова; в работе — еще не закончена.", "в работе", "После проверки обновить итоговый статус."),
    ("Комментарий сотрудника", "Свободное поле для замечаний.", "Сюда записываются ручные находки, ошибки и правки.", "Чем конкретнее комментарий, тем лучше.", "Сайт устарел, новый домен не найден.", "Оставить понятную заметку для команды."),
]

SECTION_FILLS = {
    "company": PatternFill("solid", fgColor="D9EAF7"),
    "best": PatternFill("solid", fgColor="DFF2E1"),
    "all": PatternFill("solid", fgColor="FFF1CC"),
    "signals": PatternFill("solid", fgColor="FCE5CD"),
    "review": PatternFill("solid", fgColor="EADCF8"),
}

LINK_LIKE_FONT = Font(color="0563C1", underline="single")

MAX_SITES = 6
MAX_PHONES = 8
MAX_EMAILS = 8
MAX_ADDRESSES = 4
MAX_REASON_LENGTH = 280


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a human-friendly validation workbook from final_results.xlsx")
    parser.add_argument("--input", required=True, help="Path to full run final_results.xlsx")
    parser.add_argument("--output", help="Path to output validation xlsx")
    parser.add_argument(
        "--enable-hyperlinks",
        action="store_true",
        help="Enable real Excel hyperlinks for best site and best email. Disabled by default because some Excel builds repair external links.",
    )
    return parser.parse_args()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def yes_no(value: object) -> str:
    text = normalize_text(value).lower()
    if text in {"yes", "true", "1", "да"}:
        return "Да"
    if text in {"no", "false", "0", "нет"}:
        return "Нет"
    return ""


def unique_list(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        normalized = item.strip()
        key = normalized.lower()
        if not normalized or key in seen:
            continue
        seen.add(key)
        out.append(normalized)
    return out


def split_pipe_values(value: str) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split("|") if part.strip()]


def join_lines(items: list[str], *, blank_line: bool = False) -> str:
    items = unique_list(items)
    if not items:
        return ""
    separator = "\n\n" if blank_line else "\n"
    return separator.join(items)


def format_list_block(
    items: list[str],
    *,
    bullet: bool = False,
    blank_line: bool = False,
    numbered: bool = False,
    limit: int | None = None,
) -> str:
    cleaned = unique_list(items)
    if not cleaned:
        return ""

    hidden_count = 0
    if limit is not None and len(cleaned) > limit:
        hidden_count = len(cleaned) - limit
        cleaned = cleaned[:limit]

    rendered: list[str] = []
    for index, item in enumerate(cleaned, start=1):
        if numbered:
            rendered.append(f"{index}. {item}")
        elif bullet:
            rendered.append(f"• {item}")
        else:
            rendered.append(item)

    if hidden_count > 0:
        rendered.append(f"+ еще {hidden_count}")

    return join_lines(rendered, blank_line=blank_line)


def format_pipe_block(
    value: str,
    *,
    bullet: bool = False,
    blank_line: bool = False,
    numbered: bool = False,
    limit: int | None = None,
) -> str:
    return format_list_block(
        split_pipe_values(value),
        bullet=bullet,
        blank_line=blank_line,
        numbered=numbered,
        limit=limit,
    )


def trim_text(value: str, limit: int) -> str:
    text = normalize_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def is_safe_excel_hyperlink_target(value: str) -> bool:
    if not value:
        return False
    text = value.strip()
    if not text:
        return False
    if any(ch in text for ch in ('"', "<", ">", "\n", "\r")):
        return False
    if any(ord(ch) > 127 for ch in text):
        return False
    if " " in text:
        return False
    if text.startswith(("http://", "https://")):
        return True
    if text.startswith("mailto:") and "?" not in text and "&" not in text:
        return True
    return False


def extract_source_names(*raw_values: str) -> list[str]:
    parts: list[str] = []
    for raw in raw_values:
        for token in re.split(r"[,;]+", raw or ""):
            token = token.strip()
            if not token:
                continue
            base = token.split(":", 1)[0].strip()
            if base in SOURCE_LABELS:
                parts.append(SOURCE_LABELS[base])
            elif token in SOURCE_LABELS:
                parts.append(SOURCE_LABELS[token])
            else:
                parts.append(token)
    return unique_list(parts)


def site_confidence(row: dict[str, str]) -> str:
    confirmed = int(row.get("confirmed_site_count") or 0)
    status = row.get("best_site_status", "")
    site_class = row.get("site_classes", "")
    worth = row.get("worth_crawling", "")
    if not row.get("best_site"):
        return "Нет сайта"
    if confirmed > 0 or status == "trusted_validated":
        return "Высокая"
    if site_class in {"E", "F"} and confirmed == 0:
        return "Низкая"
    if status == "trusted":
        return "Средняя"
    if status in {"domain_resolution", "merged_only"}:
        return "Низкая"
    if worth == "true":
        return "Средняя"
    return "Низкая"


def site_choice_reason(row: dict[str, str]) -> str:
    if not row.get("best_site"):
        return "Статус: сайт не выбран"

    confirmed = int(row.get("confirmed_site_count") or 0)
    status = row.get("best_site_status", "")
    sources = extract_source_names(row.get("best_site_sources", ""))
    confidence = site_confidence(row)

    if confirmed > 0 or status == "trusted_validated":
        status_line = "Статус: сайт подтвержден"
        risk_line = "Риск: низкий"
    elif status == "trusted":
        status_line = "Статус: сайт выбран по нескольким сильным сигналам"
        risk_line = "Проверить: подтверждение сайта вручную"
    elif status == "domain_resolution":
        status_line = "Статус: сайт выбран по совпадениям в источниках"
        risk_line = "Проверить: сайт не подтвержден обходом"
    elif status == "merged_only":
        status_line = "Статус: сайт взят из собранных данных"
        risk_line = "Проверить: слабое подтверждение"
    else:
        status_line = "Статус: сайт найден автоматически"
        risk_line = "Проверить: требуется ручная валидация"

    source_line = "Основание: " + (", ".join(sources) if sources else "автовыбор системы")
    confidence_line = f"Уверенность: {confidence}"
    return trim_text("\n".join([status_line, source_line, confidence_line, risk_line]), MAX_REASON_LENGTH)


def what_found_summary(row: dict[str, str]) -> str:
    confirmed_parts: list[str] = []
    if int(row.get("confirmed_site_count") or 0) > 0:
        confirmed_parts.append("сайт")
    if int(row.get("trusted_phone_count") or 0) > 0:
        confirmed_parts.append("телефон")
    if int(row.get("trusted_email_count") or 0) > 0:
        confirmed_parts.append("email")

    lines = [
        f"Лучшее: сайт {row.get('trusted_site_count') or 0}, телефон {row.get('trusted_phone_count') or 0}, email {row.get('trusted_email_count') or 0}",
        f"Всего найдено: сайты {row.get('raw_site_count') or 0}, телефоны {row.get('raw_phone_count') or 0}, email {row.get('raw_email_count') or 0}",
    ]
    if confirmed_parts:
        lines.append("Подтверждено: " + ", ".join(confirmed_parts))
    return "\n".join(lines)


def manual_check_hint(row: dict[str, str]) -> str:
    hints: list[str] = []
    if yes_no(row.get("inn_valid")) == "Нет":
        hints.append("проверить ИНН")
    if row.get("best_site"):
        if int(row.get("confirmed_site_count") or 0) == 0:
            hints.append("проверить сайт")
    elif int(row.get("raw_site_count") or 0) > 0:
        hints.append("выбрать сайт")
    if row.get("best_phone"):
        if int(row.get("trusted_phone_count") or 0) == 0 and int(row.get("raw_phone_count") or 0) > 0:
            hints.append("проверить телефон")
    elif int(row.get("raw_phone_count") or 0) > 0:
        hints.append("выбрать телефон")
    if row.get("best_email"):
        if int(row.get("trusted_email_count") or 0) == 0 and int(row.get("raw_email_count") or 0) > 0:
            hints.append("проверить email")
    elif int(row.get("raw_email_count") or 0) > 0:
        hints.append("выбрать email")
    if not hints:
        return "минимум ручной проверки"
    return format_list_block(unique_list(hints), bullet=True)


def load_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [normalize_text(ws.cell(row=1, column=i).value) for i in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: normalize_text(values[i]) for i in range(len(headers))})
    return rows


def add_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Как читать")

    ws["A1"] = "Как читать файл"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Что это за файл"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = (
        "Это таблица для ручной проверки результатов парсинга. "
        "Система уже выбрала лучший сайт, телефон, email и адрес, "
        "но сотрудник должен подтвердить или исправить эти значения."
    )

    ws["A6"] = "Как работать со строкой"
    ws["A6"].font = Font(bold=True)
    steps = [
        "1. Посмотреть колонку «Сайт (лучший вариант)».",
        "2. Сравнить его с колонкой «Все найденные сайты».",
        "3. Проверить лучший телефон, email и адрес.",
        "4. Посмотреть колонки «Почему выбран сайт» и «Что проверить вручную».",
        "5. Заполнить «Проверка сайта / телефона / email» и поставить «Итог проверки».",
    ]
    for row_index, step in enumerate(steps, start=7):
        ws.cell(row=row_index, column=1, value=step)

    ws["A13"] = "Что значат цветовые группы"
    ws["A13"].font = Font(bold=True)
    color_rows = [
        ("Синие колонки", "данные о компании"),
        ("Зеленые колонки", "лучший выбранный вариант"),
        ("Желтые колонки", "все найденные варианты"),
        ("Оранжевые колонки", "подсказки системы"),
        ("Фиолетовые колонки", "поля для сотрудника"),
    ]
    row_index = 14
    for left, right in color_rows:
        ws.cell(row=row_index, column=1, value=left).font = Font(bold=True)
        ws.cell(row=row_index, column=2, value=right)
        row_index += 1

    ws["A21"] = "Смысл статусов"
    ws["A21"].font = Font(bold=True)
    status_rows = [
        ("Сайт подтвержден", "Да — сайт подтвержден автоматикой; Нет — нужен ручной контроль."),
        ("Уверенность по сайту", "Высокая / Средняя / Низкая / Нет сайта."),
        ("Итог проверки", "не начато / в работе / проверено / отклонено."),
    ]
    row_index = 22
    for left, right in status_rows:
        ws.cell(row=row_index, column=1, value=left).font = Font(bold=True)
        ws.cell(row=row_index, column=2, value=right)
        row_index += 1

    start_row = 27
    headers = ["Колонка", "Что это", "Как использовать", "Хороший сигнал / риск", "Пример", "Что делать сотруднику"]
    for column_index, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=column_index, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    row_index = start_row + 1
    for title, meaning, how_to_use, good_bad, example, action in COLUMN_GUIDE_ROWS:
        values = [title, meaning, how_to_use, good_bad, example, action]
        for column_index, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=column_index, value=value)
        row_index += 1

    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_letter, width in {"A": 28, "B": 34, "C": 38, "D": 34, "E": 28, "F": 34}.items():
        ws.column_dimensions[column_letter].width = width

    ws.freeze_panes = f"A{start_row + 1}"


def add_header_comments_and_styles(ws) -> None:
    for idx, (title, description, width, section) in enumerate(HEADER_SPECS, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = SECTION_FILLS[section]
        cell.comment = Comment(description, "Codex")
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADER_SPECS)).column_letter}1"


def add_validations(ws, max_row: int) -> None:
    validations = {
        "S": '"подтвержден,неверный,нужно проверить,нет сайта"',
        "T": '"верный,неверный,нужно проверить,нет данных"',
        "U": '"верный,неверный,нужно проверить,нет данных"',
        "V": '"не начато,в работе,проверено,отклонено"',
    }
    for col, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{max_row}")


def add_conditional_formatting(ws, max_row: int) -> None:
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="F4CCCC")

    ws.conditional_formatting.add(f"F2:F{max_row}", CellIsRule(operator="equal", formula=['"Да"'], fill=green))
    ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator="equal", formula=['"Высокая"'], fill=green))
    ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator="equal", formula=['"Средняя"'], fill=yellow))
    ws.conditional_formatting.add(f"G2:G{max_row}", CellIsRule(operator="equal", formula=['"Низкая"'], fill=red))
    ws.conditional_formatting.add(f"V2:V{max_row}", CellIsRule(operator="equal", formula=['"проверено"'], fill=green))
    ws.conditional_formatting.add(f"V2:V{max_row}", CellIsRule(operator="equal", formula=['"в работе"'], fill=yellow))
    ws.conditional_formatting.add(f"V2:V{max_row}", CellIsRule(operator="equal", formula=['"отклонено"'], fill=red))


def build_validation_rows(rows: list[dict[str, str]]) -> list[list[str]]:
    output: list[list[str]] = []
    for row in rows:
        data_sources = extract_source_names(
            row.get("best_site_sources", ""),
            row.get("best_phone_sources", ""),
            row.get("best_email_sources", ""),
            row.get("best_address_sources", ""),
        )
        output.append(
            [
                row.get("row_index", ""),
                row.get("inn", ""),
                row.get("company_name", ""),
                yes_no(row.get("inn_valid")),
                format_pipe_block(row.get("best_site", "")),
                "Да" if int(row.get("confirmed_site_count") or 0) > 0 else "Нет",
                site_confidence(row),
                site_choice_reason(row),
                format_pipe_block(row.get("all_raw_sites", ""), bullet=True, limit=MAX_SITES),
                format_pipe_block(row.get("best_phone", "")),
                format_pipe_block(row.get("all_raw_phones", ""), limit=MAX_PHONES),
                format_pipe_block(row.get("best_email", "")),
                format_pipe_block(row.get("all_raw_emails", ""), limit=MAX_EMAILS),
                format_pipe_block(row.get("best_address", ""), blank_line=True),
                format_pipe_block(row.get("all_raw_addresses", ""), blank_line=True, numbered=True, limit=MAX_ADDRESSES),
                what_found_summary(row),
                manual_check_hint(row),
                join_lines(data_sources),
                "",
                "",
                "",
                "не начато",
                "",
            ]
        )
    return output


def build_workbook(rows: list[dict[str, str]], output_path: Path, *, enable_hyperlinks: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Проверка компаний"
    add_header_comments_and_styles(ws)

    for out_row in build_validation_rows(rows):
        ws.append(out_row)

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        site_cell = row[4]
        email_cell = row[11]

        if isinstance(site_cell.value, str) and site_cell.value.startswith(("http://", "https://")):
            if enable_hyperlinks and is_safe_excel_hyperlink_target(site_cell.value):
                site_cell.hyperlink = site_cell.value
                site_cell.style = "Hyperlink"
            else:
                site_cell.font = LINK_LIKE_FONT

        if isinstance(email_cell.value, str) and "@" in email_cell.value and "\n" not in email_cell.value:
            mailto = f"mailto:{email_cell.value}"
            if enable_hyperlinks and is_safe_excel_hyperlink_target(mailto):
                email_cell.hyperlink = mailto
                email_cell.style = "Hyperlink"
            else:
                email_cell.font = LINK_LIKE_FONT

    add_validations(ws, max_row)
    add_conditional_formatting(ws, max_row)
    add_legend_sheet(wb)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).resolve()
    if not input_path.exists():
        raise SystemExit(f"Input not found: {input_path}")
    output_path = Path(args.output).resolve() if args.output else input_path.with_name("validation_results.xlsx")
    rows = load_rows(input_path)
    build_workbook(rows, output_path, enable_hyperlinks=args.enable_hyperlinks)
    print(str(output_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
