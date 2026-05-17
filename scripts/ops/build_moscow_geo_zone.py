from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.discovery.geo_zone import classify


CSV_FIELDNAMES = [
    "source_id",
    "region",
    "region_normalized",
    "municipality",
    "municipality_normalized",
    "settlement",
    "settlement_normalized",
    "settlement_type",
    "full_name",
    "full_name_normalized",
    "population",
    "children",
    "latitude_dms",
    "longitude_dms",
    "latitude_dd",
    "longitude_dd",
    "inside_outer_polygon",
    "inside_inner_polygon",
    "geo_bucket",
    "geo_weight",
    "distance_to_moscow_km",
    "oktmo",
    "dadata",
    "rosstat",
]

BUCKET_ORDER = ("core", "outer_band", "outside")


@dataclass(frozen=True)
class ExportPaths:
    csv_path: Path
    xlsx_path: Path
    report_path: Path


def _clean_text(value: str | None) -> str:
    if value is None:
        return ""
    cleaned = value.replace("\xa0", " ").strip()
    return " ".join(cleaned.split())


def _normalize_text(value: str | None) -> str:
    return _clean_text(value).replace("\u0451", "\u0435").replace("\u0401", "\u0415")


def _parse_float(value: str | None, *, field_name: str, row_number: int) -> float:
    cleaned = _clean_text(value).replace(",", ".")
    if not cleaned:
        raise ValueError(f"row {row_number}: missing {field_name}")
    try:
        return float(cleaned)
    except ValueError as exc:
        raise ValueError(f"row {row_number}: invalid {field_name}={cleaned!r}") from exc


def _parse_int_or_blank(value: str | None) -> int | str:
    cleaned = _clean_text(value)
    if not cleaned:
        return ""
    try:
        return int(float(cleaned.replace(",", ".")))
    except ValueError:
        return cleaned


def _full_name(settlement_type: str, settlement: str) -> str:
    if settlement_type and settlement:
        return f"{settlement_type} {settlement}"
    return settlement or settlement_type


def _rounded(value: float, digits: int) -> float:
    return round(value, digits)


def _build_output_row(raw_row: dict[str, str], *, row_number: int) -> dict[str, object]:
    region = _clean_text(raw_row.get("region"))
    municipality = _clean_text(raw_row.get("municipality"))
    settlement = _clean_text(raw_row.get("settlement"))
    settlement_type = _clean_text(raw_row.get("type"))
    latitude_dd = _parse_float(raw_row.get("latitude_dd"), field_name="latitude_dd", row_number=row_number)
    longitude_dd = _parse_float(raw_row.get("longitude_dd"), field_name="longitude_dd", row_number=row_number)
    classification = classify(latitude_dd, longitude_dd)

    full_name = _full_name(settlement_type, settlement)
    return {
        "source_id": _clean_text(raw_row.get("id")),
        "region": region,
        "region_normalized": _normalize_text(region),
        "municipality": municipality,
        "municipality_normalized": _normalize_text(municipality),
        "settlement": settlement,
        "settlement_normalized": _normalize_text(settlement),
        "settlement_type": settlement_type,
        "full_name": full_name,
        "full_name_normalized": _normalize_text(full_name),
        "population": _parse_int_or_blank(raw_row.get("population")),
        "children": _parse_int_or_blank(raw_row.get("children")),
        "latitude_dms": _clean_text(raw_row.get("latitude_dms")),
        "longitude_dms": _clean_text(raw_row.get("longitude_dms")),
        "latitude_dd": _rounded(latitude_dd, 6),
        "longitude_dd": _rounded(longitude_dd, 6),
        "inside_outer_polygon": classification.inside_outer_polygon,
        "inside_inner_polygon": classification.inside_inner_polygon,
        "geo_bucket": classification.geo_bucket,
        "geo_weight": classification.geo_weight,
        "distance_to_moscow_km": _rounded(classification.distance_to_moscow_km, 2),
        "oktmo": _clean_text(raw_row.get("oktmo")),
        "dadata": _clean_text(raw_row.get("dadata")),
        "rosstat": _clean_text(raw_row.get("rosstat")),
    }


def _default_export_paths(output_dir: Path) -> ExportPaths:
    return ExportPaths(
        csv_path=output_dir / "moscow_geo_zone_settlements.csv",
        xlsx_path=output_dir / "moscow_geo_zone_settlements.xlsx",
        report_path=output_dir / "moscow_geo_zone_report.txt",
    )


def _write_report(
    *,
    report_path: Path,
    input_path: Path,
    export_paths: ExportPaths,
    total_rows: int,
    bucket_counts: Counter[str],
    inside_region_counts: Counter[str],
    inside_region_labels: dict[str, str],
    top_regions_limit: int,
) -> None:
    inside_total = bucket_counts.get("core", 0) + bucket_counts.get("outer_band", 0)
    lines = [
        "Moscow geo-zone operator export",
        "",
        f"input_csv: {input_path}",
        f"rows_total: {total_rows}",
        f"rows_inside_zone: {inside_total}",
        f"rows_core: {bucket_counts.get('core', 0)}",
        f"rows_outer_band: {bucket_counts.get('outer_band', 0)}",
        f"rows_outside: {bucket_counts.get('outside', 0)}",
        "",
        "generated_files:",
        f"- csv: {export_paths.csv_path}",
        f"- xlsx: {export_paths.xlsx_path}",
        f"- report: {export_paths.report_path}",
        "",
        "bucket_counts:",
    ]
    for bucket in BUCKET_ORDER:
        lines.append(f"- {bucket}: {bucket_counts.get(bucket, 0)}")
    lines.extend(["", "top_regions_inside_zone:"])
    for rank, (region_key, count) in enumerate(inside_region_counts.most_common(top_regions_limit), start=1):
        region_label = inside_region_labels.get(region_key, region_key)
        lines.append(f"- {rank}. {region_label}: {count}")
    if not inside_region_counts:
        lines.append("- none")
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a one-off Moscow geo-zone settlement export for manual operator review."
    )
    parser.add_argument("--input", required=True, help="Path to the source settlement CSV file.")
    parser.add_argument("--output-dir", required=True, help="Directory for generated CSV/XLSX/report files.")
    parser.add_argument(
        "--top-regions-limit",
        type=int,
        default=15,
        help="How many top regions inside the zone to include in the report outputs.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")

    args = parse_args(argv)
    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    export_paths = _default_export_paths(output_dir)

    if not input_path.is_file():
        raise SystemExit(f"Input CSV does not exist: {input_path}")
    if args.top_regions_limit < 1:
        raise SystemExit("--top-regions-limit must be >= 1")

    output_dir.mkdir(parents=True, exist_ok=True)

    bucket_counts: Counter[str] = Counter()
    inside_region_counts: Counter[str] = Counter()
    inside_region_labels: dict[str, str] = {}
    total_rows = 0

    workbook = Workbook(write_only=True)
    settlements_sheet = workbook.create_sheet("settlements")
    settlements_sheet.append(CSV_FIELDNAMES)

    with (
        input_path.open("r", encoding="utf-8-sig", newline="") as source_file,
        export_paths.csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file,
    ):
        reader = csv.DictReader(source_file)
        writer = csv.DictWriter(csv_file, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()

        for row_number, raw_row in enumerate(reader, start=2):
            total_rows += 1
            output_row = _build_output_row(raw_row, row_number=row_number)
            writer.writerow(output_row)
            settlements_sheet.append([output_row[fieldname] for fieldname in CSV_FIELDNAMES])

            geo_bucket = str(output_row["geo_bucket"])
            bucket_counts[geo_bucket] += 1
            if bool(output_row["inside_outer_polygon"]):
                region_label = str(output_row["region"] or "<blank>")
                region_key = str(output_row["region_normalized"] or region_label)
                inside_region_counts[region_key] += 1
                inside_region_labels.setdefault(region_key, region_label)

    bucket_sheet = workbook.create_sheet("bucket_counts")
    bucket_sheet.append(["geo_bucket", "rows"])
    for bucket in BUCKET_ORDER:
        bucket_sheet.append([bucket, bucket_counts.get(bucket, 0)])

    top_regions_sheet = workbook.create_sheet("top_regions_inside_zone")
    top_regions_sheet.append(["rank", "region", "rows_inside_zone"])
    for rank, (region_key, count) in enumerate(inside_region_counts.most_common(args.top_regions_limit), start=1):
        top_regions_sheet.append([rank, inside_region_labels.get(region_key, region_key), count])
    if not inside_region_counts:
        top_regions_sheet.append([1, "none", 0])

    workbook.save(export_paths.xlsx_path)

    _write_report(
        report_path=export_paths.report_path,
        input_path=input_path,
        export_paths=export_paths,
        total_rows=total_rows,
        bucket_counts=bucket_counts,
        inside_region_counts=inside_region_counts,
        inside_region_labels=inside_region_labels,
        top_regions_limit=args.top_regions_limit,
    )

    print(f"CSV {export_paths.csv_path}")
    print(f"XLSX {export_paths.xlsx_path}")
    print(f"REPORT {export_paths.report_path}")
    print(f"ROWS_TOTAL {total_rows}")
    for bucket in BUCKET_ORDER:
        print(f"BUCKET_{bucket.upper()} {bucket_counts.get(bucket, 0)}")
    for rank, (region_key, count) in enumerate(inside_region_counts.most_common(args.top_regions_limit), start=1):
        print(f"TOP_REGION_{rank} {inside_region_labels.get(region_key, region_key)} | {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
